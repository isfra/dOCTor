import os
import numpy as np
import GuiClean as GC
from openpyxl import Workbook, load_workbook
import shutil
from scipy.stats import linregress
import matplotlib.pyplot as plt

wb=load_workbook(os.path.dirname(os.path.realpath('__file__')) + '\\OCTvsNIRS 20 04 20 Final - per FP.xlsx')#"/globalLBI.xlsx")

for sheet in wb:
    print(sheet.title)

data=wb['fct']
keys={}
gui=GC.Main()
i=1


LBI=[]
for column in "BC":
    for row in range(2,35):
        cell_name = "{}{}".format(column, row)
        print(data[cell_name].value)
        if data[cell_name].value==None:
            miss=row
        else:
            try :
                LBI.append(float(str(data[cell_name].value)))
            except:
                print ('error', row,column)
                pass

#del LBI[miss-2]
Y=LBI[:len(LBI)//2]
X=LBI[len(LBI)//2:]
#del Y[miss]

plt.scatter(X, Y)
x = np.linspace(0, 1000, 1000)
R=linregress(X, Y)
print(R[0])
m=R[0]
q=R[1]
ax = plt.axes()
ax.plot(x, m*x+q)
plt.show()
print ('valori',X)

print(Y)

print(linregress(X, Y))
'''

for cell in  data[1]:
    keys[cell.value]=i
    print(i)
    i+=1
    
print(keys)

database='D:\Franz\oct_software\data\OCTvsNIRSmiss'    #'/home/eugenio/workspace/documets di pasqualino/Documents/OCTdatabase/pazienti mace'
#database='/media/fisici/Hitachi/Pazienti clima fra'

lp = 2
thr = 400

chem_path=database +'\\Chemograms'

if not os.path.exists(chem_path):
    os.mkdir(chem_path)
    

patients = next(os.walk(database))[1]

for patient in patients :

    for pb in next(os.walk(os.path.join(database,patient)))[1]:
        data_path=os.path.join(database,patient,pb)

        print(data_path)
        if os.path.exists(data_path+'/pullback.npy') and os.path.exists(data_path+'/Prediction.npy') :
        
            data.cell(lp, keys['Patient surname']).value = patient
            data.cell(lp, keys['Pullback']).value = pb

            print(data_path)
            gui.import_pullback_callback(data_path)


            l=0
            for interv in gui.interv_of_pertinence:
               l+=interv[1]-interv[0]
            data.cell(lp, keys['Length']).value = l*0.2

            #lbi_depth=20

            gui.current_section_canvas.thr.set(0.3)
            gui.current_section_canvas.thr_ca.set(0.4)
            gui.topological_build()

            gui.mark_deep_plaque_n(depth=120)

            #gui.load_labels_classification()
            gui.prediction_canvas.update(0)

            gui.update_idletasks()
            gui.update()
            gui.save_chemiogram(chem_path, 'Low ' + patient +' '+pb)

            gui.analyze_LBI(superf=0, depth=20)

            data.cell(lp,keys['Max LBI']).value = np.max(gui.lbi)
            data.cell(lp,keys['Max LBI Fr']).value = np.argmax(gui.lbi) +1
            data.cell(lp,keys['Over LBI']).value = sum(gui.lbi>thr)
            if l-20 > 0:
                data.cell(lp,keys['%Over LBI']).value = sum(gui.lbi>thr)/(l-20)
            else:
                data.cell(lp,keys['%Over LBI']).value = None

            #data.cell(lp,keys['Mean LBI']).value = np.mean(gui.lbi[lp0:lp1-lbi_depth])

            gui.analyze_LBI(superf=1)
            data.cell(lp,keys['Max LBIs']).value = np.max(gui.lbi)
            data.cell(lp,keys['Max LBIs Fr']).value = np.argmax(gui.lbi) +1
            data.cell(lp,keys['Over LBIs']).value = sum(gui.lbi>thr)
            if l-20 > 0:
                data.cell(lp,keys['%Over LBIs']).value = sum(gui.lbi>thr)/(l-20)
            else:
                data.cell(lp,keys['%Over LBIs']).value = None

            #data.cell(lp,keys['Mean LBIs']).value = np.mean(gui.lbi[lp0:lp1-lbi_depth])
            #data.cell(lp,keys['LBIs all']).value = gui.LBI_superf(lp0,lp1)

            gui.current_section_canvas.thr.set(0.65)
            gui.current_section_canvas.thr_ca.set(0.4)
            gui.topological_build()

            gui.mark_deep_plaque_n(depth=120)
            #gui.load_labels_classification()
            gui.prediction_canvas.update(0)

            gui.update_idletasks()
            gui.update()
            gui.save_chemiogram(chem_path, 'High ' + patient +' '+pb)

            gui.analyze_LBI(superf=0, depth=20)

            data.cell(lp,keys['Max LBI thr']).value = np.max(gui.lbi)
            data.cell(lp,keys['Max LBI thr Fr']).value = np.argmax(gui.lbi) +1
            data.cell(lp,keys['Over LBI thr']).value = sum(gui.lbi>thr)
            if l-20 > 0:
                data.cell(lp,keys['%Over LBI thr']).value = sum(gui.lbi>thr)/(l-20)
            else:
                data.cell(lp,keys['%Over LBI thr']).value = None

            gui.analyze_LBI(superf=1)
            data.cell(lp,keys['Max LBIs thr']).value = np.max(gui.lbi)
            data.cell(lp,keys['Max LBIs thr Fr']).value = np.argmax(gui.lbi) +1
            data.cell(lp,keys['Over LBIs thr']).value = sum(gui.lbi>thr)
            if l-20 > 0:
                data.cell(lp,keys['%Over LBIs thr']).value = sum(gui.lbi>thr)/(l-20)
            else:
                data.cell(lp,keys['%Over LBIs thr']).value = None

            lp+=1

wb.save(os.path.dirname(os.path.realpath('__file__')) + '\\OCTvsNIRS.xlsx')#"/globalLBI.xlsx")
'''