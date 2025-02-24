
from PIL import Image, ImageTk
#from tkinter import *
from tkinter import Grid, Canvas, Text, Tk, IntVar, LEFT, RIGHT, NW, W, font, Checkbutton, YES, Scrollbar, VERTICAL, Y , Menu, Radiobutton, DoubleVar,  Entry, BOTTOM, TOP, END, Toplevel,Message, Button, StringVar, Label, BooleanVar
from tkinter.filedialog import askopenfilename, askdirectory

from scipy import interpolate 
from scipy import integrate
from scipy import signal

import cv2
import skimage.io
import collections
import csv
#import DataProcessing as DP
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from cmath import sqrt
import time
from datetime import date, datetime
import sched
import os
import io
from numpy import memmap, mean
import math

import pickle
#import xlwt 
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.colors import RED
from openpyxl.styles.fills import PatternFill, FILL_SOLID
from openpyxl.styles import Font

import glob
import shutil
from keras.models import Model, load_model
from skimage.transform  import resize
from sklearn.metrics import classification_report
#from Crypto.Util.number import size
#from keras.backend.cntk_backend import variable
import tensorflow
#from gui import lipid_color_index


RED = [0.6, 0.1, 0]
YELLOW = [1.0, 1.0, 0.0]
WHITE = [0.9, 0.9, 0.9]
BLACK = [0.1, 0.1, 0.1]
ORANGE = [1. , 0.756 ,0.027]   #[0.9216,0.4157,0.0078]
number_of_colors = 4096
'''
non_lipid_color_index = number_of_colors - 4
lipid_color_index = number_of_colors - 3
calcium_color_index = number_of_colors - 2
unknown_color_index = number_of_colors - 1

calcium_color_index = number_of_colors - 4
lipid_color_index = number_of_colors - 3
non_lipid_color_index = number_of_colors - 2
unknown_color_index = number_of_colors - 1
'''

unknown_color_index = number_of_colors - 5
non_lipid_color_index = number_of_colors - 4
deep_plaque_color_index=number_of_colors-3
lipid_color_index = number_of_colors - 2
calcium_color_index = number_of_colors - 1
#number_of_colors+=1

STRIDE = 16
DEG = 64
SPICCHIO_DEPHT = 400

play=False




class _OrderedDictMaker(object):
    def __getitem__(self, keys):
        if not isinstance(keys, tuple):
            keys = (keys,)
        assert all(isinstance(key, slice) for key in keys)

        return collections.OrderedDict([(k.start, k.stop) for k in keys])


def WriteOrdDictToCSV(csv_file, Info):
    keys, values = [], []

    for key, value in Info.items():
        keys.append(key)
        values.append(value)

    with open(csv_file, "a") as outfile:
        csvwriter = csv.writer(outfile)
        if os.stat(csv_file).st_size == 0:
            csvwriter.writerow(keys)
        csvwriter.writerow(values)
    return

ordereddict = _OrderedDictMaker()
Info = ordereddict[

        'accuracy'       : 0       ,
        'TP'             : 0       ,
        'FP'             : 0       ,
        'TN'             : 0       ,
        'FN'             : 0       ,
        'FPR'            : 0       ,
        'TNR'            : 0       ,
        'precision'      : 0       ,
        'recall'         : 0       ,
        'f1'             : 0       ,
        'MCC'            : 0
        ]


def save_list(l,model_path):
    with open(model_path, "wb") as fp:   #Pickling
        pickle.dump(l, fp)
    return

def load_list(model_path):
    with open(model_path, "rb") as fp:   # Unpickling
        b = pickle.load(fp)
    return b

def build_labels_topology(Y, guide, Nf, Nr, deg=64, stride=16):
  
  
    T=np.empty((Nf, Nr),dtype=object)
    
    for i in range(Nf):
        for j in range(Nr):
            T[i,j]=[]
      
    nSlices=Nf
    n4Slice= int (Nr/stride) 
    N=int ((Nr-deg)/stride) +1
    nTot= n4Slice*nSlices
     
    count=0
    for k in range(nTot):
    
            if(k%n4Slice*stride+deg<=Nr):
                if sum(guide[k//n4Slice,  k%n4Slice*stride: k%n4Slice*stride+deg])>0 :
                    count+=1
                    T[k//n4Slice,i].append(-np.ones(3))
                else :
                    for i in range(k%n4Slice*stride , k%n4Slice*stride+deg):
                        T[k//n4Slice,i].append(Y[k-count])
            else :
                if sum(guide[k//n4Slice,  k%n4Slice*stride:  Nr]) + sum(guide[k//n4Slice, 0 : (k%n4Slice*stride+deg)%Nr ]) > 0 :
                    count+=1
                else :
                    for i in range(k%n4Slice*stride , Nr):
                                T[k//n4Slice,i].append(Y[k-count])
                    for i in range((k%n4Slice*stride+deg)%Nr):
                                T[k//n4Slice,i].append(Y[k-count])
                    
    return T

def ricostruisci_from_topology(T,guide,Nf,Nr,deg,stride,thr,thr_ca):
    labels=np.zeros([Nf,Nr])
    
    nSlices=Nf
    n4Slice= int (Nr/stride) 
    N=int ((Nr-deg)/stride) +1
    
    for i in range(2,Nf-2):
        for j in range(Nr):
            
            m=0
            n=3*len(T[i,j])
            Nl=0
            Nc=0
            Ng=0
            for k in range(-2,2):
                for p in T[i+k,j]:
                    if p[1]>thr:
                        Nl+=1
                    elif p[2]>thr_ca:
                        Nc+=1
                    elif p[1]==-1:
                        Ng+=1
            if Nl>(n-Ng)/3:
                labels[i,j]=1
            elif Nc>(n-Ng)/3:
                labels[i,j]=3
            '''
            'k=0
            while k*deg/stride +deg > j :
                k+=1
            i0= k*deg/stride
            for k in range(n4Slice):
                if k*stride < j and j< k*stride +deg: 
                    if guide
            '''
            labels[guide==1]=2
    return labels

def build_matrix_topology(Y, guide, Nf, Nr, deg=64, stride=16):
  
    guide_counter=np.zeros([Nf,Nr])
    
    
    nSlices=Nf
    n4Slice= int (Nr/stride) 
    N=int ((Nr-deg)/stride) +1
    nTot= n4Slice*nSlices
     
    T=np.zeros([Nf, n4Slice,3])
    count=0
    for k in range(nTot):
    
            if(k%n4Slice*stride+deg<=Nr):
                if sum(guide[k//n4Slice,  k%n4Slice*stride: k%n4Slice*stride+deg])>0 :
                    count+=1
                    T[k//n4Slice,k%n4Slice,:]=np.zeros(3)
                    guide_counter[k//n4Slice,  k%n4Slice*stride: k%n4Slice*stride+deg]+=1
                else :
                    T[k//n4Slice,k%n4Slice,:]=Y[k-count]
            else :
                if sum(guide[k//n4Slice,  k%n4Slice*stride:  Nr]) + sum(guide[k//n4Slice, 0 : (k%n4Slice*stride+deg)%Nr ]) > 0 :
                    count+=1
                    T[k//n4Slice,k%n4Slice,:]=-np.zeros(3)
                    guide_counter[k//n4Slice,  k%n4Slice*stride:  Nr]+=1
                    guide_counter[k//n4Slice, 0 : (k%n4Slice*stride+deg)%Nr ]+=1
                else :
                    T[k//n4Slice,k%n4Slice,:]=Y[k-count]
    
    #print(T.shape,T[:,-n4Slice+N:,:].shape)
    #for i in range (N,n)
    T=np.concatenate([T[:,-n4Slice+N:,:],T],axis=1)
    #print(T.shape)
    return T, guide_counter

def ricostruisci_from_topology_matrix(T, guide_counter, guide,Nf,Nr,deg,stride,thr,thr_ca):
    labels=np.zeros([Nf,Nr])
    lipid=np.zeros([Nf,Nr])
    calcium=np.zeros([Nf,Nr])
    
    nSlices=Nf
    n4Slice= int (Nr/stride) 
    N=int ((Nr-deg)/stride) +1
    
    
    l=2
    h=2
    #print(T.shape)
    for i in range(h):
        T=np.concatenate([T[0,:,:][np.newaxis,:,:] ,T,T[-1,:,:][np.newaxis,:,:]] ,axis=0)
    #print(T.shape)
    f=np.ones([2*l+1,2*h])
    labels_slice_l=signal.convolve2d(T[:,:,1], f, 'valid')
    labels_slice_c=signal.convolve2d(T[:,:,2], f, 'valid')
    #print('wooooooo',labels_slice_c.shape,n4Slice)
    #labels=labels
    for i in range(Nf):
        for j in range(Nr//stride):
           lipid[i,j*stride:j*stride+deg]=labels_slice_l[i,j]/ ( (2*l+1)*(2*h) - sum(guide_counter[np.max([i-h,0]):np.min([i+h,Nf]),j*stride]) )
           calcium[i,j*stride:j*stride+deg]=labels_slice_c[i,j]/( (2*l+1)*(2*h) - sum(guide_counter[np.max([i-h,0]):np.min([i+h,Nf]),j*stride]) )
    for i in range(Nf):
        for j in range(Nr//stride):
            if(calcium[i,j*stride]>thr_ca):
                labels[i,j*stride:j*stride+stride]=3
            elif(lipid[i,j*stride]>thr):
                labels[i,j*stride:j*stride+stride]=1
                
    labels[guide==1]=2
    #print(Nr%stride, labels[:,-(Nr%stride):].shape,labels[:,-Nr%stride-1].shape,np.reshape(labels[:,-Nr%stride-1],[Nf,1]).shape)
    if Nr%stride:
        labels[:,-(Nr%stride):]=np.reshape(labels[:,-Nr%stride-1],[Nf,1])
    return labels
        
def pred_to_label_multiclass2(preds, thr=0.5, thr_ca=0.5):
    l = np.zeros(preds.shape[0])

    l = np.argmax(preds, axis=1)

    THR = np.zeros(3)
    THR[0] = 0.
    THR[1] = thr
    THR[2] = thr_ca
    for i in range(preds.shape[0]):

        if (preds[i, l[i]] < THR[l[i]]):
            s = np.argmax(preds[i, np.delete([0, 1, 2], l[i])])

            if (s != 0):
                if preds[i, np.delete([0, 1, 2], np.array([0, l[i]]))] > THR[np.delete([0, 1, 2], np.array([0, l[i]]))]:
                    l[i] = np.delete([0, 1, 2], [0, l[i]])
            else:
                l[i] = 0

    return l


class SectionCanvas(Canvas):
    def __init__(self, parent, *args, **kwargs):
        Canvas.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        self.sRadius=self.parent.canvas_w//2
        self.index = IntVar(master=self.parent, value=0)
        self.zoom_now = DoubleVar(master=self.parent, value=1)#self.screen_radius/(self.parent.Np+120))#self.parent.Np)

        self.zoom_on = IntVar()

        #self.zoom_on = BooleanVar(master=self.parent, value=0)
        #self.zoom_on.set(False)

        self.measure_on = BooleanVar()
        self.measure_on.set(False)
        self.deleting_points =BooleanVar()
        self.deleting_points.set(False)
        self.zoomed_vertex=[0,0]
        '''
        self.zoom_shift1=[0,0]
        self.zoom_shift2 = [0, 0]
        self.zoom_shift = [0, 0]
        '''
        self.radius_index=IntVar(master=self.parent, value=0)
        self.thr = DoubleVar(master=self.parent, value=0.65)#self.parent.pullback.shape[2])
        self.thr_ca = DoubleVar(master=self.parent, value=0.4)
        self.create_image(0, 0, image=[], anchor='nw', tag='section_canvas_image')
        self.create_line(0, 0, 0, 0, fill='blue', width=1, tag='section_canvas_line')

        self.show_labels = IntVar(master=self.parent, value=1)
        self.labels_button = Checkbutton(self, text='Show labels', variable=self.show_labels)#, command=self.update())
        #self.create_window(10, 10, anchor=NW, window=self.labels_button)
        #self.chk.pack(side=RIGHT, anchor=W,expand=True)
        self.classification = IntVar(master=self.parent, value=1)
        self.classification_button = Checkbutton(self, text='Binary classification', variable=self.classification)
        #self.create_window(10, 30, anchor=NW, window=self.classification_button)
        self.old_classification = IntVar(master=self.parent, value=1)#self.classification)
        self.show_lumen_fc=1
        
        self.rescaling = IntVar(master=self.parent, value=0)
        self.rescaling_button = Checkbutton(self, text='Rescale labels', variable=self.rescaling)
        #self.create_window(10, 50, anchor=NW, window=self.rescaling_button)
        self.old_rescaling = IntVar(master=self.parent, value=0)#self.classification)
        
        self.spicchio = IntVar(master=self.parent, value=0)
        self.create_arc(0, 0, 0 , 0,start=0, extent=0, outline='red', fill='', width=1, tag='spicchio_curvac')
        self.create_arc(0, 0, 0 , 0,start=0, extent=0, outline='red', fill='', width=1, tag='spicchio_curval')
        self.create_arc(0, 0, 0 , 0,start=0, extent=0, outline='red', fill='', width=1, tag='spicchio_curvae')
        
        self.starting_angle=IntVar(master=self.parent, value=0)
        self.stopping_angle=IntVar(master=self.parent, value=0)
        

        self.intima_points=[]
        self.create_line(0,0,0,0, fill='red', width=2, tag='intima_line')
        self.create_line(0,0,0,0, fill='green', width=2, tag='calibration_line')
        self.starting_rho=DoubleVar(master=self.parent, value=0)
        self.stopping_rho=DoubleVar(master=self.parent, value=0)
        self.old_intima = []
        self.old_fc = []
        self.old_fct = []
        self.shots=IntVar(master=self.parent, value=0)
        self.reference_shot=IntVar(master=self.parent, value=0)
        self.fixed_points=[]
        self.fixed_points_coord=[]
        self.fixed_points_fc=[]
        self.fixed_points_fc_coord=[]
        
        
        self.old_fixed_points=[np.copy(self.fixed_points)]
        self.old_fixed_points_fc=[np.copy(self.fixed_points_fc)]
        self.select_point=IntVar(master=self.parent, value=0)
        self.interp_start=IntVar(master=self.parent, value=0)
        self.interp_end=IntVar(master=self.parent, value=0)
        self.select_arc=IntVar(master=self.parent, value=0)
        self.selected_arcs=[]
        self.old_selected_arcs=[np.copy(self.selected_arcs)]
        self.selected_arcs_fc=[]
        self.old_selected_arcs_fc=[np.copy(self.selected_arcs_fc)]
        
        self.fc_points=[]
        self.plaque_arcs=[]
        self.fct=[0,np.nan,np.nan]
        
    def zommed_immage(self):
        
        
        k0= self.parent.Np / (self.parent.Np+120) 
        center = (self.sRadius, self.sRadius)
        center_x = self.sRadius
        center_y = self.sRadius
        x1=self.zoom_vertex1[0]
        y1=self.zoom_vertex1[1]
        x2=self.zoom_vertex2[0]
        y2=self.zoom_vertex2[1]

        
        
        if np.sqrt(np.power(x1-x2,2)+np.power(y1-y2,2))>100:
            #if not self.zoom_now.get()==1:
                
            l=np.abs(x1-x2)
            if self.zoom_now.get()==1:
                k=2*self.sRadius/l*k0
            else:
                k=2*self.sRadius/l
            #print(l,np.abs(y1-y2),k)
            self.frame=self.parent.pullback[self.index.get(),:,:]
            self.frame=self.frame/(np.max([np.max(self.frame[300:]),np.max(self.frame[150:]),np.max(self.frame[0:])]))+0.005#(self.parent.pullback)
                #print(np.max(self.frame))
            self.frame[self.frame>1]=1
            self.zoomed_section = cv2.linearPolar(cv2.resize(self.frame,(int(2*self.sRadius*k),int(2*self.sRadius*k)) ,interpolation=cv2.INTER_CUBIC ), (int(self.sRadius*k),int(self.sRadius*k)), int(self.sRadius*k), cv2.WARP_INVERSE_MAP + cv2.WARP_FILL_OUTLIERS)
            
            self.zoom_now.set(k*self.zoom_now.get())
            v=[ int(k*y1) , int(k*x1)  ]
            self.zoomed_vertex=v
            
            #self.frameCart=cv2.resize(self.zoomed_section,(int(2*radius),int(2*radius)),interpolation=cv2.INTER_NEAREST)
            #print(self.zoomed_section.shape, np.max(self.zoomed_section),np.min(self.zoomed_section),np.max(self.zoomed_section[ v[0]:v[0]+2*self.sRadius, v[1]:v[1]+2*self.sRadius ]),np.mean(self.zoomed_section))
            self.frameCart=self.zoomed_section[ v[0]:v[0]+2*self.sRadius, v[1]:v[1]+2*self.sRadius ]
    
            #self.frameCart=np.vstack([np.zeros((100,self.frameCart.shape[1])),self.frameCart ,np.zeros((50,self.frameCart.shape[1]))])
    
            #print(self.zoomed_section.shape,v[0],v[0]+2*self.sRadius, v[1],v[1]+2*self.sRadius)
            #print(self.frameCart.shape, np.max(self.frameCart),np.min(self.frameCart), type(self.frameCart))
        return

    def move_zoom (self, s):
        #s[0] *= self.zoom_now.get()
        #s[1] *= self.zoom_now.get()
        self.zoomed_vertex [0] -= s[1]
        self.zoomed_vertex [1] -= s[0]
        #print(self.zoomed_vertex[0],self.zoomed_vertex[1])
        # self.frameCart=cv2.resize(self.zoomed_section,(int(2*radius),int(2*radius)),interpolation=cv2.INTER_NEAREST)
        # print(self.zoomed_section.shape, np.max(self.zoomed_section),np.min(self.zoomed_section),np.max(self.zoomed_section[ v[0]:v[0]+2*self.sRadius, v[1]:v[1]+2*self.sRadius ]),np.mean(self.zoomed_section))
        self.frameCart = self.zoomed_section[self.zoomed_vertex[0]:self.zoomed_vertex[0] + 2 * self.sRadius, self.zoomed_vertex[1]:self.zoomed_vertex[1] + 2 * self.sRadius]

        #self.frameCart = np.vstack([np.zeros((100, self.frameCart.shape[1])), self.frameCart, np.zeros((50, self.frameCart.shape[1]))])
        
    def update(self):
        if not self.zoom_on.get():
            self.frame=self.parent.pullback[self.index.get(),:,0:self.zoom.get()]
            #self.frame[self.frame>100]=0
            #self.frame[self.frame<50]=0
            #self.frame[(self.frame>0) & (np.std(self.frame,axis=0)>50)]=0
            #self.frame[(self.frame>0) & (np.mean(self.frame,axis=0)<15)]=0
            
            #self.frame=self.frame/(np.max([np.max(self.frame[300:]),np.max(self.frame[150:]),np.max(self.frame[0:])]))+0.005#(self.parent.pullback)
            self.frame=self.frame/(np.max(self.frame))+0.005#(self.parent.pullback)
            
            #print(np.max(self.frame))
            self.frame[self.frame>1]=1
            
            double_colors=1
            neutralize=0
            if self.classification.get()==0 :
                neutralize=number_of_colors-1
                double_colors=2
            '''if self.classification.get()==0 :
                if self.old_classification.get()==1:
                    self.parent.annotation=self.parent.load_labels_continue()
                    self.old_classification.set(0)
                    self.parent.labels=self.parent.labels/(2)+0.5
                #print(self.parent.labels[self.index.get()-0,:])
                elif self.rescaling.get()==1 and self.old_rescaling.get()==0:
                    self.parent.annotation=self.parent.load_labels_continue()
                    self.indx=(self.parent.labels == -number_of_colors + unknown_color_index)
                    self.parent.labels=(self.parent.labels-np.mean(self.parent.labels[[self.parent.labels != -number_of_colors + unknown_color_index]]))/np.std(self.parent.labels[self.parent.labels != -number_of_colors + unknown_color_index])
                    #-np.mean(self.parent.labels)
                    self.parent.labels[self.parent.labels<-1]=-1
                    self.parent.labels[self.parent.labels>1]=1
                    self.parent.labels=self.parent.labels/(4)+0.75
                    #print(np.max(self.parent.labels))
                    #print(np.min(self.parent.labels))
                    self.parent.labels[self.indx]= (-number_of_colors + unknown_color_index)/2+0.5
                    self.old_rescaling.set(1)
                elif self.rescaling.get()==0 and self.old_rescaling.get()==1:
                    self.parent.annotation,self.parent.labels=load_labels_continue()
                    self.old_rescaling.set(0)
                    self.parent.labels=self.parent.labels/(2)+0.5
                neutralize=number_of_colors-1
                double_colors=2
                
            elif self.classification.get()==1 and self.old_classification.get()==0:#self.classification.get()!=self.old_classification.get():
                 self.parent.annotation,self.parent.labels=load_labels_classification()
                 self.old_classification.set(1)
            '''

            b=False
            for i in range (len(self.parent.interv_of_pertinence)):
                if self.index.get() >= self.parent.interv_of_pertinence[i][0] and self.index.get() <self.parent.interv_of_pertinence[i][1]:
                    b=True
            if self.show_labels.get()==1 :
                if b:
                    l=self.parent.labels[self.index.get(),:]
                    a=self.parent.annotation[self.index.get(), :]
                else:
                    l=np.ones(self.parent.Nr)*unknown_color_index
                    a=np.ones(self.parent.Nr) * unknown_color_index
                for i in range(50):                     #anche usando insert(self.frame, z)
                    self.frame=np.hstack([self.frame,np.reshape(l/(number_of_colors-neutralize), [self.parent.Nr,1])]) #np.transpose(labels[section_index,:])
                self.frame=np.hstack([self.frame,np.zeros([self.parent.Nr,20])])  
                for i in range(50):
                    self.frame=np.hstack([self.frame,np.reshape(a/(number_of_colors*double_colors), [self.parent.Nr,1])])

            self.frameCart=self.convert_polar_to_cartesian(self.frame)
            #self.frameCart=np.vstack([np.zeros((100,self.frameCart.shape[1])),self.frameCart ,np.zeros((50,self.frameCart.shape[1]))])
        #cmap = mpl.colors.ListedColormap([RED, YELLOW, WHITE, BLACK])
        #print(self.frameCart.shape, np.max(self.frameCart),np.min(self.frameCart))
        
        if self.classification.get()==0 :
            cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.copper_continue(), N=2*number_of_colors)
        else :
            cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.copper(), N=number_of_colors)
        self.img = Image.fromarray(cmap(self.frameCart, bytes=True))#self.frameCart)
        
        self.parent.prediction_canvas.update(0)
        self.build_measures()
        self.show_plaque_margin(self.index.get())
        self.build_macrophages()
        self.delete("max_diameter")
        self.delete("min_diameter")
        if self.show_lumen_fc:#self.parent.current_class.get()==4 or self.parent.current_class.get()==5):
            self.build_intima()
            self.refresh_points()
            self.build_fc()
            self.show_calibration()
        else:
            self.coords('intima_line',0,0,0,0)
            self.delete('fc_lines')
            self.delete('fct_line')
            self.delete('angle_margin')
            self.delete('point_bar')
            self.delete('section_canvas_line')
            self.delete('calibration_line')

        self.photoImage = ImageTk.PhotoImage(self.img) 
        self.itemconfig('section_canvas_image', image=self.photoImage)

        self.parent.measure_canvas.update()
        if self.zoom_on.get():
            self.delete('angle_margin')
            self.delete('section_canvas_line')
            self.delete("point_bar")
            self.delete("spicchio_curvac")
            self.delete("spicchio_curvae")
            self.delete("spicchio_curval")
            
    def update_line(self, y):
        center_x = self.sRadius
        center_y = self.sRadius
        x = self.sRadius * np.cos(2*np.pi*y / self.parent.Nr) + center_x
        y = self.sRadius * np.sin(2*np.pi*y / self.parent.Nr) + center_y
        self.coords('section_canvas_line', center_x, center_y, x, y)
    '''    
    def _create_circle_arc(self, x, y, r, **kwargs):
        if "start" in kwargs and "end" in kwargs:
            kwargs["extent"] = kwargs["end"] - kwargs["start"]
            del kwargs["end"]
        return self.create_arc(x-r, y-r, x+r, y+r, **kwargs)
    
    def update_circle_arc(self) :
        self.create_circle_arc = self._create_circle_arc
        scaled_depht=(300./1120)*300
        intima=np.min(self.parent.lumen[self.index.get(),self.spicchio.get()*STRIDE:self.spicchio.get()*STRIDE+DEG])*300./1120
        self.create_circle_arc(300, 450, intima+scaled_depht, style="arc",fill='white', width=scaled_depht, start=self.spicchio.get()*STRIDE, end=self.spicchio.get()*STRIDE+DEG)
    '''
    def update_circle_arc(self) :

        theta1=self.spicchio.get()*STRIDE*360/self.parent.Nr
        theta2=(self.spicchio.get()*STRIDE+DEG)*360/self.parent.Nr
        center_x = self.sRadius
        center_y = self.sRadius
        scaled_depht=(SPICCHIO_DEPHT/(self.zoom.get()+120))*self.sRadius
        intima=np.min(self.parent.lumen[self.index.get(),self.spicchio.get()*STRIDE:self.spicchio.get()*STRIDE+DEG])*self.sRadius/(self.zoom.get()+120)
        self.coords('spicchio_curvac', center_x-intima, center_y-intima, center_x+intima, center_y+intima)
        self.coords('spicchio_curval', center_x-intima-scaled_depht, center_y-intima-scaled_depht, center_x+intima+scaled_depht, center_y+intima+scaled_depht)
        self.coords('spicchio_curvae', center_x-self.sRadius , center_y-self.sRadius, center_x+self.sRadius, center_y+self.sRadius)
        self.itemconfig('spicchio_curvac', start=360-theta2,extent=theta2-theta1)
        self.itemconfig('spicchio_curval', start=360-theta2,extent=theta2-theta1)
        self.itemconfig('spicchio_curvae', start=360-theta2,extent=theta2-theta1)
        
    def convert_polar_to_cartesian(self, polar_section):
        center = (self.sRadius, self.sRadius)
        cartesian_section = cv2.linearPolar(cv2.resize(polar_section, (2*self.sRadius,2*self.sRadius),interpolation=cv2.INTER_NEAREST ), center, self.sRadius, cv2.WARP_INVERSE_MAP + cv2.WARP_FILL_OUTLIERS)
        return cartesian_section


    @staticmethod
    def copper():
        x = np.linspace(0, 1, number_of_colors)
        r = 30
        R = np.arctan(r*x)/np.arctan(r)
        g = 10
        G = np.arctan(g*x)/np.arctan(g)
        b = 0.01
        B = np.arctan(b*x)/np.arctan(b)

        cmap = np.vstack((R, G, B))
        cmap = cmap.transpose()

        cmap[non_lipid_color_index, :] = RED
        cmap[lipid_color_index, :] = YELLOW
        cmap[calcium_color_index, :] = WHITE
        cmap[unknown_color_index, :] = BLACK
        cmap[deep_plaque_color_index, :] = ORANGE
        return cmap


    @staticmethod
    def copper_continue():
        x = np.linspace(0, 1, number_of_colors)
        r = 30
        R = np.arctan(r*x)/np.arctan(r)
        g = 10
        G = np.arctan(g*x)/np.arctan(g)
        b = 0.01
        B = np.arctan(b*x)/np.arctan(b)

        cmap = np.vstack((R, G, B))

        x = np.linspace(0, 1, number_of_colors)
        r = 30
        R = np.arctan(r*(1-x))/np.arctan(r)
        g = 30
        G = np.arctan(g*x)/np.arctan(g)
        b = 0.000000000
        B = np.arctan(b*x)/np.arctan(g)

        cmap2 = np.vstack((R, G, B))
        cmap = np.hstack((cmap,cmap2))
        cmap = cmap.transpose()

        cmap[non_lipid_color_index, :] = RED
        cmap[lipid_color_index, :] = YELLOW
        cmap[calcium_color_index, :] = WHITE
        cmap[unknown_color_index, :] = BLACK

        return cmap

    def get_angle_index(self,X,Y):
        x= X - self.sRadius
        y= Y - self.sRadius #- 100
        if x==0 and y>0:
            a = np.pi/2 
        elif x==0 and y<0:
            a = 3*np.pi/2
        else:        
            a=np.arctan(y/x)
            if x<0:
                a+= np.pi 
            if x>0 and y<0:
                a+= 2*np.pi
        b=int(a*self.parent.Nr//(2*np.pi))
        return b
    
    def calculate_area(self,f):
        k=0.8185/160
        h=np.pi*2/self.parent.Nr
        I= h / 2 * integrate.simps(np.power(k*self.parent.lumen[f,:],2) ,even='avg')# x=np.arange(self.parent.lumen.shape[0]),even='avg')
        self.parent.areas[f]=I
        return
    
    def calculate_plaque_angle(self,frame):
        l=self.parent.lipid[frame,:]
        bx,by=self.baricentre(frame)
        angle=0
        s=[]
        e=[]
        
        i=0
        while l[i]==1 and i<len(l)-1 :
            i+=1
        if (i==len(l)-1) :
            angle=360
        else:
            if i==0:
                #and not(i%len(l) in e)
                while l[i%len(l)]==0  and i<len(l) :
                    i+=1
                if (i==len(l)) :
                    angle=0
                else :
                    s.append(i)
                    while l[i%len(l)]==1 and i<len(l)-1:
                        i+=1
                    e.append(i)
                    f=0
                    while f==0:
                        ns,ne=self.find_next_arc(i, l)
                        if not(ns in s):
                            s.append(ns)
                            e.append(ne)
                            i=e[-1]
                            #print('arc found',e[-1],s,e)
                        else:
                            f=1
            else:
                #e.append(i)
                f=0
                while f==0:
                    ns,ne=self.find_next_arc(i, l)
                    if not(ns in s) and not(ne in e):
                        s.append(ns) 
                        e.append(ne)
                        i=e[-1]
                    else:
                        f=1
            center_x = self.sRadius
            center_y = self.sRadius
            #print(s)
            #print(e)
            self.plaque_arcs=[]
            if len(s):# and e :
                self.plaque_arcs=[ (s[i],e[i]) for i in range(len(s)) ]
                #print('yo',self.plaque_arcs)
                
            self.delete('angle_margin')
            for j in range(len(s)):
                r1=self.parent.lumen[frame,s[j]]/(self.zoom.get()+120)*self.sRadius
                x1 = r1 * np.cos(2*np.pi*s[j] / self.parent.Nr) + center_x
                y1 = r1 * np.sin(2*np.pi*s[j] / self.parent.Nr) + center_y
                r2=self.parent.lumen[frame,e[j]]/(self.zoom.get()+120)*self.sRadius
                x2 = r2 * np.cos(2*np.pi*e[j] / self.parent.Nr) + center_x
                y2 = r2 * np.sin(2*np.pi*e[j] / self.parent.Nr) + center_y

                det = (x1 - bx) * (y2 - by) - (y1 - by) * (x2 - bx)
                '''
                dot = (x1-bx) *(y1-by) + (x2-bx) * (y2-by)
                
                a = math.atan2(det, dot)
                print(a,a / np.pi * 180)
                if a<0:
                    a=2*np.pi + a

                a = a / np.pi * 180
                print(a)
                '''


                a=np.arccos( ( (x1-bx)*(x2-bx) + (y1-by)*(y2-by) ) / (np.sqrt((x1-bx)*(x1-bx)+(y1-by)*(y1-by)) * np.sqrt((x2-bx)*(x2-bx)+(y2-by)*(y2-by)) ) )

                #print('estremi placca',s,e)
                #print('arccos',a/np.pi*180)

                a=a/np.pi*180

                if det <0 :
                    a=360-a
                '''
                if s[j]<e[j]:
                    if e[j] - s[j]>len(l)/2:
                        a=360-a
                else:
                    if - s[j] + len(l) + e[j]>len(l)/2:
                        a=360-a
                
                '''
                #print(a)

                angle+=a
                #print(angle)
                self.create_line(bx, by, x1, y1, fill='green', width=1, tag='angle_margin')
                self.create_line(bx, by, x2, y2, fill='green', width=1, tag='angle_margin')
                self.parent.plaque_angle[frame]=angle

        self.parent.plaque_arcs[frame]=np.copy(self.plaque_arcs)
        if np.isnan(angle):
            angle=0
            print("angle is nan")
        self.parent.plaque_angle[frame]=angle
        
        return angle
                
    def find_next_arc(self,i,l):
        while l[i%len(l)]==1:
            i+=1
        while l[i%len(l)]==0  :
            i+=1
            if(i>self.parent.Nr*2):
                return 0,0
        s=i%len(l)
        while l[i%len(l)]==1:
            i+=1
        e=i%len(l)
        return s,e

    def find_guide_arc(self,f):
        s,e=self.find_next_arc(0, self.parent.guide[f])
        if not(s==0 and e==0):
            self.parent.guide_arcs[f]=[]
            self.parent.guide_arcs[f].append(tuple((int(e),int(s))))
        
        
        
    def baricentre(self,f):
        center_x = self.sRadius
        center_y = self.sRadius
        l=self.parent.lumen[f,:]
        x=0
        y=0
        tw=0
        for i in range(len(l)):
            r=self.parent.lumen[f,i]#/(self.zoom.get()+120)*radius
            w=r*2*np.pi / self.parent.Nr
            tw+=w
            x += w * r * np.cos(2*np.pi*i / self.parent.Nr) #+ center_x
            y += w * r * np.sin(2*np.pi*i / self.parent.Nr) #+ center_y
        x/=tw#len(l)
        y/=tw#len(l)
        x=x*self.sRadius/(self.zoom.get()+120)#/np.sqrt(x*x+y*y)
        y=y*self.sRadius/(self.zoom.get()+120)#//np.sqrt(x*x+y*y)
        x=x+ center_x
        y=y+ center_y
        point_tag="point_bar"#len(self.fixed_points)-1)
        #print(point_tag)
        self.delete("point_bar")
        self.create_oval(x-2,y-2,x+2,y+2, tags=point_tag, fill='green')
        #print(x,y)
        return x,y
    
    
    def test_diameter(self,f):
        bx,by=self.baricentre(f)
        B=np.array([bx,by])
        
        self.delete("max_diameter")
        
        eps=0.01
        m=2000
        M=0
        i=0
        d=self.parent.Nr//2-1
        a=1
        while np.abs(a)> eps:
            #print('azz')
            r=self.parent.lumen[f,i]*self.sRadius/(self.zoom.get()+120)
            P=np.array([r*np.cos(2*np.pi*i / self.parent.Nr),r*np.sin(2*np.pi*i / self.parent.Nr)])+self.sRadius
            r=self.parent.lumen[f,d]*self.sRadius/(self.zoom.get()+120)
            D=np.array([r*np.cos(2*np.pi*d / self.parent.Nr),r*np.sin(2*np.pi*d / self.parent.Nr)])+self.sRadius
            BP= P - B
            BD= D - B
            print(BP,BD)
            a=np.arcsin( np.cross(BP,BD) / (np.linalg.norm(BP)*np.linalg.norm(BD)) )
            print(a)
            d=int((d + a*self.parent.Nr/(2*np.pi)/2) %self.parent.Nr)
            print(d)
            MP=P
            r=self.parent.lumen[f,d]*self.sRadius/(self.zoom.get()+120)
            D=np.array([r*np.cos(2*np.pi*d / self.parent.Nr),r*np.sin(2*np.pi*d / self.parent.Nr)])+self.sRadius
            MD=D  
            self.create_line(P[0],P[1],B[0],B[1], fill='blue', width=1, tag='max_diameter')
            self.create_line(D[0],D[1],B[0],B[1], fill='blue', width=1, tag='max_diameter')
            self.create_line(MP[0],MP[1],MD[0],MD[1], fill='red', width=1, tag='max_diameter')
            self.parent.wait_variable(self.parent.v_stop)
        
        
    def calculate_diameter(self,f):
        
        bx,by=self.baricentre(f)
        B=np.array([bx,by])
        
        self.delete("max_diameter")
        self.delete("min_diameter")
        
        last_diameter=self.parent.Nr//2
        eps=0.02
        m=2000
        M=0
        i=0
        d=self.parent.Nr//2-1
        while i < last_diameter:
            a=1
            d=(d+1)%self.parent.Nr
            k=0
            while abs(a)> eps and k<10:
                #self.delete("diameter")
                r=self.parent.lumen[f,i]*self.sRadius/(self.zoom.get()+120)
                P=np.array([r*np.cos(2*np.pi*i / self.parent.Nr),r*np.sin(2*np.pi*i / self.parent.Nr)])+self.sRadius
                r=self.parent.lumen[f,d]*self.sRadius/(self.zoom.get()+120)
                D=np.array([r*np.cos(2*np.pi*d / self.parent.Nr),r*np.sin(2*np.pi*d / self.parent.Nr)])+self.sRadius
                BP= P - B
                BD= D - B
                a=np.arcsin( np.cross(BP,BD) / (np.linalg.norm(BP)*np.linalg.norm(BD)) )
                d=int((d + a*self.parent.Nr/(2*np.pi)/2)%self.parent.Nr)
                k+=1
                #self.create_line(P[0],P[1],B[0],B[1], fill='blue', width=1, tag='diameter')
                #self.create_line(D[0],D[1],B[0],B[1], fill='blue', width=1, tag='diameter')
                #self.create_line(P[0],P[1],D[0],D[1], fill='red', width=1, tag='diameter')
                #self.parent.wait_variable(self.parent.v_stop)
            
            l = np.linalg.norm(P-D)
            if l > M:
                max_diameter = [l,i,d]
                M = l
                MP = P
                MD = D
            if l < m:
                min_diameter = [l,i,d]
                m = l
                mP = P
                mD = D
            if i == 0:
                last_diameter=d
            i += 1

        self.create_line(MP[0],MP[1],MD[0],MD[1], fill='red', width=1, tag='max_diameter')
        self.create_line(mP[0],mP[1],mD[0],mD[1], fill='red', width=1, tag='min_diameter')
        '''
        font_object = font.Font(family='Helvetica', size='20')
        self.parent.measure_canvas.last_line+= self.parent.measure_canvas.spaceline
        self.parent.measure_canvas.create_text(0, self.parent.measure_canvas.last_line, text='Max diameter' +str(round(max_diameter[0] * 0.88 / 196 * 1000)) + 'm-6', 
                                               tags='diameters', fill='green', font=font_object,anchor=NW)
        self.parent.measure_canvas.last_line+= self.parent.measure_canvas.spaceline
        self.parent.measure_canvas.create_text(0, self.parent.measure_canvas.last_line, text='Min diameter' +str(round(min_diameter[0] * 0.88 / 196 * 1000)) + 'm-6',
                                               tags='diameters', fill='green', font=font_object,anchor=NW)
        '''
        #self.parent.measure_canvas.print_diameters(min_diameter[0],max_diameter[0])
        self.parent.diameters[0,f,:]=np.asarray(min_diameter)
        self.parent.diameters[1,f,:]=np.asarray(max_diameter)
        self.parent.measure_canvas.update()
    
    def show_plaque_margin(self,f):
        self.delete('angle_margin')
        center_x = self.sRadius
        center_y = self.sRadius
        bx,by=self.baricentre(f)
        p=self.parent.plaque_arcs[f]
        if len(p):
            for j in range(len(p)):
                r1=self.parent.lumen[f,p[j][0]]/(self.zoom.get()+120)*self.sRadius
                x1 = r1 * np.cos(2*np.pi*p[j][0] / self.parent.Nr) + center_x
                y1 = r1 * np.sin(2*np.pi*p[j][0] / self.parent.Nr) + center_y
                r2=self.parent.lumen[f,p[j][1]]/(self.zoom.get()+120)*self.sRadius
                x2 = r2 * np.cos(2*np.pi*p[j][1] / self.parent.Nr) + center_x
                y2 = r2 * np.sin(2*np.pi*p[j][1] / self.parent.Nr) + center_y
                self.create_line(bx, by, x1, y1, fill='green', width=1, tag='angle_margin')
                self.create_line(bx, by, x2, y2, fill='green', width=1, tag='angle_margin')
                
    def show_calibration(self):
        center_x = self.sRadius
        center_y = self.sRadius
        k0=self.parent.Np/(self.parent.Np+120)
        #if not self.zoom_on.get():
        points=[]
        
        self.create_line(0,0,0,0, fill='green', width=2, tag="calibration_line")
        if not self.zoom_on.get():
            for i in range(self.parent.Nr):
                r=self.parent.calibration*self.sRadius/(self.zoom.get()+120)
                x = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x +self.parent.cal_center_x.get()*self.sRadius/(self.zoom.get()+120)
                y = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y +self.parent.cal_center_y.get()*self.sRadius/(self.zoom.get()+120)
                points.append(x)
                points.append(y)
            self.coords('calibration_line', points)
            
        else : 
            R=self.zoomed_section.shape[0]/2
            for i in range(self.parent.Nr):
                r=self.parent.calibration/(self.zoom.get())*R
                x = r * np.cos(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[1] +self.parent.cal_center_x.get()/(self.zoom.get())*R
                y = r * np.sin(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[0] +self.parent.cal_center_y.get()/(self.zoom.get())*R#+ 100
                points.append(x)
                points.append(y)
            self.coords('calibration_line', points)
    
        
                
    def build_intima (self):
        center_x = self.sRadius
        center_y = self.sRadius
        k0=self.parent.Np/(self.parent.Np+120)
        #if not self.zoom_on.get():
        self.intima_points=[]
        if not self.zoom_on.get():
            for i in range(self.parent.Nr):
                r=self.parent.lumen[self.index.get(),i]*self.sRadius/(self.zoom.get()+120)
                x = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x 
                y = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                self.intima_points.append(x)
                self.intima_points.append(y)
            self.coords('intima_line', self.intima_points)
            
        else : 
            R=self.zoomed_section.shape[0]/2
            for i in range(self.parent.Nr):
                r=self.parent.lumen[self.index.get(),i]/(self.zoom.get())*R
                x = r * np.cos(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[1] 
                y = r * np.sin(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                self.intima_points.append(x)
                self.intima_points.append(y)
            self.coords('intima_line', self.intima_points)
    
    def build_fc (self):
        
        center_x = self.sRadius
        center_y = self.sRadius
        self.delete('fc_lines')
        self.delete('fct_line')
        self.fc_points=[]
        self.plaque_arcs=self.parent.plaque_arcs[self.index.get()]
        #print(self.plaque_arcs)
        if not self.zoom_on.get():
            for k in range(len(self.plaque_arcs)):
                fc_points=[]
                if self.plaque_arcs[k][0]<self.plaque_arcs[k][1] :  
                    for i in range (self.plaque_arcs[k][0],self.plaque_arcs[k][1]):
                        r=self.parent.fc[self.index.get(),i]/(self.zoom.get()+120)*self.sRadius
                        x = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x
                        y = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                        fc_points.append(x)
                        fc_points.append(y)
                else:
                    for i in range (self.plaque_arcs[k][0],self.parent.Nr):
                        r=self.parent.fc[self.index.get(),i]/(self.zoom.get()+120)*self.sRadius
                        x = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x
                        y = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                        fc_points.append(x)
                        fc_points.append(y)
                    for i in range (self.plaque_arcs[k][1]):
                        r=self.parent.fc[self.index.get(),i]/(self.zoom.get()+120)*self.sRadius
                        x = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x
                        y = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                        fc_points.append(x)
                        fc_points.append(y)
                line_tag="line"+str(k)
                self.create_line(fc_points, fill='green', width=2, tags=(line_tag,"fc_lines"))
                self.build_fct()
        else:
            R=self.zoomed_section.shape[0]/2
            
            for k in range(len(self.plaque_arcs)):
                fc_points=[]
                if self.plaque_arcs[k][0]<self.plaque_arcs[k][1] :  
                    for i in range (self.plaque_arcs[k][0],self.plaque_arcs[k][1]):
                        r=self.parent.fc[self.index.get(),i]/(self.zoom.get())*R
                        x = r * np.cos(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[1] 
                        y = r * np.sin(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                
                        fc_points.append(x)
                        fc_points.append(y)
                else:
                    for i in range (self.plaque_arcs[k][0],self.parent.Nr):
                        r=self.parent.fc[self.index.get(),i]/(self.zoom.get())*R
                        x = r * np.cos(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[1] 
                        y = r * np.sin(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                
                        fc_points.append(x)
                        fc_points.append(y)
                    for i in range (self.plaque_arcs[k][1]):
                        r=self.parent.fc[self.index.get(),i]/(self.zoom.get())*R
                        x = r * np.cos(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[1] 
                        y = r * np.sin(2*np.pi*i / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                
                        fc_points.append(x)
                        fc_points.append(y)
                line_tag="line"+str(k)
                self.create_line(fc_points, fill='green', width=2, tags=(line_tag,"fc_lines"))
            self.build_fct()
        
    def build_fct(self):
        
        center_x = self.sRadius
        center_y = self.sRadius
        self.fc=self.parent.fc[self.index.get()]
        self.fct=self.parent.fct[self.index.get()]
        #print(self.index.get(),self.fct,self.fct[1],self.fct[2])
        if not np.isnan(self.fct[1]):
            if not self.zoom_on.get():
                r=self.parent.lumen[self.index.get(),int(self.fct[1])]/(self.zoom.get()+120)*self.sRadius        
                x1 = r * np.cos(2*np.pi*self.fct[1] / self.parent.Nr) + center_x
                y1 = r * np.sin(2*np.pi*self.fct[1] / self.parent.Nr) + center_y
                r=self.parent.fc[self.index.get(),int(self.fct[2])]/(self.zoom.get()+120)*self.sRadius
                x2 = r * np.cos(2*np.pi*self.fct[2] / self.parent.Nr) + center_x
                y2 = r * np.sin(2*np.pi*self.fct[2] / self.parent.Nr) + center_y
                self.create_line(x1,y1,x2,y2, fill='blue', width=2, tags=("fct_line"))
            else:
                R=self.zoomed_section.shape[0]/2
                r=self.parent.lumen[self.index.get(),int(self.fct[1])]/(self.zoom.get())*R     
                x1 = r * np.cos(2*np.pi*self.fct[1] / self.parent.Nr) + R - self.zoomed_vertex[1] 
                y1 = r * np.sin(2*np.pi*self.fct[1] / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                r=self.parent.fc[self.index.get(),int(self.fct[2])]/(self.zoom.get())*R
                x2 = r * np.cos(2*np.pi*self.fct[2] / self.parent.Nr) + R - self.zoomed_vertex[1] 
                y2 = r * np.sin(2*np.pi*self.fct[2] / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                self.create_line(x1,y1,x2,y2, fill='blue', width=2, tags=("fct_line"))
            
            
    def refresh_points (self):
        self.delete('points')
        
        center_x = self.sRadius
        center_y = self.sRadius
        if not self.zoom_on.get():
            for i in range(len(self.fixed_points)):
                r=self.fixed_points[i][1]*self.sRadius/(self.zoom.get()+120)
                t=self.fixed_points[i][0]*2*np.pi/self.parent.Nr
                x=r*np.cos(t)
                y=r*np.sin(t)
                point_tag="point"+str(i)#len(self.fixed_points)-1)
                #print(point_tag)
                self.create_oval(x-3+center_x,y-3+center_y,x+3+center_x,y+3+center_y, tags=(point_tag,"points"), fill='blue')

            for i in range(len(self.fixed_points_fc)):
                r=self.fixed_points_fc[i][1]*self.sRadius/(self.zoom.get()+120)
                t=self.fixed_points_fc[i][0]*2*np.pi/self.parent.Nr
                x=r*np.cos(t)
                y=r*np.sin(t)
                point_tag="point_fc"+str(i)#len(self.fixed_points)-1)
                #print(point_tag)
                self.create_oval(x-3+center_x,y-3+center_y,x+3+center_x,y+3+center_y, tags=(point_tag,"points"), fill='blue')
        else:
            self.delete('points')
            
    def build_macrophages(self):
        self.delete('mph')
        center_x = self.sRadius
        center_y = self.sRadius
        for X in self.parent.macrophages:
            if(X[0]==self.index.get()):
                
                if not self.zoom_on.get():
                    r = X[2]/(self.zoom.get()+120)*self.sRadius        
                    x = r * np.cos(2*np.pi*X[1] / self.parent.Nr) + center_x
                    y = r * np.sin(2*np.pi*X[1] / self.parent.Nr) + center_y
                    self.create_oval(x-3,y-3,x+3,y+3, fill='white', width=2, tags=("mph"))
                else:
                    R=self.zoomed_section.shape[0]/2
                    r=X[2]/(self.zoom.get())*R     
                    x = r * np.cos(2*np.pi*X[1] / self.parent.Nr) + R - self.zoomed_vertex[1] 
                    y = r * np.sin(2*np.pi*X[1] / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                    self.create_oval(x-3,y-3,x+3,y+3, fill='white', width=2, tags=("mph"))
            
            
    def build_measures(self):
        self.delete('measure_points','measure')
        l=len(self.parent.measure_points[self.index.get()])
        if not self.zoom_on.get():
            for i in range(l):
                if i%2==0:
                    r=self.parent.measure_points[self.index.get()][i][1]*self.sRadius/(self.zoom.get()+120)
                    x1=r * np.cos(2*np.pi*self.parent.measure_points[self.index.get()][i][0] / self.parent.Nr) + self.sRadius
                    y1=r * np.sin(2*np.pi*self.parent.measure_points[self.index.get()][i][0] / self.parent.Nr) + self.sRadius
                    #x1=self.parent.measure_points_cart[self.index.get()][i][0]
                    #y1=self.parent.measure_points_cart[self.index.get()][i][1]

                    point_tag = "measure_point" + str(i)
                    self.create_oval(x1 - 3, y1 - 3, x1 + 3 , y1 + 3 , tags=(point_tag, "measure_points"), fill='green')
                else:
                    line_tag = "measure" + str(i//2)
                    #x0=self.parent.measure_points_cart[self.index.get()][i-1][0]
                    #y0=self.parent.measure_points_cart[self.index.get()][i-1][1]
                    r=self.parent.measure_points[self.index.get()][i][1]*self.sRadius/(self.zoom.get()+120)
                    x0=r * np.cos(2*np.pi*self.parent.measure_points[self.index.get()][i][0] / self.parent.Nr) + self.sRadius
                    y0=r * np.sin(2*np.pi*self.parent.measure_points[self.index.get()][i][0] / self.parent.Nr) + self.sRadius

                    d=np.sqrt(np.power(x0 - x1,2) + np.power(y0 - y1,2))
                    d= d *(self.zoom.get()+120)/self.sRadius

                    self.create_line( x0 , y0, x1, y1, fill='green', width=1, tags=(line_tag, "measure"))
                    self.create_text(60, 20*(i//2) +30 , text='Measure '+chr(65+i//2)+'= ' +str(int(d* self.parent.calibration_scale)), tags=['show_measure',line_tag, "measure"], fill='green')
                    self.create_text(x0*1.01,y0*1.01 , text=chr(65+(i//2)), tags=['letter',line_tag, "measure"], fill='green')
        else:
            for i in range(l):
                if i%2==0:
                    R=self.zoomed_section.shape[0]/2
                    r=self.parent.measure_points[self.index.get()][i][1]/(self.zoom.get())*R     
                    x1 = r * np.cos(2*np.pi*self.parent.measure_points[self.index.get()][i][0] / self.parent.Nr) + R - self.zoomed_vertex[1] 
                    y1 = r * np.sin(2*np.pi*self.parent.measure_points[self.index.get()][i][0] / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                    #x1c=self.parent.measure_points_cart[self.index.get()][i][0]
                    #y1c=self.parent.measure_points_cart[self.index.get()][i][1]
                        
                    point_tag = "measure_point" + str(i)
                    self.create_oval(x1 - 3, y1 - 3, x1 + 3 , y1 + 3 , tags=(point_tag, "measure_points"), fill='green')
                else:
                    line_tag = "measure" + str(i//2)
                    r=self.parent.measure_points[self.index.get()][i][1]/(self.zoom.get())*R     
                    x0 = r * np.cos(2*np.pi*self.parent.measure_points[self.index.get()][i][0] / self.parent.Nr) + R - self.zoomed_vertex[1] 
                    y0 = r * np.sin(2*np.pi*self.parent.measure_points[self.index.get()][i][0] / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                    self.create_line( x0 , y0, x1, y1, fill='green', width=1, tags=(line_tag, "measure"))
                    self.create_text(x0*1.01,y0*1.01 , text=chr(65+(i//2)), tags=['letter',line_tag, "measure"], fill='green')
                    
                    #x0=self.parent.measure_points_cart[self.index.get()][i-1][0]
                    #y0=self.parent.measure_points_cart[self.index.get()][i-1][1]
                    d=np.sqrt(np.power(x0 - x1,2) + np.power(y0 - y1,2))
                    d= d *(self.zoom.get()+120)/self.sRadius
                    self.create_text(60, 20*(i//2)  +30 , text='Measure '+chr(65+i//2)+'= ' +str(int(d* self.parent.calibration_scale/self.zoom_now.get()*1000/1120)), tags=['show_measure',line_tag, "measure"], fill='green')

    def add_shot(self):
        self.old_intima.append(np.copy(self.parent.lumen))
        self.old_fc.append(np.copy(self.parent.fc))
        self.old_fixed_points.append(np.copy(self.fixed_points))
        self.old_fixed_points_fc.append(np.copy(self.fixed_points_fc))
        self.old_selected_arcs.append(np.copy(self.selected_arcs))
        self.old_selected_arcs_fc.append(np.copy(self.selected_arcs_fc))
        self.old_fct.append(np.copy(self.parent.fct))
        self.shots.set(self.shots.get()+1)
       
    def cut_history(self):
        self.old_intima=self.old_intima[:self.shots.get()+1]
        self.old_fc=self.old_fc[:self.shots.get()+1]
        self.old_fixed_points=self.old_fixed_points[:self.shots.get()+1]
        self.old_fixed_points_fc=self.old_fixed_points_fc[:self.shots.get()+1]
        self.old_selected_arcs=self.old_selected_arcs[:self.shots.get()+1]
        self.old_selected_arcs_fc=self.old_selected_arcs_fc[:self.shots.get()+1]
        
        
    def guida (self,s,e):
        
        G=np.zeros([e-s,self.parent.Nr])
        for i in range(s,e):
            for j in range(self.parent.Nr):
                c=0
                for k in range(170,self.parent.Np-10):
                    if (np.mean(self.parent.pullback[i,j,k:k+10])>40):
                        c=1
                        break
                G[i-s,j]=1-c
        return G        
        
    def guida_fit (self,s,e):
        
        G=np.ones([e-s,self.parent.Nr])
        for i in range(s,e):
            for j in range(self.parent.Nr):
                v1=0
                v2=0
                for k in range(100,self.parent.Np):
                    if (self.parent.pullback[i,j,k]>100):
                        if v1==0:
                            v1=k
                        else :
                            v2=k
                    if v2-v1>50:
                            G[i-s,j]=0
                            break
                      
        return G                        
    
    def intensities_grad(self,I,d=1):
        G=np.zeros(I.shape)#self.parent.pullback.shape[1:])
        s=18
        K=np.arange(41)-20
        K=-1/(np.sqrt(2*np.pi)*s*s*s) * np.multiply(K , np.exp(-1/2*np.multiply(K,K)/(s*s))) 
        for t in range(G.shape[0]):
            G[t,:]=np.convolve(I[t],K,'same')
        G=-d*G
        G=G-np.min(G)
        G=G/np.max(G)
        return G
    
    
    def intensities_grad_fc(self,I,d=1):
        G=np.zeros(I.shape)#self.parent.pullback.shape[1:])
        s_light=18
        s_dark=30#25
        s=s_light
        K=np.arange(20)-20
        K=-1/(np.sqrt(2*np.pi)*s*s*s) * np.multiply(K , np.exp(-1/2*np.multiply(K,K)/(s*s))) 
        #print(K)
        s=s_dark
        L=np.arange(31)
        L=-1/(np.sqrt(2*np.pi)*s*s*s) * np.multiply(L , np.exp(-1/2*np.multiply(L,L)/(s*s))) 
        #print(L)
        K=np.concatenate([K,L])
        #print(K.shape,K)
        for t in range(G.shape[0]):
            G[t,:]=np.convolve(I[t],K,'same')
        G=-d*G
        G=G-np.min(G)
        G=G/np.max(G)
        return G

    '''
    def best_paths(self,P,C,G,n,t,f):
        Q=np.zeros((P.shape[0]+1,P.shape[1]))
        Q[P.shape[0],:]=np.arange(P.shape[1])
        #Q=np.concatenate([P,np.reshape(np.arange(P.shape[1]),[1,P.shape[1]])],axis=1)
        print(P[:,200].shape,Q[:-1,200].shape)
        for j in range(60,900):
            bp=np.argmax(C[j-n:j+n])
            C[j]+=G[t,j]
            Q[:-1,j]=P[:,bp]
        if t+1 < G.shape[0]:
            Q,C=self.best_paths(Q, C, G, n, t+1, f)
        return Q,C
     '''
    '''   
    def best_paths(self,P,C,G,n,t,r1,r2): #funziiona versione base
        Q=np.zeros((P.shape[0]+1,P.shape[1]))
        Q[P.shape[0],:]=np.arange(P.shape[1])
        C2=np.zeros(C.shape)
        #Q=np.concatenate([P,np.reshape(np.arange(P.shape[1]),[1,P.shape[1]])],axis=1)
        print(P.shape,Q.shape)
        for j in range(r1,r2):
            bp=np.argmax( C[ np.max([j-n,0]): np.min([j+n,P.shape[1]]) ] ) + np.max([j-n,0])
            C2[j]=C[bp]+G[t,j]
            Q[:-1,j]=P[:,bp]
        return Q,C2    
    '''
    def best_paths(self,P,C,G,n,t,r1,r2,d): 
        Q=np.zeros((P.shape[0]+1,P.shape[1]))
        Q[P.shape[0],:]=np.arange(P.shape[1])
        C2=np.zeros(C.shape)
        for j in range(r1,r2):
            Cp=np.zeros(  np.min([ j+n , r2 ]) - np.max ([j-n+1,r1]) )
            for k in range( np.max([j-n+1,r1]) , np.min([ j+n , r2 ]) ):
                Cp[k-int(np.max([j-n+1,r1]))]= C[k] + (G[t,j] + G[t-1,k])*d[np.abs(k-j)]
            bp=np.argmin(Cp) + np.max([j-n+1,r1])
            C2[j]=Cp[bp-np.max([j-n+1,r1])]  #C[bp]+ (G[t,j] + G[t-1,bp])*(1+alpha*np.power(np.abs(bp-j),beta))
            Q[:-1,j]=P[:,bp]
            #if(j==200):
            #   print(bp,P[:,bp],Q[:,j])
            #print(Q[:,r1:r2])
            #print(Cp,C2[j])
            #print(bp)
        
        return Q,C2    

    def best_paths_opt(self,P,C,G,n,t,r1,r2,d): 
        Q=np.zeros((P.shape[0]+1,P.shape[1]))
        Q[P.shape[0],:]=np.arange(P.shape[1])
        C2=np.ones(C.shape)*1000
        Cp=np.zeros( 2*n-1 )
        for j in range(r1 + n,r2-n):
            e1=j-n+1
            e2=j+n
            Cp= C[e1:e2] + np.multiply(G[t,j] + G[t-1,e1: e2],d)  
            bp=np.argmin(Cp) 
            C2[j]=Cp[bp]  #C[bp]+ (G[t,j] + G[t-1,bp])*(1+alpha*np.power(np.abs(bp-j),beta))
            Q[:-1,j]=P[:,bp+ e1]
        return Q,C2    

    def best_paths_rec(self,P,C,G,n,t,stop,r1,r2,d): 
        Q=np.zeros((P.shape[0]+1,P.shape[1]))
        Q[P.shape[0],:]=np.arange(P.shape[1])
        C2=np.ones(C.shape)*1000
        Cp=np.zeros( 2*n-1 )
        for j in range(r1 + n,r2-n):
            e1=j-n+1
            e2=j+n
            Cp= C[e1:e2] + np.multiply(G[t,j] + G[t-1,e1: e2],d)  
            bp=np.argmin(Cp) 
            C2[j]=Cp[bp]  #C[bp]+ (G[t,j] + G[t-1,bp])*(1+alpha*np.power(np.abs(bp-j),beta))
            Q[:-1,j]=P[:,bp+ e1]
        if t+1 < stop:#G.shape[0]:
            Q,C2=self.best_paths_rec(Q, C2, G, n, t+1,stop, r1,r2, d)
        
        return Q,C2    


    def lumen_weights(self):
        n=30
        alpha=0.01
        beta=1
        d=np.zeros(2*n-1)
        for j in range(0,2*n-1):
            d[j]=(1+alpha*np.power(np.abs(j-n+1),beta))
        return d
   
    def boundary(self,f,d):

        n=30
        r1=45
        r2=900
        pad=10
        
        T=time.time()
        
        self.parent.lumen[f,:]=0
        #gs,ge=self.find_next_arc(0, self.parent.guide[f,] )
        gs=self.parent.guide_arcs[f][0][1]        
        ge=self.parent.guide_arcs[f][0][0]        
        I=np.copy(self.parent.pullback[f,:,:])#np.zeros(self.parent.pullback[f,:,:].shape)
        #I[self.parent.guide[f,:]==1,:]=0
        gs = (gs -pad + self.parent.Nr)%self.parent.Nr
        ge = (ge + pad)%self.parent.Nr
        print('frame, guide arc',f,gs,ge)
        
        if gs > ge :
            I[gs:,:]=0
            I[:ge,:]=0
        else:
            I[gs:ge,:]=0
        
        G=self.intensities_grad(I)#self.parent.pullback[f,:,:])
        #G[self.parent.guide[f,:]==1,:]=0.1
        #G=self.intensities_grad(self.parent.pullback[f,:,:],-1)
        #C=np.zeros(G.shape[1])
        C=G[ge,:]
        P=np.reshape(np.arange(G.shape[1]), [1,G.shape[1]])
        T=time.time()
        
        
        if gs  > ge :
            #for t in range(1,G.shape[0]):#1,200):
            P,C=self.best_paths_rec(P, C, G, n, ge+1,gs,r1,r2,d) #for spezzato per escludere la guida 
            #print(P)
            bp=np.argmin(C[r1:r2])+r1
            print(f,'path',time.time()-T)
        
            #bp=np.argmax(C[r1:r2])+r1
        
            #print(bp,C[bp],P[:,bp])
            #self.parent.lumen[f,]=P[:,bp]
            self.parent.lumen[f,ge:gs]=P[:,bp]
        else:
            P,C=self.best_paths_rec(P, C, G, n, ge+1,G.shape[0],r1,r2,d)
            P,C=self.best_paths_rec(P, C, G, n,0, gs,r1,r2,d)
            #I=np.concatenate([I0[:,-a:],I0,I0[:,0:a]],axis=1)
            bp=np.argmin(C[r1:r2])+r1
            #bp=np.argmax(C[r1:r2])+r1
            #print('path',time.time()-T)
            #print(bp,C[bp],P[:,bp])
        
            self.parent.lumen[f,:gs]=P[G.shape[0]-(ge):,bp]
            self.parent.lumen[f,ge+1:]=P[:G.shape[0]-(ge+1),bp]
          
        self.interpolate_guide_arc(f, (gs-1 + self.parent.Nr)%self.parent.Nr, (ge+1)%self.parent.Nr)
        
        print('lumen: ',time.time()-T)
        #self.fc_boundary(f,self.fc_weights())
        
        self.add_shot()
        #self.build_intima()
        #self.update()
        
        '''
        line=np.zeros(G[:,r1:r2].shape)
        for i in range(len(self.parent.lumen[f])):
            line[i,int(self.parent.lumen[f,i])-r1]=1
        plt.imshow(np.asarray(np.maximum(  G[:,r1:r2], np.max(G[:,r1:r2])*line)))
        plt.show()
        '''
        '''for h in range(100,800,20):
            self.parent.lumen[f,]=P[:,h]
            self.add_shot()
            self.build_intima()
        '''
                
        #self.guide_boundary(916,self.parent.Nf)
        #for i in range(G.shape[0]):
        #    p=self.best_paths(P, C, G, n, t, f)
        return

    '''
    def best_paths_opt_c(self,P,C,G,n,t,r1,r2,d): 
        Q=np.zeros((P.shape[0]+1,P.shape[1]))
        Q[P.shape[0],:]=np.arange(P.shape[1])
        C2=np.ones(C.shape)*np.inf
        Cp=np.zeros( 2*n-1 )
        for j in range(r1 + n,r2-n):
            e1=j-n+1
            e2=j+n
            Cp= C[e1:e2] + np.multiply(G[t,j] + G[(t-1+self.parent.Nr)%self.parent.Nr,e1: e2],d)  
            bp=np.argmin(Cp) 
            C2[j]=Cp[bp]  #C[bp]+ (G[t,j] + G[t-1,bp])*(1+alpha*np.power(np.abs(bp-j),beta))
            Q[:-1,j]=P[:,bp+ e1]
        return Q,C2    
    
    def lumen_boundary_c(self,f,d):

        n=30
        r1=45
        r2=900
        pad=10
        v=100
        
        T=time.time()
        
        self.parent.lumen[f,:]=0
        #gs,ge=self.find_next_arc(0, self.parent.guide[f,] )
        gs=self.parent.guide_arcs[f][0][1]        
        ge=self.parent.guide_arcs[f][0][0]        
        I=np.copy(self.parent.pullback[f,:,:])#np.zeros(self.parent.pullback[f,:,:].shape)
        #I[self.parent.guide[f,:]==1,:]=0
        gs = (gs -pad + self.parent.Nr)%self.parent.Nr
        ge = (ge + pad)%self.parent.Nr
        print('frame, guide arc',f,gs,ge)
        
        if gs > ge :
            I[gs:,:]=0
            I[:ge,:]=0
        else:
            I[gs:ge,:]=0
        
        G=self.intensities_grad(I)#self.parent.pullback[f,:,:])
        #G[self.parent.guide[f,:]==1,:]=0.1
        print('grads: ',time.time()-T)
        #G=self.intensities_grad(self.parent.pullback[f,:,:],-1)
        #C=np.zeros(G.shape[1])
        C=G[ge,:]
        P=np.reshape(np.arange(G.shape[1]), [1,G.shape[1]])
        T=time.time()
        
        
        if gs  > ge :
            #r1=np.max([np.guide[f-1,:]])
            for t in range(ge+1,gs):#1,200):
                P,C=self.best_paths_opt_c(P, C, G, n, t,r1,r2,d) #for spezzato per escludere la guida 
            #print(P)
            bp=np.argmin(C[r1:r2])+r1
            print(f,'path',time.time()-T)
        
            #bp=np.argmax(C[r1:r2])+r1
        
            #print(bp,C[bp],P[:,bp])
            #self.parent.lumen[f,]=P[:,bp]
            self.parent.lumen[f,ge:gs]=P[:,bp]
        else:
            for t in range(ge+1,G.shape[0]):
                P,C=self.best_paths_opt_c(P, C, G, n, t,r1,r2,d)
            for t in range(ge+1,gs):
                P,C=self.best_paths_opt_c(P, C, G, n,t,r1,r2,d)
            #I=np.concatenate([I0[:,-a:],I0,I0[:,0:a]],axis=1)
            bp=np.argmin(C)#[r1:r2])+r1
            #bp=np.argmax(C[r1:r2])+r1
            #print('path',time.time()-T)
            #print(bp,C[bp],P[:,bp])
        
            print(bp,P.shape)
            self.parent.lumen[f,:gs]=P[G.shape[0]-ge:,bp]
            self.parent.lumen[f,ge+1:]=P[:G.shape[0]-(ge+1),bp]
          
        self.interpolate_guide_arc(f, (gs-1 + self.parent.Nr)%self.parent.Nr, (ge+1)%self.parent.Nr)
        
        print('lumen: ',time.time()-T)
        #self.fc_boundary(f,self.fc_weights())
        
        self.add_shot()
        self.build_intima()
        self.update()
        
        
        #for i in range(G.shape[0]):
        #    p=self.best_paths(P, C, G, n, t, f)
        return
'''


    def best_paths_rec_c(self,P,C,G,n,t,stop, vr1,vr2,R1,R2,d): 
        Q=np.zeros((P.shape[0]+1,P.shape[1]))
        Q[P.shape[0],:]=np.arange(P.shape[1])
        C2=np.ones(C.shape)*np.inf
        
        if t==0:
            #print(vr1[t-1],vr2[t-1])
            bp=np.argmin(C[vr1[t-1]:vr2[t-1]])+vr1[t-1]
        else:
            bp=0
            
        r1=np.max([ R1, np.min([bp -n, vr1[t] ]) ])
        r2=np.min([ R2, np.max([ bp +n, vr2[t] ]) ])
        vr1[t]=r1
        vr2[t]=r2
        G[t,:r1]=np.inf
        G[t,r2:]=np.inf
        Cp=np.zeros( 2*n-1 )
        #print(G[t])
        #print(r1,r2)
        for j in range(r1 ,r2):
            e1=j-n+1
            e2=j+n
            Cp= C[e1:e2] + np.multiply(G[t,j] + G[t-1,e1: e2],d)  
            bp=np.argmin(Cp) 
            C2[j]=Cp[bp]  #C[bp]+ (G[t,j] + G[t-1,bp])*(1+alpha*np.power(np.abs(bp-j),beta))
            Q[:-1,j]=P[:,bp+ e1]
        if t+1 < stop:#G.shape[0]:
            Q,C2=self.best_paths_rec_c(Q, C2, G, n, t+1,stop, vr1,vr2,R1,R2, d)
        
        return Q,C2    


    def lumen_boundary_c(self,f,d):
        n=30
        R1=75
        R2=870 
        pad=10
        
        T=time.time()
        
        self.parent.lumen[f,:]=0
        gs=self.parent.guide_arcs[f][0][1]        
        ge=self.parent.guide_arcs[f][0][0]        
        I=np.copy(self.parent.pullback[f,:,:])
        gs = (gs -pad + self.parent.Nr)%self.parent.Nr
        ge = (ge + pad)%self.parent.Nr
        print('frame, guide arc',f,gs,ge)
        
        if gs > ge :
            I[gs:,:]=0
            I[:ge,:]=0
        else:
            I[gs:ge,:]=0
        
        G=self.intensities_grad(I)
        print('grads: ',time.time()-T)
        C=G[ge,:]
        P=np.reshape(np.arange(G.shape[1]), [1,G.shape[1]])
        T=time.time()
        
        vr1=self.parent.lumen[f-1]-100#np.max([self.guide[f]-100,R1])
        vr2=self.parent.lumen[f-1]+100#np.min([self.guide[f]+100,R2])
        #print(vr1.shape,vr1[200:300])
        
        if ge==self.parent.Nr - 1:
            P,C=self.best_paths_rec_c(P, C, G, n, 0, gs, vr1, vr2, R1, R2, d)  
            bp=np.argmin(C[R1:R2])+R1
            print(f,'path',time.time()-T)

            self.parent.lumen[f,:gs]=P[1:,bp]
            self.parent.lumen[f,ge]=P[0,bp]

        elif gs==0:
            P,C=self.best_paths_rec_c(P, C, G, n, ge, self.parent.Nr, vr1, vr2, R1, R2, d)  
            bp=np.argmin(C[R1:R2])+R1
            print(f,'path',time.time()-T)

            self.parent.lumen[f,ge:]=P[:-1,bp]
            self.parent.lumen[f,gs]=P[-1,bp]

        
        elif gs  > ge :
            P,C=self.best_paths_rec_c(P, C, G, n, ge+1, gs, vr1, vr2, R1, R2, d)  
            bp=np.argmin(C[R1:R2])+R1
            print(f,'path',time.time()-T)

            self.parent.lumen[f,ge:gs]=P[:,bp]
        else:
            P,C=self.best_paths_rec_c(P, C, G, n, ge+1, G.shape[0], vr1, vr2, R1, R2, d)
            P,C=self.best_paths_rec_c(P, C, G, n, 0, gs, vr1, vr2, R1, R2, d)
            bp=np.argmin(C[R1:R2])+R1
        
            self.parent.lumen[f,:gs]=P[G.shape[0]-(ge):,bp]
            self.parent.lumen[f,ge+1:]=P[:G.shape[0]-(ge+1),bp]
          
        self.interpolate_guide_arc(f, (gs-1 + self.parent.Nr)%self.parent.Nr, (ge+1)%self.parent.Nr)
        
        print('lumen: ',time.time()-T)
        
        self.add_shot()
        self.build_intima()
        self.update()
      
        return
    
    def fc_weights(self):
        n=6
        alpha=0.01
        beta=1
        d=np.zeros(2*n-1)
        for j in range(0,2*n-1):
            d[j]=(1+alpha*np.power(np.abs(j-n+1),beta))
        return d
    
    def fc_boundary(self,f,d):
        
        F=self.index.get()
        n=6
        r1=12
        r2=140
        
        self.parent.fc[f,:]=0    
        fct=[] 
        
        self.plaque_arcs=self.parent.plaque_arcs[f]
        #print('frame',f)
        for arc in range(len(self.plaque_arcs)):
        
            
            ls,le=self.plaque_arcs[arc][0],self.plaque_arcs[arc][1]
            l=le-ls
            if le<ls:
                l= le + self.parent.Nr - ls
            
            I=self.fc_roi(f,ls,le,150)
            G=self.intensities_grad_fc(I,-1)#self.parent.pullback[f,:,:])
            C=G[0,:]
            P=np.reshape(np.arange(G.shape[1]), [1,G.shape[1]])
            P,C=self.best_paths_rec(P, C, G, n, 0,l,r1,r2,d) #for spezzato per escludere la guida 
            bp=np.argmin(C[r1:r2])+r1
                
            if ls<le:   
                for i in range(ls,le+1):
                    self.parent.fc[f,i]= P[i-ls,bp] + self.parent.lumen[f,i]
                fct.append([np.min(P[:,bp]),(np.argmin(P[:,bp])+ls)%self.parent.Nr,(np.argmin(P[:,bp])+ls)%self.parent.Nr])        
            else:
                for i in range(ls,self.parent.Nr):
                    self.parent.fc[f,i]= P[i-ls,bp] + self.parent.lumen[f,i]
                for i in range(0,le+1):    
                    self.parent.fc[f,i]= P[ i+ self.parent.Nr -ls ,bp ] + self.parent.lumen[f,i]
                fct.append([np.min(P[:,bp]),(np.argmin(P[:,bp])+ls)%self.parent.Nr,(np.argmin(P[:,bp])+ls)%self.parent.Nr])        
            line=np.zeros(G[:,r1:r2].shape)
            '''for i in range(G.shape[0]):
                #print(P[i,bp]-r1)
                line[i , int(P[i,bp]-r1) ]=1
            plt.imshow(np.asarray(np.maximum(  G[:,r1:r2], np.max(G[:,r1:r2])*line)))
            plt.show()
            '''
        self.calculate_fct_cart(f)
        #self.calculate_fct(f,fct)
        #print('min fct',fct , self.fct)
        
        #self.add_shot()
        self.build_fc()
        #self.update()
        
        return
    
    def calculate_fct(self,f,fct):
        if fct:
            self.fct=min(fct,key=lambda i : i[0])
        else :
            self.fct=[0,np.nan,np.nan]
        self.parent.fct[f]=np.copy(self.fct)   
        
    def calculate_fct_cart(self,f):
        n=10
        
        center_x = self.sRadius
        center_y = self.sRadius
        fc_points=[]
        lumen_points=[]
        self.plaque_arcs=self.parent.plaque_arcs[f]
        fct=[]
        for k in range(len(self.plaque_arcs)):
            fc_points=[]
            lumen_points=[]
            if self.plaque_arcs[k][0]<self.plaque_arcs[k][1] :  
                for i in range (self.plaque_arcs[k][0],self.plaque_arcs[k][1]):
                    r=self.parent.fc[f,i]/(self.zoom.get()+120)*self.sRadius
                    xf = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x
                    yf = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                    fc_points.append(xf)
                    fc_points.append(yf)
                    r=self.parent.lumen[f,i]/(self.zoom.get()+120)*self.sRadius
                    xl = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x
                    yl = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                    lumen_points.append(xl)
                    lumen_points.append(yl)

            else:
                for i in range (self.plaque_arcs[k][0],self.parent.Nr):
                    r=self.parent.fc[f,i]/(self.zoom.get()+120)*self.sRadius
                    xf = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x
                    yf = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                    fc_points.append(xf)
                    fc_points.append(yf)
                    r=self.parent.lumen[f,i]/(self.zoom.get()+120)*self.sRadius
                    xl = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x
                    yl = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                    lumen_points.append(xl)
                    lumen_points.append(yl)

                for i in range (self.plaque_arcs[k][1]):
                    r=self.parent.fc[f,i]/(self.zoom.get()+120)*self.sRadius
                    xf = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x
                    yf = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                    fc_points.append(xf)
                    fc_points.append(yf)
                    r=self.parent.lumen[f,i]/(self.zoom.get()+120)*self.sRadius
                    xl = r * np.cos(2*np.pi*i / self.parent.Nr) + center_x
                    yl = r * np.sin(2*np.pi*i / self.parent.Nr) + center_y
                    lumen_points.append(xl)
                    lumen_points.append(yl)

            m=1000
            x1=0
            x2=0
            y1=0
            y2=0
            ls=self.plaque_arcs[k][0]
            min_fct=49
            bfc=1000
            bl , bf = [ls,ls]
            for i in range(len(fc_points)//2):
                local_m=1000
                for j in range( np.max([0,i-n]), np.min([len(fc_points)//2,i+n]) ):
                    #l=np.linalg.norm([  fc_points[2*i]-lumen_points[2*j], fc_points[2*i+1]-lumen_points[2*j+1] ])
                    x1=fc_points[2*i]
                    y1=fc_points[2*i+1]
                    x2=lumen_points[2*j]
                    y2=lumen_points[2*j+1]
                    l=np.sqrt((x1-x2)*(x1-x2) +(y1-y2)*(y1-y2))*(self.zoom.get() + 120)/self.sRadius
                    check_ok=False

                    if l<local_m:
                        if self.check_fc(f, (j+ls)%self.parent.Nr, (i+ls)%self.parent.Nr, x2, y2, x1, y1) and l* self.parent.calibration_scale >min_fct: 
                            local_m=l
                            check_ok=True
                    if l<m and check_ok:
                        m=l
                        bfc=l
                        bf=i
                        bl=j
                self.parent.all_fct[f,(i+ls)%self.parent.Nr]=local_m
            fct.append([bfc,(bl+ls)%self.parent.Nr,(bf+ls)%self.parent.Nr]) 
            #print(x1,x2,y1,y2)
        self.calculate_fct(f, fct)
        #print(self.fct, fct,len(self.parent.plaque_arcs[f]))
    
    
    def check_fc(self, f, tl, tfc, xl,yl, xfc ,yfc):

        '''
        xl *= (self.zoom.get() + 120)/self.sRadius - self.sRadius
        yl *= (self.zoom.get() + 120)/self.sRadius - self.sRadius 
        xfc *= (self.zoom.get() + 120)/self.sRadius - self.sRadius
        yfc *= (self.zoom.get() + 120)/self.sRadius - self.sRadius
        m= (yl-yfc)/(xl-xfc)
        q= yl-m*xl
        '''
        
        #h=np.abs(tl-tfc)/10
        
        rl=self.parent.lumen[f,tl]
        rfc=self.parent.fc[f,tfc]
        
        thr_jump=20
        #print(np.abs(self.parent.lumen[f,(tl-3)%self.parent.Nr]-self.parent.lumen[f,(tl+3)%self.parent.Nr]) , thr_jump)
        j=np.abs(self.parent.lumen[f,(tl-3)%self.parent.Nr]-self.parent.lumen[f,(tl+3)%self.parent.Nr])
        
        if j > thr_jump:
            j1=np.abs(self.parent.lumen[f,(tl-20)%self.parent.Nr]-self.parent.lumen[f,(tl+20)%self.parent.Nr])
        
            if np.abs(j/6 -j1/40) > thr_jump/6 * 0.2 :
                #print('jump',j,j1,tl )
                return False
        s=0
        n=6
        #if tl==tfc:
        for r in np.linspace(rl,rfc,n):
            #t=tl
            #if not tl==tfc:
            for t in np.linspace(tl,tfc,n):
                #print(np.linspace(rl,rfc,10))
                s+=self.parent.pullback[ f, int(np.rint(t))%self.parent.Nr, np.min([int(np.rint(r)),self.parent.Np-1]) ]
                #print(self.parent.pullback[ f, int(np.rint(tl))%self.parent.Nr, np.min([int(np.rint(r)),self.parent.Np-1]) ])
        '''else :
            for t in np.linspace(tl+self.parent.Nr,tfc+self.parent.Nr,10):
                #print(np.linspace(tl+self.parent.Nr,tfc+self.parent.Nr,10))
                theta=t*2*np.pi/self.parent.Nr
                r=q/(np.sin(theta)-m*np.cos(theta))
                print(f, int(t)%self.parent.Nr, int(r),s )
                s+=self.parent.pullback[ f, int(t)%self.parent.Nr, int(r) ]
        '''
        
        n*=n
        #print(s/n)
        return (s/n > 150)
        
    def fc_roi(self,f,s,e,depth):
        #print('fc roi',s,e)
        d=e-s
        if e<s:
            d= e + self.parent.Nr - s
        intima_zone=np.zeros([d,depth])
        
        if s<e:
            for j in range(e-s) :
                max_depth=self.parent.Np - self.parent.lumen[f,j+s]
                
                if(max_depth > depth) :
                    intima_zone[j,:] = self.parent.pullback[f ,j+s, self.parent.lumen[f,j+s]: self.parent.lumen[f,j+s]+depth]
                else :
                    intima_zone[j , 0:max_depth-1]=self.parent.pullback[f , j+s, self.parent.lumen[f,j+s]:-1]
        elif s>e:
            for j in range(s,self.parent.Nr) :
                
                max_depth=self.parent.Np - self.parent.lumen[f,j]
                
                if(max_depth > depth) :
                    intima_zone[j-s,:] = self.parent.pullback[f ,j, self.parent.lumen[f,j]: self.parent.lumen[f,j]+depth]
                else :
                    intima_zone[j-s , 0:max_depth-1]=self.parent.pullback[f ,j, self.parent.lumen[f,j]:-1]
        
            for j in range( e ) :
                
                max_depth=self.parent.Np - self.parent.lumen[f,j]
                
                if(max_depth > depth) :
                    intima_zone[j+ self.parent.Nr - s,:] = self.parent.pullback[f , j, self.parent.lumen[f,j]: self.parent.lumen[f,j]+depth]
                else :
                    intima_zone[j+ self.parent.Nr - s , 0:max_depth-1]=self.parent.pullback[f , j, self.parent.lumen[f,j]:-1]
        #else:
        #    return np.zeros([1,depth])
        return intima_zone
            

            
    def guide_boundary(self, s,e):
        
        f=self.index.get()
        n=15        
        a=100
        nb=20
        r1=0
        r2=self.parent.Nr+2*(a-nb)
        alpha=0.05
        beta=1
        
        d=np.zeros(2*n-1)
        for j in range(0,2*n-1):
            d[j]=(1+alpha*np.power(np.abs(j-n+1),beta))
        
        I0= np.sum(self.parent.pullback[s:e,:,200:],axis=2)
        I=np.concatenate([I0[:,-a:],I0,I0[:,0:a]],axis=1)
    
        G=self.intensities_grad(I)
        #G=self.intensities_grad(I,-1)
        
        G=G[:,nb:-nb]
        
        P1=np.reshape(np.arange(G.shape[1]), [1,G.shape[1]])
        C1=G[0,:]
        #C1=np.zeros(G.shape[1])
 
        for t in range(1,G.shape[0]):#1,200):
            P1,C1=self.best_paths_opt(P1, C1, G, n, t,r1,r2,d)
        bp1=np.argmin(C1)#[r1:r2])+r1
        #print(C1)
        #bp1=np.argmax(C1[r1:r2])+r1
        #print(bp1,C1[bp1],P1[:,bp1])
        
        G=self.intensities_grad(I,-1)
        #G=self.intensities_grad(I,1)
        G=G[:,nb:-nb]

        P2=np.reshape(np.arange(G.shape[1]), [1,G.shape[1]])
        C2=G[0,:]
        
        for t in range(1,G.shape[0]):#1,200):
            P2,C2=self.best_paths_opt(P2, C2, G, n, t,r1,r2,d)
        bp2=np.argmin(C2)#[r1:r2])+r1
        #bp2=np.argmax(C2[r1:r2])+r1
        #print(bp2,C2[bp2],P2[:,bp2])
        #start=np.min([P1[0,bp1],bp1])
        
        lines=np.zeros([G.shape[0],self.parent.guide.shape[1]])
        self.parent.guide[s:e,:]=0
        
        scale= G.shape[1] -a + nb
        aa=self.parent.guide.shape[1]
        for i in range(s,e):
            if ( P1[i-s,bp1] +aa - a + nb )%aa > ( P2[i-s,bp2] +aa - a + nb )%aa :
                self.parent.guide[i,int((P2[i-s,bp1] +aa - a + nb )%aa):int((P1[i-s,bp2] +aa - a + nb )%aa)]=1
            else:
                self.parent.guide[i,int((P2[i-s,bp1] +aa - a + nb )%aa):]=1
                self.parent.guide[i,:int((P1[i-s,bp2] +aa - a + nb )%aa)]=1
            lines[ i-s,int( (P1[i-s,bp1] +aa - a + nb )%aa ) ]=1
            lines[ i-s,int( (P2[i-s,bp2] +aa - a + nb )%aa ) ]=1
        self.parent.labels[self.parent.guide==1]=unknown_color_index
        
        #np.save(model_file + "/sum.npy",lines)
        #np.save(model_file + "/lines.npy",lines)
        '''plt.imshow(np.asarray(np.maximum(  G[:,a-nb:self.parent.guide.shape[1]+a-nb],np.max(G)*lines)))
        plt.show()
        
        plt.imshow(np.asarray(np.maximum(  I0,np.max(I0)*lines)))
        plt.show()
        '''
        
        return
    '''    
    def best_paths_double(self,Q,P,C,Gx,Gy,n,t,r1,r2,m,M,d): 
        C2=np.ones(C.shape)*np.inf
        for x in range(r1 + n,r2-n-1):
            e1=x-n+1
            Cx=np.multiply(Gx[t,x] + Gx[t-1,e1: x+n],d)
            #print(x,r2-n)
            for y in range(x + m + n, x + M -n):
                Cp=np.zeros( [2*n-1, 2*n-1])
                e2=y-n+1
                Cy=np.multiply(Gy[t,y] + Gy[t-1,e2: y+n],d)
                Cp= C[e1: x+n,e2: y+n] + Cx[np.newaxis,:] + Cy[:,np.newaxis]  
                bp=np.unravel_index(np.argmin(Cp, axis=None), Cp.shape)
                C2[x,y]=Cp[bp]  #C[bp]+ (G[t,j] + G[t-1,bp])*(1+alpha*np.power(np.abs(bp-j),beta))
                #print(bp,e1,e2)
                #print(x, y, Cp[bp], P.shape, type(P[bp[0]+ e1,bp[1]+ e2]), bp[0]+ e1, bp[1]+ e2, P[bp[0]+ e1,bp[1]+e2])#,type(P[bp[0]+ e1,bp[1]+ e2][0]))
                #Q[bp[0]+ e1,bp[1]+ e2]=np.copy(P[bp[0]+ e1,bp[1]+ e2].insert(0,(x,y)))
                Q[x,y]=P[bp[0]+ e1-1,bp[1]+ e2-1]
                Q[x,y].insert(0,(x,y))
                
                #print(Q[x,y])
        return Q,C2    

    def guide_boundary_double(self, s,e):
        
        n=15        
        a=140
        nb=20
        M=80
        m=20
        r1=0
        r2=self.parent.Nr+2*(a-nb)-M
        alpha=0.05
        beta=1
        
        d=np.zeros(2*n-1)
        for j in range(0,2*n-1):
            d[j]=(1+alpha*np.power(np.abs(j-n+1),beta))
        
        I0= np.sum(self.parent.pullback[s:e,:,200:],axis=2)
        I=np.concatenate([I0[:,-a:],I0,I0[:,0:a]],axis=1)
    
        Gx=self.intensities_grad(I,-1)
        Gx=Gx[:,nb:-nb]

        Gy=self.intensities_grad(I)
        Gy=Gy[:,nb:-nb]
        
        
        Q=np.empty((r2-r1 , r2+M ),dtype=object)
        
        for i in range(Q.shape[0]):
            for j in range(Q.shape[1]):
                Q[i,j]=[(i,j)]
        
        C=np.ones([r2-r1 , r2+M])*np.inf
        for i in range(r2-r1):
            for j in range(i+m, i+M ):
                #print(C[i,j])
                #print(i,Gx[0,i])
                #print(r2+M,j,Gy[0,j])
                C[i,j]=Gx[0,i]+ Gy[0,j]
        
        P=np.copy(Q)
        print(Gx.shape,Q.shape)
        
        for t in range(1,Gx.shape[0]):#1,200):
            print('frame',t)
            P,C=self.best_paths_double(Q,P, C, Gx,Gy, n, t,r1,r2,m,M,d)
            print(P[200,260])
        bp = np.unravel_index(np.argmin(C, axis=None), P.shape)

        
        lines=np.zeros([Gx.shape[0],self.parent.guide.shape[1]])
        self.parent.guide[s:e,:]=0
        
        print(bp)
        print(C[bp])
        #print(P[bp])
        #print(type(P),type(P[bp]),len(P[bp]),P[bp][0],P[bp])
        
        ns_lines=np.zeros(I.shape)
        scale= Gx.shape[1] -a + nb
        aa=self.parent.guide.shape[1]
        for i in range(s,e):
            if ( P[bp][i-s][1] +aa - a + nb )%aa > ( P[bp][i-s][0] +aa - a + nb )%aa :
                self.parent.guide[i,int((P[bp][i-s][0] +aa - a + nb )%aa):int((P[bp][i-s][1] +aa - a + nb )%aa)]=1
            else:
                self.parent.guide[i,int((P[bp][i-s][0] +aa - a + nb )%aa):]=1
                self.parent.guide[i,:int((P[bp][i-s][1] +aa - a + nb )%aa)]=1
            lines[ i-s,int( (P[bp][i-s][1] +aa - a + nb )%aa ) ]=1
            lines[ i-s,int( (P[bp][i-s][0] +aa - a + nb )%aa ) ]=1
            ns_lines[i,P[bp][i-s][0]+r1]=1
            ns_lines[i,P[bp][i-s][1]]=1
        self.parent.labels[self.parent.guide==1]=unknown_color_index
        self.parent.labels[self.parent.guide==0]=non_lipid_color_index
        
        #np.save(model_file + "/sum.npy",lines)
        #np.save(model_file + "/lines.npy",lines)
        #plt.imshow(np.asarray(np.maximum(  G[:,a-nb:self.parent.guide.shape[1]+a-nb],np.max(G)*lines)))
        #plt.show()
        
        #plt.imshow(np.asarray(np.maximum(  I0,np.max(I0)*lines)))
        #plt.show()
        
        
        return I,ns_lines#I0,lines
'''
    
    
    def best_paths_double(self,P,C,Gx,Gy,n,t,r1,r2,m,M,d): 
        C2=np.ones(C.shape)*np.inf
        Q=np.zeros((P.shape[0],P.shape[1] ,P.shape[2]+1,2))
        
        for x in range(r1 + n,r2-n-1):
            e1=x-n+1
            Cx=np.multiply(Gx[t,x] + Gx[t-1,e1: x+n],d)
            #print(x,r2-n)
            for y in range(x + m + n, x + M -n):
                Cp=np.zeros( [2*n-1, 2*n-1])
                e2=y-n+1
                Cy=np.multiply(Gy[t,y] + Gy[t-1,e2: y+n],d)
                Cp= C[e1: x+n,e2: y+n] + Cx[np.newaxis,:] + Cy[:,np.newaxis]  
                bp=np.unravel_index(np.argmin(Cp, axis=None), Cp.shape)
                C2[x,y]=Cp[bp]  #C[bp]+ (G[t,j] + G[t-1,bp])*(1+alpha*np.power(np.abs(bp-j),beta))
                #print(bp,e1,e2)
                #print(x, y, Cp[bp], P.shape, type(P[bp[0]+ e1,bp[1]+ e2]), bp[0]+ e1, bp[1]+ e2, P[bp[0]+ e1,bp[1]+e2])#,type(P[bp[0]+ e1,bp[1]+ e2][0]))
                #Q[bp[0]+ e1,bp[1]+ e2]=np.copy(P[bp[0]+ e1,bp[1]+ e2].insert(0,(x,y)))
                Q[x,y,:-1,:]=P[bp[0]+ e1,bp[1]+ e2,:,:]
                Q[x,y,P.shape[2],0]=x
                Q[x,y,P.shape[2],1]=y
        
                #print(Q[x,y])
        return Q,C2    

    def gradients_variance(self,sense=-1):
        l_min=90
        d=2.5
        s=4
        h=15
        d=0.3
        l_min=80
        I_thr=15
        T = time.time()
        #G=self.pullback
        I0=self.parent.pullback
        I=np.concatenate([I0[:,:,-h:],I0,I0[:,:,0:h]],axis=2)
        I[I<I_thr]=0
        
        G=np.zeros(self.parent.pullback.shape)

        K=np.arange(2*h+1)-h
        K=-1/(np.sqrt(2*np.pi)*s*s*s) * np.multiply(K , np.exp(-1/2*np.multiply(K,K)/(s*s)))
        for t in range(G.shape[0]):
            for z in range(G.shape[1]):
                G[t,z,:]=np.convolve(I[t,z,:],K,'valid')

        G=-G
        G[G<d]=0
        Gsum= np.sum(G[:,:,l_min:],axis=2)
        
        P = np.arange(0, self.parent.Np)
        P[:l_min] = 0
        E = np.dot(G,P)
        L = np.divide(E,Gsum)
        mask = np.isnan(L)
        L[mask] = np.interp(np.flatnonzero(mask), np.flatnonzero(~mask), L[~mask])

        P = np.multiply(P,P)
        V = np.dot(G,P)
        V = np.divide(V,Gsum)
        V = V - np.multiply(L,L)

        mask = np.isnan(V)
        V[mask] = np.interp(np.flatnonzero(mask), np.flatnonzero(~mask), V[~mask])

        #print('non positivi prima di log' ,sum(sum(V<=0)))
        V=np.log2(V)
        
        mask = np.isnan(V) ^ np.isinf(V)
        V[mask] = np.interp(np.flatnonzero(mask), np.flatnonzero(~mask), V[~mask])
        
        #print('prima di mean',sum(sum(np.isnan(V))) ,V.size , np.mean(V))
        
        V=V-np.mean(V)
        V=V/np.std(V)
        mask = np.isnan(V)
        V[mask] = np.interp(np.flatnonzero(mask), np.flatnonzero(~mask), V[~mask])
        
        V[V<-1]=-1
        V[V>1]=1
        V=(V+1)/2
        #if np.any(np.isnan(V)):
        #    print( 'brutte cose')
        return V

    def guide_boundary_double(self, s,e):
        
        n=15        
        a=140
        nb=20
        M=80
        m=15
        r1=0
        r2=self.parent.Nr+2*(a-nb)-M
        alpha=0.05
        beta=1
        
        d=np.zeros(2*n-1)
        for j in range(0,2*n-1):
            d[j]=(1+alpha*np.power(np.abs(j-n+1),beta))
        
        I0= np.sum(self.parent.pullback[s:e,:,230:],axis=2)
        I=np.concatenate([I0[:,-a:],I0,I0[:,0:a]],axis=1)
        #print('',I.shape)
        
        fg=np.concatenate([self.parent.guide_fix[0,:,-a:],self.parent.guide_fix[0,:,:],self.parent.guide_fix[0,:,0:a]],axis=1)
        fg=fg[:,nb:-nb]
        #print('',I.shape)
        
        G0=self.gradients_variance()
        #print('V shape',G0.shape)
        
        G=np.concatenate([G0[:,-a:],G0,G0[:,0:a]],axis=1)
        
        #print('V shape',G.shape)
        #Gx=self.intensities_grad(I,-1)
        Gx=self.intensities_grad(G,-1)
        Gx=Gx[:,nb:-nb]
        Gx[fg==1]=np.inf
        
        
        fg=np.concatenate([self.parent.guide_fix[1,:,-a:],self.parent.guide_fix[1,:,:],self.parent.guide_fix[1,:,0:a]],axis=1)
        fg=fg[:,nb:-nb]
        
        #Gy=self.intensities_graqd(I)
        Gy=self.intensities_grad(G)
        Gy=Gy[:,nb:-nb]
        Gy[fg==1]=np.inf
        
        
        P=np.zeros([r2-r1 , r2+M ,1,2])
        for i in range(P.shape[0]):
            for j in range(P.shape[1]):
                P[i,j,0,0]=i
                P[i,j,0,1]=j
        
        
        C=np.ones([r2-r1 , r2+M])*np.inf
        for i in range(r2-r1):
            for j in range(i+m, i+M ):
                #print(C[i,j])
                #print(i,Gx[0,i])
                #print(r2+M,j,Gy[0,j])
                C[i,j]=Gx[0,i] + Gy[0,j]
        
        #print(Gx.shape,P.shape)
        
        for t in range(1,Gx.shape[0]):#1,200):
            #print('checking guide: frame',t)
            P,C=self.best_paths_double(P, C, Gx,Gy, n, t,r1,r2,m,M,d)
            #print(P[200,260])
        bp = np.unravel_index(np.argmin(C, axis=None), C.shape)

        
        lines=np.zeros([Gx.shape[0],self.parent.guide.shape[1]])
        self.parent.labels[self.parent.guide==1]=non_lipid_color_index
        self.parent.guide[s:e,:]=0
        
        #print(bp)
        #print(C[bp])
        #print(P[bp])
        #print(type(P),type(P[bp]),len(P[bp]),P[bp][0],P[bp])
        
        ns_lines=np.zeros(I.shape)
        scale= Gx.shape[1] -a + nb
        aa=self.parent.guide.shape[1]
        for i in range(s,e):
            if ( P[bp[0],bp[1],i-s,1] +aa - a + nb )%aa > ( P[bp[0],bp[1],i-s,0] +aa - a + nb )%aa :
                self.parent.guide[i,int((P[bp[0],bp[1],i-s,0] +aa - a + nb )%aa):int((P[bp[0],bp[1],i-s,1] +aa - a + nb )%aa)]=1
            else:
                self.parent.guide[i,int((P[bp[0],bp[1],i-s,0] +aa - a + nb )%aa):]=1
                self.parent.guide[i,:int((P[bp[0],bp[1],i-s,1] +aa - a + nb )%aa)]=1
            self.parent.guide_arcs[i].append((int( (P[bp[0],bp[1],i-s,1] +aa - a + nb )%aa),int( (P[bp[0],bp[1],i-s,0] +aa - a + nb )%aa)))
            lines[ i-s,int( (P[bp[0],bp[1],i-s,1] +aa - a + nb )%aa ) ]=1
            lines[ i-s,int( (P[bp[0],bp[1],i-s,0] +aa - a + nb )%aa ) ]=1
            #ns_lines[i,P[bp[0],bp[1],i-s,0]+r1]=1
            #ns_lines[i,P[bp[0],bp[1],i-s,1]]=1
        self.parent.labels[self.parent.guide==1]=unknown_color_index
        
        #np.save(model_file + "/sum.npy",lines)
        #np.save(model_file + "/lines.npy",lines)
        '''plt.imshow(np.asarray(np.maximum(  G[:,a-nb:self.parent.guide.shape[1]+a-nb],np.max(G)*lines)))
        plt.show()
        
        plt.imshow(np.asarray(np.maximum(  I0,np.max(I0)*lines)))
        plt.show()
        '''
        
        self.parent.guide_lines=lines
        return I0,lines



    def choise_fixed_points(self,I,G):
        
        hi=20
        Ni=self.parent.Nr//hi
        interv=np.arange(0,self.parent.Nr,hi)
        l_mean=10
        dt=2
        dr=8
        s=10
        
        ordered_points=self.fixed_points
        ordered_points=sorted(ordered_points, key=lambda i: i[0])
        maxi=np.ones(Ni+1,dtype=np.int)*10000
        indi=np.ones(Ni+1,dtype=np.int)*10000
        for i in range(len(ordered_points)):
            if  maxi[ordered_points[i][0]//hi]>G[ordered_points[i][0],int(ordered_points[i][1])]  :
                maxi[ordered_points[i][0]//hi]=G[ordered_points[i][0],int(ordered_points[i][1])]
                indi[ordered_points[i][0]//hi]=i
        indilist=list(indi)
        deleted=0
        i=0
        
        while i<len(indilist):
            if indilist[i]<10000:
                t=ordered_points[indilist[i]][0]
                r=ordered_points[indilist[i]][1]
                if np.mean(I[int(t-dt)%self.parent.Nr:int(t+dt)%self.parent.Nr,int(r-dr)-s:int(r)-s])>30:
                    del indilist[i]
                    deleted+=1
                elif np.mean(I[int(t-dt)%self.parent.Nr:int(t+dt)%self.parent.Nr,int(r)+s:int(r+dr)+s])<40:
                    del indilist[i]
                    deleted+=11
                else:
                    i+=1
            else:
                i+=1
        self.fixed_points = [ordered_points[i] for i in indilist if i<10000]
        '''
        maxi=[]
        indi=[]
        for i in range(len(ordered_points)):
            
            l_mean=hi*np.mean( np.asarray([ ordered_points[i+j][1] for j in range(-l_mean,l_mean) if np.abs(ordered_points[j][0]-ordered_points[i][0]) < hi]))
            maxi.append(G[ordered_points[i][0],int(ordered_points[i][1])])
            indi.append(i)
        self.fixed_points= [ordered_points[i] for i in list(indi)]
        '''
        #print(indi)
        #print('final',len(self.fixed_points), 340%350)
    
    
    def choise_fixed_points_interv(self,I,G):
        self.parent.wait_variable(self.parent.v_stop)
        h_max=30
        h_min=5
        dh=4
        h0=0
        h1=10
        
        dt=2
        dr=8
        s=10
        
        ordered_points=self.fixed_points
        ordered_points=sorted(ordered_points, key=lambda i: i[0])
        indi=[]
        interv=[]
        
        i=0
        idx=-1
        m=2
        while i<len(ordered_points):
            if ordered_points[i][0]<h0:
                i+=1
            elif ordered_points[i][0]>h1:
                if not idx==h0-1:
                    indi.append(idx)
                    interv.append((h0,h1))
                    d=int((h_max-h_min)/(855)*(950 -ordered_points[idx][1]))
                    h0=ordered_points[idx][0] + int(d/2)
                    h1=h0 + d
                    idx=h0-1
                    m=2
                    #print(h0,h1)
                else:
                    h1=ordered_points[i][0]
            else:
                if  m>G[ordered_points[i][0],int(ordered_points[i][1])]  :
                    m=G[ordered_points[i][0],int(ordered_points[i][1])]
                    idx=i
                i+=1

        deleted=0
        i=0
        
        #print(len(indi),interv)
        self.fixed_points = [ordered_points[i] for i in indi]
        self.refresh_points()
        self.parent.wait_variable(self.parent.v_stop)

        while i<len(indi):
            t=ordered_points[indi[i]][0]
            r=ordered_points[indi[i]][1]
            if np.mean(I[int(t-dt)%self.parent.Nr:int(t+dt)%self.parent.Nr,int(r-dr)-s:int(r)-s])>30:
                del indi[i]
                deleted+=1
            elif np.mean(I[int(t-dt)%self.parent.Nr:int(t+dt)%self.parent.Nr,int(r)+s:int(r+dr)+s])<40:
                del indi[i]
                deleted+=1
            else:
                i+=1
        
        self.fixed_points = [ordered_points[i] for i in indi]
        self.refresh_points()
        self.parent.wait_variable(self.parent.v_stop)



    def choise_fixed_points_interv2(self,I,G):
        #self.parent.wait_variable(self.parent.v_stop)
        h_max=30
        h_min=5
        dh=4
        h0=0
        h1=10
        
        #print('Intensities shape', I.shape)
        dt=2
        dr=8
        s=10
        dr_bound=2
        
        ordered_points=self.fixed_points
        ordered_points=sorted(ordered_points, key=lambda i: i[0])
        indi=[]
        interv=[]
        
        i=0
        idx=-1
        m=2
        while i<len(ordered_points):
            if ordered_points[i][0]<h0:
                i+=1
            elif ordered_points[i][0]>h1:
                if not idx==h0-1:
                    indi.append(idx)
                    interv.append((h0,h1))
                    d=int(h_min+(h_max-h_min)*np.power((950 -ordered_points[idx][1])/855,2) )
                    h0=ordered_points[idx][0] + int(d/2)
                    h1=h0 + d
                    idx=h0-1
                    m=2
                    #print(h0,h1)
                else:
                    h1=ordered_points[i][0]
            else:
                if  m>G[ordered_points[i][0],int(ordered_points[i][1])]  :
                    t=int(ordered_points[i][0])
                    r=int(ordered_points[i][1])
                    if np.mean(I[(t-dt)%self.parent.Nr:(t+dt)%self.parent.Nr,(r-dr)-s:(r)-s])>30 or np.mean(I[(t-dt)%self.parent.Nr:(t+dt)%self.parent.Nr,(r)+s:int(r+dr)+s])<40:
                        del ordered_points[i]
                    else:
                        m=G[t,r]
                        idx=i
                        i+=1
                        if indi:
                            if np.abs((r-ordered_points[indi[-1]][1])/(t-ordered_points[indi[-1]][0]))>dr_bound:
                                h1=t
                else:
                    i+=1

        deleted=0
        i=0
        
        #print(len(indi),interv)
        self.fixed_points = [ordered_points[i] for i in indi]
        #self.refresh_points()
        #self.parent.wait_variable(self.parent.v_stop)
        '''
        while i<len(indi):
            t=ordered_points[indi[i]][0]
            r=ordered_points[indi[i]][1]
            if np.mean(I[int(t-dt)%self.parent.Nr:int(t+dt)%self.parent.Nr,int(r-dr)-s:int(r)-s])>30:
                del indi[i]
                deleted+=1
            elif np.mean(I[int(t-dt)%self.parent.Nr:int(t+dt)%self.parent.Nr,int(r)+s:int(r+dr)+s])<40:
                del indi[i]
                deleted+=1
            else:
                i+=1
        
        self.fixed_points = [ordered_points[i] for i in indi]
        self.refresh_points()
        self.parent.wait_variable(self.parent.v_stop)
        '''


    def validate_interpolation(self, f,lumen_splines):

        if len(self.fixed_points) < 5 :
            return

        t=np.asarray([i[0] for i in self.fixed_points])#,dtype=np.int)
        der=lumen_splines.derivative(1)
        d=der(t)

        margin=4
        bound=-2
        x = []
        y = []

        if der(t[0]+margin)*der(t[1]-margin)<bound:
            #print('jjcale')
            m=np.mean([self.fixed_points[len(self.fixed_points)-1][1],self.fixed_points[0][1]])
            if  np.abs(m-self.fixed_points[0][1])/m>0.2 :
                del self.fixed_points[0]
                self.parent.lumen[f,:],lumen_splines=self.interpolate_intima(show=False)
                self.validate_interpolation(f,lumen_splines)
                return
            x.append(self.fixed_points[0][0])
            y.append(self.fixed_points[0][1])
            x.append(self.fixed_points[1][0])
            y.append(self.fixed_points[1][1])
            X=np.arange(t[0],t[1])
            self.parent.lumen[f,t[0]:t[1]]=np.interp(X,x,y)
            x = []
            y = []
        
        
        for i in range(1,len(self.fixed_points)-1):
            if der(t[i]+margin)*der(t[i+1]-margin)<bound:
                #print('jjcale')
                m=np.mean([self.fixed_points[i-1][1],self.fixed_points[i+1][1]])
                if  np.abs(m-self.fixed_points[i][1])/m>0.2 and (self.fixed_points[i][1] - self.fixed_points[i-1][1])*(self.fixed_points[i][1] - self.fixed_points[i+1][1])>0:
                    del self.fixed_points[i]

                    self.parent.lumen[f,:],lumen_splines=self.interpolate_intima(show=False)
                    #self.refresh_points()
                    #self.build_intima()
                    #self.parent.wait_variable(self.parent.v_stop)
                    self.validate_interpolation(f,lumen_splines)
                    return
                x.append(self.fixed_points[i][0])
                y.append(self.fixed_points[i][1])
                x.append(self.fixed_points[i+1][0])
                y.append(self.fixed_points[i+1][1])
                X=np.arange(t[i],t[i+1])
                self.parent.lumen[f,t[i]:t[i+1]]=np.interp(X,x,y)
                x = []
                y = []
        i=len(self.fixed_points)-1
        if der((t[i]+margin)%self.parent.Nr)*der( (t[0]-margin)%self.parent.Nr)<bound:
            #print('jjcale')
            m=np.mean([self.fixed_points[i-1][1],self.fixed_points[0][1]])
            if  np.abs(m-self.fixed_points[i][1])/m>0.2 :
                del self.fixed_points[i]
                self.parent.lumen[f,:],lumen_splines=self.interpolate_intima(show=False)
                self.validate_interpolation(f,lumen_splines)
                return
            x.append(self.fixed_points[i][0])
            y.append(self.fixed_points[i][1])
            x.append(self.fixed_points[0][0] + self.parent.Nr)
            y.append(self.fixed_points[0][1])
            X=np.arange(t[i],t[0]+self.parent.Nr)
            self.parent.lumen[f,t[i]:]=np.interp(X,x,y)[:self.parent.Nr-t[i]]
            self.parent.lumen[f, :t[0]] = np.interp(X, x, y)[self.parent.Nr-t[i]:]
            x = []
            y = []
        return

    def lumen_ds(self,f,single_frame=False):
        
        it=3500
        center_x = self.sRadius
        center_y = self.sRadius
        pad=0
        min_r=75
        r0=np.random.randint(100,200)
        t0=np.random.randint(0,self.parent.Nr)
        thr=50
        v=1
        alpha=50
        
        gs=self.parent.guide_arcs[f][0][1]        
        ge=self.parent.guide_arcs[f][0][0]        
        #print('guida' , gs , ge)
        gs = (gs -pad + self.parent.Nr)%self.parent.Nr
        ge = (ge + pad)%self.parent.Nr
        I=np.copy(self.parent.pullback[f,:,:])
        if gs > ge :
            I[gs:,:]=0
            I[:ge,:]=0
        else:
            I[gs:ge,:]=0
        G=self.intensities_grad(I)
        ds_points=[]
        self.delete("points")
        self.delete("ds")
        self.delete("return_points")
        first_light=0
        first_light_r=100
        x=0
        y=0
        
        for k in range(2):
            r0=np.random.randint(100,200)
            t0=np.random.randint(0,self.parent.Nr)
            i=t0
            if k==0:
                spin=-1
            else:
                spin=1
            while i < it+t0:
                if r0>950:
                    i=first_light + 10
                    #print(r0,first_light_r)
                    if first_light_r>200:
                        r0=first_light_r - 100
                    else :
                        r0=np.random.randint(100,200)
                    first_light=i
                    first_light_r=r0
                t=(spin*i)%self.parent.Nr
                if I[t,int(r0)]>thr:
                    v=-1
                    #first_light=i
                    #first_light_r=r0
                else :
                    v=1
                r1=r0 + v*G[t,int(r0)]*G[t,int(r0)]*alpha
                #print('r1',r1)
                if(r1<min_r):
                    r1=min_r
                r=r1/(self.zoom.get()+120)*self.sRadius        
                x = r * np.cos(2*np.pi*t / self.parent.Nr) + center_x
                y = r * np.sin(2*np.pi*t / self.parent.Nr) + center_y
                ds_points.append(x)
                ds_points.append(y)
        
                if I[t,int(r0)]>60:
                    first_light=i
                    first_light_r=r0
                if(G[t,int(r0)]<0.55 and I[t,int(r0)]> 70):
                    self.fixed_points.append((t,r0))
                    #if r0==first_light_r:
                    #self.create_oval(x - 3, y - 3, x + 3 , y + 3 , tags=("return_points"), fill='green')
                    #print()
                
                r0=r1
                i+=1

            #self.create_line(ds_points, fill='blue', width=2, tags=("ds"))
            #self.parent.wait_variable(self.parent.v_stop)
            #self.delete("ds")

        #print('len',len(self.fixed_points))
        #self.refresh_points()
        #self.fit_lumen(f)
        
        self.choise_fixed_points_interv2(I,G)
        if  len(self.fixed_points)>4:
            self.parent.lumen[f,:],lumen_splines=self.interpolate_intima(show=False)
        else :
            lumen_splines=0
        self.validate_interpolation(f,lumen_splines)
        
        if single_frame:
            self.refresh_points()
            
            #self.interpolate_guide_arc(f, (gs-1 + self.parent.Nr)%self.parent.Nr, (ge+1)%self.parent.Nr)
            #self.delete("ds")
            #self.delete("return_points")
            #self.refresh_points()
            self.build_intima()
            
        if np.any([self.parent.lumen[f]<40]) or len(self.fixed_points)<5 or np.any([self.parent.lumen[f]>self.parent.Np-1]):
            self.fixed_points=[]
            d=self.lumen_weights()
            self.boundary(f, d)
        if  single_frame:
            self.add_shot()
        else :
            self.fixed_points=[]
        return


    def fit_lumen(self, f):
        ordered_points=sorted(self.fixed_points, key=lambda i: i[0])
        x=np.zeros(len(ordered_points))
        y=np.zeros(len(ordered_points))
        for i in range(len(ordered_points)):
            x[i]=ordered_points[i][0]
            y[i]=ordered_points[i][1]
        lumenp=np.polyfit(x, y, 15)
        self.parent.lumen[f,:]=np.polyval(lumenp, np.arange(self.parent.Nr))
        
        
    def lumen_ds_1(self,f):
        it=3000
        center_x = self.sRadius
        center_y = self.sRadius
        pad=0
        min_r=95
        r0=150
        thr=50
        v=1
        alpha=8

        gs=self.parent.guide_arcs[f][0][1]        
        ge=self.parent.guide_arcs[f][0][0]        
        I=np.copy(self.parent.pullback[f,:,:])
        gs = (gs -pad + self.parent.Nr)%self.parent.Nr
        ge = (ge + pad)%self.parent.Nr
        if gs > ge :
            I[gs:,:]=0
            I[:ge,:]=0
        else:
            I[gs:ge,:]=0
        G=self.intensities_grad(I)
        ds_points=[]
        self.delete("ds")
        G_gs=0
        first_light=ge
        first_light_r=0
        last_light=ge
        last_light_r=0
        
        for i in range(ge,it):
            if r0>900:
                break
                #r0=np.random.randint(100,500)
                #t=t-40
                #i=last_light 
                #r0=last_light_r
            t=i%self.parent.Nr
            #if(t==gs):
            #    G_dg=v*G[t,int(r0)]
            if (gs<ge and gs<t and t<ge) or (gs>ge and ((gs<t and t<self.parent.Nr)or (0<t and t<ge))):
                v=G_dg/G[t,int(r0)] /alpha*5
            else:
                if I[t,int(r0)]>thr:
                    v=-1
                    if(first_light==ge):
                        first_light=i
                        first_light_r=r0
                else :
                    v=1
                if G[t,int(r0)]<0.3:
                    last_light=i
                    last_light_r=r0
            if(t==gs-1):
                G_dg=-(r0-first_light_r)/(i-first_light)

            r1=r0 + v*G[t,int(r0)]*alpha
            if(r1<min_r):
                r1=min_r
            r0=r1
            r=r1/(self.zoom.get()+120)*self.sRadius        
            x = r * np.cos(2*np.pi*t / self.parent.Nr) + center_x
            y = r * np.sin(2*np.pi*t / self.parent.Nr) + center_y
            ds_points.append(x)
            ds_points.append(y)
    
            if(i%self.parent.Nr==ge):
                first_light=ge
                first_light_r=0
            
        self.create_line(ds_points, fill='blue', width=2, tags=("ds"))
            

   
        return

    
    def lumen_detect (self,s,e):

        I=np.zeros([e-s,self.parent.Nr])
        G=np.zeros([e-s,self.parent.Nr])
        self.fixed_points=[]
        h=15
        p=int(self.parent.Nr/h)
    
        l=5
        m=10
    
        I=np.zeros([e-s,self.parent.Nr],dtype=int)
        for i in range(s,e):
            for j in range(1,p-1):
                #c=0
                c=0
                f=0
                for k in range(10+m,self.parent.Np-80):
                    if (np.mean(self.parent.pullback[i, j*h, self.parent.Np-k : self.parent.Np-k+10]) > 100):
                        c=0
                        f=1
                        I[i-s, j*h] = self.parent.Np-k
                    elif c<35 and f==1:
                        c+=1
                    elif c>=35 and f==1:
                        break
            #print(type())
                #print(I[i-s, j*h])
                #print(I[i-s, j*h] - m, I[i-s, j*h] -m -2*l )
                M1= np.mean(self.parent.pullback[i, j*h -l: j*h +l , I[i-s, j*h] - m -2*l : I[i-s, j*h] -m ])
                M2= np.mean(self.parent.pullback[i, j*h -l: j*h +l , I[i-s, j*h] + m : I[i-s, j*h] +m +2*l])
                c=0
                for k in range(I[i-s, j*h]+50,self.parent.Np-10):
                    if (np.mean(self.parent.pullback[i,j*h,k:k+10])>30):
                        c=1
                        break
                c=1
                if np.mean(self.parent.pullback[i, j*h , I[i-s, j*h] +10: I[i-s, j*h] +20])>250 and np.mean(self.parent.pullback[i, j*h , I[i-s, j*h] + 60 : I[i-s, j*h] +100])<30:
                    c=0
                    if((j*h-10)%G.shape[1]<(j*h+10)%G.shape[1]):
                        G[i-s,(j*h-10)%G.shape[1]:(j*h+10)%G.shape[1]]=1
                    else:
                        G[i-s,(j*h-10)%G.shape[1]:]=1
                        G[i-s,:(j*h+10)%G.shape[1]]=1
                #print(j,M2-M1>50 and I[i-s, j*h]!=0 and c==1,M2-M1,c)
                if M2-M1>50 and I[i-s, j*h]!=0 and c==1:
                    self.fixed_points.append((j*h,I[i-s, j*h]))
            #I[i-s]
            I[i-s],_=self.interpolate_intima(show=False)
            #self.refresh_points()
            self.fixed_points=[]
        self.add_shot()
        self.build_intima()  
        return I,G

    
    
    def lumen_detect (self,s,e):

        I=np.zeros([e-s,self.parent.Nr])
        G=np.zeros([e-s,self.parent.Nr])
        self.fixed_points=[]
        h=15
        p=int(self.parent.Nr/h)
    
        l=5
        m=10
    
        I=np.zeros([e-s,self.parent.Nr],dtype=int)
        for i in range(s,e):
            for j in range(1,p-1):
                #c=0
                c=0
                f=0
                for k in range(10+m,self.parent.Np-80):
                    if (np.mean(self.parent.pullback[i, j*h, self.parent.Np-k : self.parent.Np-k+10]) > 100):
                        c=0
                        f=1
                        I[i-s, j*h] = self.parent.Np-k
                    elif c<35 and f==1:
                        c+=1
                    elif c>=35 and f==1:
                        break
            #print(type())
                #print(I[i-s, j*h])
                #print(I[i-s, j*h] - m, I[i-s, j*h] -m -2*l )
                M1= np.mean(self.parent.pullback[i, j*h -l: j*h +l , I[i-s, j*h] - m -2*l : I[i-s, j*h] -m ])
                M2= np.mean(self.parent.pullback[i, j*h -l: j*h +l , I[i-s, j*h] + m : I[i-s, j*h] +m +2*l])
                c=0
                for k in range(I[i-s, j*h]+50,self.parent.Np-10):
                    if (np.mean(self.parent.pullback[i,j*h,k:k+10])>30):
                        c=1
                        break
                c=1
                if np.mean(self.parent.pullback[i, j*h , I[i-s, j*h] +10: I[i-s, j*h] +20])>250 and np.mean(self.parent.pullback[i, j*h , I[i-s, j*h] + 60 : I[i-s, j*h] +100])<30:
                    c=0
                    if((j*h-10)%G.shape[1]<(j*h+10)%G.shape[1]):
                        G[i-s,(j*h-10)%G.shape[1]:(j*h+10)%G.shape[1]]=1
                    else:
                        G[i-s,(j*h-10)%G.shape[1]:]=1
                        G[i-s,:(j*h+10)%G.shape[1]]=1
                #print(j,M2-M1>50 and I[i-s, j*h]!=0 and c==1,M2-M1,c)
                if M2-M1>50 and I[i-s, j*h]!=0 and c==1:
                    self.fixed_points.append((j*h,I[i-s, j*h]))
            #I[i-s]
            I[i-s],_=self.interpolate_intima(show=False)
            #self.refresh_points()
            self.fixed_points=[]
        self.add_shot()
        self.build_intima()  
        return I,G


    def intima_detect (self,s,e):

        I=np.zeros([e-s,self.parent.Nr])
        for i in range(s,e):
            print(i)
            for j in range(self.parent.Nr):
                for k in range(110,self.parent.Np):
                    if (np.mean(self.parent.pullback[i,j,k:k+5])>100):
                        I[i-s,j]=k
                        break
                if (I[i-s,j] == 0):
                    I[i-s,j]=I[i-s,j-1]
        return I
    
    def intima_detect_bis (self,s,e):
    
        I=np.zeros([e-s,self.parent.Nr])
        for i in range(s,e):
            for j in range(self.parent.Nr):
                #c=0
                for k in range(10,self.parent.Np-100):
                    if (np.mean(self.parent.pullback[i, j, self.parent.Np-k : self.parent.Np-k+10]) > 100):
                        I[i-s, j] = self.parent.Np-k
                if (I[i-s, j] == 0):
                    I[i-s, j] = I[i-s, j - 1]
        return I
    '''
    def find_intima(self):
        Ni=0
        Nf=60
        #J=self.intima_detect(Ni, Nf)
        K= self.intima_detect_bis(Ni,Nf)
        #I= (J + K)/2
        self.parent.lumen[Ni:Nf]=K
        self.add_shot()
        self.build_intima()
        Ni=0#self.index.get()
        Nf=LF#+permutation[self.index.get()-FF]+1
        #J=self.intima_detect(Ni, Nf)
        K,G= self.lumen_detect(Ni,Nf)
        #G=self.guida_fit(Ni,Nf)
        #I= (J + K)/2
        #print(np.max(G))
        self.parent.guide[Ni:Nf]=G
        #print(self.parent.labels.shape, self.parent.guide.shape)
        self.parent.labels[self.parent.guide==1]=unknown_color_index
        self.parent.lumen[Ni:Nf]=K
        self.add_shot()
        self.build_intima()
        self.update()
    '''
    def find_intima_interp(self):
        Ni=0
        Nf=60
        
        J=self.intima_detect(Ni, Nf)
        K= self.intima_detect_bis(Ni,Nf)
        I= (J + K)/2
        for i in range(Nf-Ni):
            for j in range(self.parent.Nr):
                if (np.abs(J[i,j] - K[i,j])<25):
                    self.fixed_points.append((j,I[i,j]))
            I[i],_=self.interpolate_intima(show=False)
            self.fixed_points=[]
            
        self.parent.lumen[Ni:Nf]=I
        self.add_shot()
        self.build_intima()
    
    def print_dist(self):
        
        center_x = self.sRadius
        center_y = self.sRadius
        
        d=np.sqrt(np.power(self.fixed_points_coord[0][0] - self.fixed_points_coord[1][0],2) + np.power(self.fixed_points_coord[0][1] - self.fixed_points_coord[1][1],2))
        d= int(d *(self.zoom.get()+120)/self.sRadius)
        print(d)
        
    def select_arc_manually(self):
        if not self.parent.current_class.get()==5:
            self.parent.current_class.set(4)
        if(self.select_arc.get()==0):
            self.interp_start.set(0)
            self.interp_end.set(0)
        self.select_arc.set((self.select_arc.get()+1)%2)
        button_text=['Modfy LA/FC','Arc selected']
        self.parent.button_canvas.button_arc.config(text=button_text[self.select_arc.get()])
        
    def find_arc(self):
        if self.parent.current_class.get()==4:
            arcs=self.selected_arcs
        if self.parent.current_class.get()==5:
            arcs=self.selected_arcs_fc
        for arc in arcs:
            if ( (self.starting_angle.get() > arc[0] and self.starting_angle.get() < arc[1]) and arc[0] < arc[1] ) or ( (self.starting_angle.get() > arc[0] or self.starting_angle.get() < arc[1]) and arc[0] > arc[1] ):
                self.interp_start.set(arc[0])
                self.interp_end.set(arc[1])
    
    
    def interpolate_guide_arc(self,f,a,b):

        frame_intima=np.concatenate([self.parent.lumen[f,-5:],self.parent.lumen[f,:],self.parent.lumen[f,0:5]])
        a+=5
        b+=5
        count=0

        d1=(frame_intima[a] - frame_intima[a-5])/5#np.mean(frame_intima[a-5:a]))/5
        d2=-(frame_intima[b] - frame_intima[b+5])/5#np.mean(frame_intima[b:b+5]))/5
        
        #print('bc: ',d1,d2)
        sense=1
        h=b-a
        if(a>b):
            sense=-1
            h=b+(self.parent.Nr-a)
        
        x=[]
        y=[]
        if(a>b):
            x=[a-5, (a-5+ b-5+self.parent.Nr)/2 , b-5+self.parent.Nr]
        else:
            x=[a-5,(a+b-10)/2,b-5]
        y=[self.parent.lumen[f,a-5],(self.parent.lumen[f,a-5]+self.parent.lumen[f,b-5])/2,self.parent.lumen[f,b-5]]
            
        x=np.asarray(x)       
        y=np.asarray(y)  
        #print("interpolate arc" , x , y)
        new_intima=interpolate.CubicSpline(x,y,bc_type=((1,d1),(1,d2)))
        nx=np.arange(a-5,a-5+h)
        
        if(a>b):
            self.parent.lumen[f,a-5:]=new_intima(nx)[:self.parent.Nr-a+5]
            self.parent.lumen[f,:b-5]=new_intima(nx)[self.parent.Nr-a+5:]
        else:
            self.parent.lumen[f,a-5:b-5]=new_intima(nx)
        
        
        
    def interpolate_arc(self,a,b):

            
        a+=5
        b+=5
        
        frame_intima=np.concatenate([self.parent.lumen[self.index.get(),-5:],self.parent.lumen[self.index.get(),:],self.parent.lumen[self.index.get(),0:5]])
        
        d1=(frame_intima[a] - frame_intima[a-5])/5#np.mean(frame_intima[a-5:a]))/5
        d2=-(frame_intima[b] - frame_intima[b+5])/5#np.mean(frame_intima[b:b+5]))/5
        
        count=0
        sense=1
        h=b-a
        if(a>b):
            sense=-1
            h=b+(self.parent.Nr-a)
        
        x=[]
        y=[]

        ordered_points=self.fixed_points
        ordered_points=sorted(ordered_points, key=lambda i: i[0])

        for i in range(len(self.fixed_points)):
            #print(i)
            #print(ordered_points[i][0] , a)
            if(((ordered_points[i][0]>=a-5 and ordered_points[i][0]<=b-5)and sense==1) or ((ordered_points[i][0]>=a-5 or ordered_points[i][0]<=b-5)and sense==-1)):
                if(a>b and ordered_points[i][0]<a-5):
                    x.append(ordered_points[i][0]+self.parent.Nr)
                    count+=1
                else:
                    x.append(ordered_points[i][0])
                y.append(ordered_points[i][1])
        
        if(a>b):
            x=x[count:] + x[:count]
            y=y[count:] + y[:count]
        
        x=np.asarray(x)       
        y=np.asarray(y)  
        #print("interpolate arc" , x)
        #print(ordered_points)
        new_intima=interpolate.CubicSpline(x,y,bc_type=((1,d1),(1,d2)))
        nx=np.arange(a-5,a-5+h)
    
        if(a>b):
            self.parent.lumen[self.index.get(),a-5:]=new_intima(nx)[:self.parent.Nr-a+5]
            self.parent.lumen[self.index.get(),:b-5]=new_intima(nx)[self.parent.Nr-a+5:]
        else:
            self.parent.lumen[self.index.get(),a-5:b-5]=new_intima(nx)

        '''
        if  self.parent.plaque_arcs[self.index.get()].size:          
            #self.parent.fct[self.index.get()]=np.copy([np.min(np.abs(self.parent.fc[self.index.get(),:]-self.parent.lumen[self.index.get(),:])),np.argmin(np.abs(self.parent.fc[self.index.get(),:]-self.parent.lumen[self.index.get(),:]))])
            self.calculate_fct_cart(self.index.get())
            save_list(self.parent.fct,self.parent.pullback_dir +'/Fct.txt')
        '''
        self.add_shot()
        self.build_intima()
        #self.build_fc()
            
            
           
    def interpolate_arc_fc(self,a,b):

            
        a+=5
        b+=5
        
        frame_intima=np.concatenate([self.parent.fc[self.index.get(),-5:],self.parent.fc[self.index.get(),:],self.parent.fc[self.index.get(),0:5]])
        if(frame_intima[a-5]==0):
            frame_intima[a-5:a+1]=np.ones(6)*self.parent.fc[self.index.get(),a]
        if(frame_intima[b+5]==0):
            frame_intima[b-1:b+5]=np.ones(6)*self.parent.fc[self.index.get(),b]
        
        d1=(frame_intima[a] - frame_intima[a-5])/5#np.mean(frame_intima[a-5:a]))/5
        d2=-(frame_intima[b] - frame_intima[b+5])/5#np.mean(frame_intima[b:b+5]))/5
        
        #print("deriv", d1, d2)
        #print(frame_intima[a-5:b+5])
        
        count=0
        sense=1
        h=b-a
        if(a>b):
            sense=-1
            h=b+(self.parent.Nr-a)
        
        x=[]
        y=[]

        ordered_points=self.fixed_points_fc
        ordered_points=sorted(ordered_points, key=lambda i: i[0])

        for i in range(len(self.fixed_points_fc)):
            #print(i)
            #print(ordered_points[i][0] , a)
            if(((ordered_points[i][0]>=a-5 and ordered_points[i][0]<=b-5)and sense==1) or ((ordered_points[i][0]>=a-5 or ordered_points[i][0]<=b-5)and sense==-1)):
                if(a>b and ordered_points[i][0]<a-5):
                    x.append(ordered_points[i][0]+self.parent.Nr)
                    count+=1
                else:
                    x.append(ordered_points[i][0])
                y.append(ordered_points[i][1])
        
        if(a>b):
            x=x[count:] + x[:count]
            y=y[count:] + y[:count]
        
        x=np.asarray(x)       
        y=np.asarray(y)  
        #print("interpolate arc" , x)
        #print(ordered_points)
        new_intima=interpolate.CubicSpline(x,y,bc_type=((1,d1),(1,d2)))
        nx=np.arange(a-5,a-5+h)

        if(a>b):
            self.parent.fc[self.index.get(),a-5:]=new_intima(nx)[:self.parent.Nr-a+5]
            self.parent.fc[self.index.get(),:b-5]=new_intima(nx)[self.parent.Nr-a+5:]
        else:
            self.parent.fc[self.index.get(),a-5:b-5]=new_intima(nx)
        
        if  self.parent.plaque_arcs[self.index.get()].size:          
            #self.parent.fct[self.index.get()]=np.copy([np.min(np.abs(self.parent.fc[self.index.get(),:]-self.parent.lumen[self.index.get(),:])),np.argmin(np.abs(self.parent.fc[self.index.get(),:]-self.parent.lumen[self.index.get(),:]))])
            self.calculate_fct_cart(self.index.get())
            save_list(self.parent.fct,self.parent.pullback_dir +'/Fct.txt')
        self.add_shot()
        self.build_intima()
        self.build_fc()
                    
    


    def interpolate_intima(self, show=True):
        ordered_points=self.fixed_points
        ordered_points=sorted(ordered_points, key=lambda i: i[0])
        x=[]
        y=[]
        '''
        x.append(0)
        y.append((ordered_points[0][1]+ordered_points[-1][1])/2)
        '''
        #print(len(self.fixed_points))
        for i in range(len(self.fixed_points)):
            x.append(ordered_points[i][0])
            y.append(ordered_points[i][1])
        ''''x.append(self.parent.Nr)
        y.append((ordered_points[0][1]+ordered_points[-1][1])/2) 
        '''
        x.append(ordered_points[0][0]+self.parent.Nr)
        y.append(ordered_points[0][1]) 
        
        x=np.asarray(x)       
        y=np.asarray(y)  
        new_intima=interpolate.CubicSpline(x,y,bc_type="periodic")
        nx=np.arange(self.parent.Nr)
        if show:
            self.parent.lumen[self.index.get(),:]=new_intima(nx)
            self.add_shot()
            self.build_intima()
        else:
            return new_intima(nx) , new_intima
        
    def regulirize_intima(self,l,n):
        box = np.ones(l+1)/(l+1)
        frame_intima=np.concatenate([self.parent.lumen[self.index.get(),-int(l/2):],self.parent.lumen[self.index.get(),:],self.parent.lumen[self.index.get(),0:int(l/2)]])
        #y_smooth = np.convolve(self.parent.lumen[self.index.get(),np.max([self.starting_angle.get()-n,0]):np.min([self.starting_angle.get()+n,self.parent.lumen.shape[1]])], box, mode='valid')
        y_smooth = np.convolve(frame_intima, box, mode='valid')
        self.parent.lumen[self.index.get(),np.max([self.starting_angle.get()-n,0]):np.min([self.starting_angle.get()+n,self.parent.Nr])]=np.copy(y_smooth)
        self.add_shot()
        self.build_intima()    
        

    def mouse_callback_press_left(self,event):

        if self.measure_on.get():
            center_x = self.sRadius
            center_y = self.sRadius
            i=0
            nearest_i=0
            min_dist=10000
            if not self.zoom_on.get() :
                for point in self.parent.measure_points[self.index.get()]:
                    r=self.parent.measure_points[self.index.get()][i][1]*self.sRadius/(self.zoom.get()+120)
                    x=r * np.cos(2*np.pi*point[0] / self.parent.Nr) + self.sRadius
                    y=r * np.sin(2*np.pi*point[0] / self.parent.Nr) + self.sRadius
                    d=np.sqrt(np.power(event.x-x,2)+np.power(event.y-y,2))
                    if( d<min_dist):
                        min_dist=d
                        nearest_i=i
                    i+=1
                self.select_point.set(nearest_i)
            else:
                for point in self.parent.measure_points[self.index.get()]:
                    R=self.zoomed_section.shape[0]/2
                    r=point[1]/(self.zoom.get())*R     
                    x = r * np.cos(2*np.pi*point[0] / self.parent.Nr) + R - self.zoomed_vertex[1] 
                    y = r * np.sin(2*np.pi*point[0] / self.parent.Nr) + R - self.zoomed_vertex[0] #+ 100
                    d=np.sqrt(np.power(event.x-x,2)+np.power(event.y-y,2))
                    if(d<min_dist):
                        min_dist=d
                        nearest_i=i
                    i+=1
                self.select_point.set(nearest_i)
            
            if self.deleting_points.get():
                if nearest_i%2==0:
                    del self.parent.measure_points[self.index.get()][nearest_i:nearest_i+2]
                    #del self.parent.measure_points[self.index.get()][nearest_i]
                else:
                    del self.parent.measure_points[self.index.get()][nearest_i-1:nearest_i+1]
                    #del self.parent.measure_points[self.index.get()][nearest_i-1]   
                self.build_measures()
            
        else :    
            if not self.zoom_on.get() :
                b=self.get_angle_index(event.x,event.y)
                self.starting_angle.set(b)
                if(self.parent.current_class.get()==4 or self.parent.current_class.get()==5 or self.parent.current_class.get()==6):
                    
                    if self.parent.current_class.get()==4 :
                        points=np.copy(self.fixed_points)
                    else:
                        points=np.copy(self.fixed_points_fc)
                        
                        
                    center_x = self.sRadius
                    center_y = self.sRadius
                    self.reference_shot.set(self.shots.get())
                    self.starting_rho.set(np.sqrt((event.x-center_x)*(event.x-center_x) + (event.y-center_y)*(event.y-center_y)))
                    
                    if not self.parent.current_class.get()==6: 
                        i=0
                        nearest_i=0
                        min_dist=np.abs(b-points[0][0])
                        for point in points:
                            if(np.abs(b-point[0])<min_dist):
                                min_dist=np.abs(b-point[0])
                                nearest_i=i
                            i+=1
                        self.select_point.set(nearest_i)
                        #print(self.select_point.get())
                        if self.deleting_points.get():
                            del self.fixed_points[nearest_i]
                            self.interpolate_intima(True)
                        #self.update()    
                    else:
                        self.starting_calibration=self.parent.calibration
            else:
                if self.parent.current_class.get()==6:
                    self.starting_point=(event.x,event.y)
                    self.starting_calibration=self.parent.calibration
                else:
                    self.rect_vert1 =np.array([event.x,event.y])
                    #-100
            
        
    def on_mouse_motion_callback_left(self,event):
        
        if self.measure_on.get() and not self.deleting_points.get():

            center_x = self.sRadius
            center_y = self.sRadius
            x=event.x
            y=event.y
            if self.zoom_on.get():
                #R=self.zoomed_section.shape[0]/2
                x=(x  + self.zoomed_vertex[1])/(self.zoom_now.get()) 
                y=(y  + self.zoomed_vertex[0])/(self.zoom_now.get()) 

            t=self.get_angle_index(x,y)
            if self.zoom_on.get():
                r=np.sqrt(np.power(x-center_x,2)+np.power(y-center_y,2))
                r= int(r *(self.zoom.get())/self.sRadius)
            else:
                r=int(np.sqrt(np.power(x-center_x,2)+np.power(y-center_y,2))/self.sRadius*(self.zoom.get()+120))
            self.parent.measure_points[self.index.get()][self.select_point.get()][0] = t
            self.parent.measure_points[self.index.get()][self.select_point.get()][1] = r
            self.build_measures()
            
        elif not self.zoom_on.get() and (not self.deleting_points.get()):
            b=self.get_angle_index(event.x,event.y)
        
            if (self.parent.current_class.get()==4 or self.parent.current_class.get()==5 or self.parent.current_class.get()==6):
                self.cut_history()
                center_x = self.sRadius
                center_y = self.sRadius
                rho=np.sqrt((event.x-center_x)*(event.x-center_x) + (event.y-center_y)*(event.y-center_y))
                d= rho - self.starting_rho.get()
                
                if self.parent.current_class.get()==6:
                    self.parent.calibration=self.starting_calibration + d*(self.zoom.get()+120)/self.sRadius
                    self.parent.calibration_scale=450/self.parent.calibration 
                    np.save(self.parent.pullback_dir + '/Calibration.npy',self.parent.calibration)
                    self.show_calibration()
                    
                else:
                    if self.parent.current_class.get()==4:
                        point_tag="point"+str(self.select_point.get())
                        r=(self.fixed_points[self.select_point.get()][1])*self.sRadius/(self.zoom.get()+120)#+d
                        t=self.fixed_points[self.select_point.get()][0]*2*np.pi/self.parent.Nr
                    else:
                        point_tag="point_fc"+str(self.select_point.get())
                        r=(self.fixed_points_fc[self.select_point.get()][1])*self.sRadius/(self.zoom.get()+120)#+d
                        t=self.fixed_points_fc[self.select_point.get()][0]*2*np.pi/self.parent.Nr
    
                    x=r*np.cos(t)
                    y=r*np.sin(t)
                    self.coords(point_tag,x-3+center_x,y-3+center_y,x+3+center_x,y+3+center_y)    
                    
                    d= int(d *(self.zoom.get()+120)/self.sRadius)
                    if self.parent.current_class.get()==4:
                        self.fixed_points[self.select_point.get()]=(self.fixed_points[self.select_point.get()][0],self.old_fixed_points[self.reference_shot.get()][self.select_point.get()][1]+d)
                        if not self.selected_arcs :
                            self.interpolate_intima()
                        else :
                            self.starting_angle.set(b)
                            self.find_arc()
                            self.interpolate_arc(self.interp_start.get(), self.interp_end.get())
                            
                        self.stopping_rho.set(self.old_intima[self.shots.get()][self.index.get(),self.starting_angle.get()] + d)
                    else :
                        self.fixed_points_fc[self.select_point.get()]=(self.fixed_points_fc[self.select_point.get()][0],self.old_fixed_points_fc[self.reference_shot.get()][self.select_point.get()][1]+d)
                        self.starting_angle.set(b)
                        self.find_arc()
                        self.interpolate_arc_fc(self.interp_start.get(), self.interp_end.get())
                        self.stopping_rho.set(self.old_fc[self.shots.get()][self.index.get(),self.starting_angle.get()] + d)
                
            else:
                self.update_line(b)
        elif not self.deleting_points.get():
            
            if self.parent.current_class.get()==6:
                R=self.zoomed_section.shape[0]/2
                
                xc= self.parent.cal_center_x.get()/(self.zoom.get())*R - self.zoomed_vertex[1]
                yc= self.parent.cal_center_y.get()/(self.zoom.get())*R - self.zoomed_vertex[0]
                
                x0=self.starting_point[0]/(self.zoom.get())*R - self.zoomed_vertex[1]
                y0=self.starting_point[1]/(self.zoom.get())*R - self.zoomed_vertex[0]
                
                x1=event.x/(self.zoom.get())*R - self.zoomed_vertex[1]
                y1=event.y/(self.zoom.get())*R - self.zoomed_vertex[0]
                
                r0=np.sqrt(np.power(x0-xc,2)+np.power(y0-yc,2))
                r1=np.sqrt(np.power(x1-xc,2)+np.power(y1-yc,2))
                d=r1-r0
                
                self.parent.calibration=self.starting_calibration + d*(self.zoom.get())/R*self.parent.Nf / (self.parent.Nf+120)#/self.zoom_now.get()/(self.parent.Nf / (self.parent.Nf+120) )
                self.parent.calibration_scale=450/self.parent.calibration 
                np.save(self.parent.pullback_dir + '/Calibration.npy',self.parent.calibration)
                self.show_calibration()
                
            else:
                l=np.max([ np.abs(event.x-self.rect_vert1[0]) , np.abs(event.y-self.rect_vert1[1]) ]) #-100
                self.rect_vert2=np.array([self.rect_vert1[0] - np.sign(self.rect_vert1[0]-event.x)* l, self.rect_vert1[1] - np.sign(self.rect_vert1[1]-event.y)*l] )
                
                v1=np.array([ np.min([self.rect_vert1[0],self.rect_vert2[0]]) , np.min([self.rect_vert1[1],self.rect_vert2[1]]) ])
                v2=np.array([ np.max([self.rect_vert1[0],self.rect_vert2[0]]) , np.max([self.rect_vert1[1],self.rect_vert2[1]]) ])
                self.delete('zsquare')
                self.create_rectangle(v1[0],v1[1],v2[0],v2[1],outline="#05f",tags='zsquare')
                self.zoom_vertex1 =np.array([(v1[0]+self.zoomed_vertex[1])/self.zoom_now.get(),(v1[1]+self.zoomed_vertex[0])/self.zoom_now.get()])
                self.zoom_vertex2 =np.array([(v2[0]+self.zoomed_vertex[1])/self.zoom_now.get(),(v2[1]+self.zoomed_vertex[0])/self.zoom_now.get()])
                    
    def mouse_callback_release_left(self,event):
            
        if not self.measure_on.get():
            if not self.zoom_on.get():
                b=self.get_angle_index(event.x,event.y)
                self.stopping_angle.set(b)
                
                if(self.parent.current_class.get()==1):
            
                    if self.stopping_angle.get()<self.starting_angle.get():
                        self.parent.lipid[self.index.get() , self.starting_angle.get():] =1#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                        self.parent.lipid[self.index.get() , 0:self.stopping_angle.get()] =1#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                        self.parent.labels[self.index.get()-0,self.starting_angle.get():]= lipid_color_index
                        self.parent.labels[self.index.get()-0,0:self.stopping_angle.get()]= lipid_color_index
                    else :
                        self.parent.lipid[self.index.get() , self.starting_angle.get():self.stopping_angle.get()] =1#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                        self.parent.labels[self.index.get()-0,self.starting_angle.get():self.stopping_angle.get()]= lipid_color_index
                    self.calculate_plaque_angle(self.index.get())
                    np.save(self.parent.pullback_dir +'/PlaqueAngle.npy',self.parent.plaque_angle)
                    save_list(self.parent.plaque_arcs,self.parent.pullback_dir +'/LipidArcs.txt')
                if(self.parent.current_class.get()==2):
            
                    if self.stopping_angle.get()<self.starting_angle.get():
                        self.parent.guide[self.index.get() , self.starting_angle.get():] =1#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                        self.parent.guide[self.index.get() , 0:self.stopping_angle.get()] =1#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                        self.parent.labels[self.index.get()-0,self.starting_angle.get():]= unknown_color_index
                        self.parent.labels[self.index.get()-0,0:self.stopping_angle.get()]= unknown_color_index
                    else :
                        self.parent.guide[self.index.get() , self.starting_angle.get():self.stopping_angle.get()] =1#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                        self.parent.labels[self.index.get()-0,self.starting_angle.get():self.stopping_angle.get()]= unknown_color_index
                    self.find_guide_arc(self.index.get())
                    save_list(self.parent.guide_arcs,self.parent.pullback_dir +'/GuideArcs.txt')
                    
                if(self.parent.current_class.get()==3):
            
                    if self.stopping_angle.get()<self.starting_angle.get():
                        self.parent.calcium[self.index.get() , self.starting_angle.get():] =1#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                        self.parent.calcium[self.index.get() , 0:self.stopping_angle.get()] =1#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                        self.parent.labels[self.index.get()-0,self.starting_angle.get():]= calcium_color_index
                        self.parent.labels[self.index.get()-0,0:self.stopping_angle.get()]= calcium_color_index
                    else :
                        self.parent.calcium[self.index.get() , self.starting_angle.get():self.stopping_angle.get()] =1#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                        self.parent.labels[self.index.get()-0,self.starting_angle.get():self.stopping_angle.get()]= calcium_color_index
                
                if(self.parent.current_class.get()==4):
                
                    self.calculate_area(self.index.get())
                    np.save(self.parent.pullback_dir +'/Area.npy',self.parent.areas)
                
                    self.parent.measure_canvas.update()
                
                '''if(self.parent.current_class.get()==4): quando modifichi con gaussiana
                
                    self.fixed_points_coord.append((event.x, event.y))
                    self.fixed_points.append((self.starting_angle.get(), self.stopping_rho.get()))
                '''
            else:
                '''#l=np.max([event.x-self.zoom_vertex1[0],event.y-self.zoom_vertex1[1]]) #-100
                l=np.max([ np.abs(event.x-self.zoom_vertex1[0]) , np.abs(event.y-self.zoom_vertex1[1]) ]) #-100
                self.zoom_vertex2=np.array([self.zoom_vertex1[0] + l,self.zoom_vertex1[1]+l])
                self.zommed_immage()'''
                #v1=np.array([ np.min([self.zoom_vertex1[0],self.zoom_vertex2[0]]) , np.min([self.zoom_vertex1[1],self.zoom_vertex2[1]]) ])
                #v2=np.array([ np.max([self.zoom_vertex1[0],self.zoom_vertex2[0]]) , np.max([self.zoom_vertex1[1],self.zoom_vertex2[1]]) ])
                #self.zoom_vertex1=v1
                #self.zoom_vertex2=v2
                if not self.parent.current_class.get()==6:
                    self.delete('zsquare')
                    self.zommed_immage()
        else:
            #save_list(self.parent.measure_points, self.parent.pullback_dir + '/FCMeasurement.txt')
            save_list([self.parent.measure_points,self.parent.measure_points_cart],self.parent.pullback_dir + "/FCMeasurement.txt")            
        self.update()

    def mouse_callback_press_right(self,event):

        if not self.zoom_on.get():

            b=self.get_angle_index(event.x,event.y)
            self.starting_angle.set(b)
            
            if self.measure_on.get():

                self.parent.measure_points_cart[self.index.get()].append([event.x,event.y])

                center_x = self.sRadius
                center_y = self.sRadius
                rho=np.sqrt((event.x-center_x)*(event.x-center_x) + (event.y-center_y)*(event.y-center_y))
                rho= int( rho *(self.zoom.get()+120)/self.sRadius )
                self.parent.measure_points[self.index.get()].append( [b, rho] )
                save_list([self.parent.measure_points,self.parent.measure_points_cart],self.parent.pullback_dir + "/FCMeasurement.txt")
                self.build_measures()
                ''''l=len(self.parent.measure_points[self.index.get()])
                point_tag = "measure_point" + str(l - 1)
                self.create_oval(event.x - 3, event.y - 3, event.x + 3 , event.y + 3 , tags=(point_tag, "measure_points"), fill='green')
                if len(self.parent.measure_points[self.index.get()])%2==0:
                    line_tag = "measure" + str(l//2 - 1)
                    #print(self.parent.measure_points_cart[self.index.get()])
                    d=np.sqrt(np.power(self.parent.measure_points_cart[self.index.get()][l-2][0] - self.parent.measure_points_cart[self.index.get()][l-1][0],2) + np.power(self.parent.measure_points_cart[self.index.get()][l-2][1] - self.parent.measure_points_cart[self.index.get()][l-1][1],2))
                    d= int(d *(self.zoom.get()+120)/self.sRadius)
                    
                    self.create_line( self.parent.measure_points_cart[self.index.get()][l-2][0] , self.parent.measure_points_cart[self.index.get()][l-2][1], event.x, event.y, fill='blue', width=1, tags=(line_tag, "measure"))
                    self.create_text(60, 20*(l//2) , text='Measure '+chr(65+(l//2-1))+'= ' +str(int(d* 0.88/196 *1000)), tags=['show_measure',line_tag], fill='blue')
                    self.create_text(self.parent.measure_points_cart[self.index.get()][l-2][0]*1.01,self.parent.measure_points_cart[self.index.get()][l-2][1]*1.01 , text=chr(65+(l//2-1)), tags=['letter',line_tag], fill='blue')
                '''
            
            if self.parent.current_class.get()==6:
                self.starting_point=(event.x,event.y)
                self.starting_cal_center=(self.parent.cal_center_x.get(),self.parent.cal_center_y.get())
            
            if(self.parent.current_class.get()==4 or self.parent.current_class.get()==5 ):

                self.cut_history()

                center_x = self.sRadius
                center_y = self.sRadius
                rho=np.sqrt((event.x-center_x)*(event.x-center_x) + (event.y-center_y)*(event.y-center_y))
                rho= int(rho *(self.zoom.get()+120)/self.sRadius)

                if (self.select_arc.get()==1):
                    if(self.interp_start.get()==self.interp_end.get()):
                        self.interp_start.set(b)
                    else:
                        self.interp_end.set(b)
                        if self.parent.current_class.get()==4:
                            self.selected_arcs.append((self.interp_start.get(),self.interp_end.get()))
                            
                        else:
                            self.selected_arcs_fc.append((self.interp_start.get(),self.interp_end.get()))
                            
                    if self.parent.current_class.get()==4:
                        self.fixed_points_coord.append((event.x, event.y))
                        self.fixed_points.append((self.starting_angle.get(), self.parent.lumen[self.index.get(),self.starting_angle.get()]))
                        r=self.parent.lumen[self.index.get(),self.starting_angle.get()]/(self.zoom.get()+120)
                        point_tag="point"+str(len(self.fixed_points)-1)
                    else:
                        self.fixed_points_fc.append((self.starting_angle.get(), self.parent.fc[self.index.get(),self.starting_angle.get()]))
                        self.fixed_points_fc_coord.append((event.x, event.y))
                        r=self.parent.fc[self.index.get(),self.starting_angle.get()]/(self.zoom.get()+120)
                        point_tag="point_fc"+str(len(self.fixed_points_fc)-1)
                    self.add_shot()
                    t=self.starting_angle.get()*2*np.pi/self.parent.Nr
                    x=r*np.cos(t)
                    y=r*np.sin(t)
                    self.create_oval(x-3+center_x,y-3+center_y,x+3+center_x,y+3+center_y, tags=(point_tag,"points"), fill='blue')

                else:
                    if self.parent.current_class.get()==4:
                        self.parent.lumen[self.index.get(),b]=rho
                        self.fixed_points_coord.append((event.x, event.y))
                        self.fixed_points.append((self.starting_angle.get(), rho))
                        point_tag="point"+str(len(self.fixed_points)-1)
                    else:
                        self.parent.fc[self.index.get(),b]=rho
                        self.fixed_points_fc_coord.append((event.x, event.y))
                        self.fixed_points_fc.append((self.starting_angle.get(), rho))
                        point_tag="point_fc"+str(len(self.fixed_points_fc)-1)
                    #amplitude=abs(rho-self.parent.lumen[self.index.get(),b])*3
                    #self.regulirize_intima(10,int(self.parent.lumen.shape[1]/2))#(np.min([5,amplitude]),amplitude)

                    self.add_shot()
                    r=rho*self.sRadius/(self.zoom.get()+120)
                    t=self.starting_angle.get()*2*np.pi/self.parent.Nr
                    x=r*np.cos(t)
                    y=r*np.sin(t)
                    self.create_oval(x-3+center_x,y-3+center_y,x+3+center_x,y+3+center_y, tags=(point_tag,"points"), fill='blue')
                    if self.parent.current_class.get()==4:
                        if not self.selected_arcs:
                            if(len(self.fixed_points)>4):
                               self.interpolate_intima()
                        else:
                            self.find_arc()
                            if(len(self.fixed_points)>2):
                                #print(self.interp_start)
                                self.interpolate_arc(self.interp_start.get(), self.interp_end.get())
                    elif self.selected_arcs_fc:
                        self.find_arc()
                        if(len(self.fixed_points_fc)>2):
                            #print(self.interp_start)
                            self.interpolate_arc_fc(self.interp_start.get(), self.interp_end.get())
                            #np.save(self.parent.lumen_path,self.parent.lumen)
        else:
            if self.measure_on.get():
                k=self.zoom_now.get()
                center_x = self.sRadius
                center_y = self.sRadius
                R=self.zoomed_section.shape[0]/2
                X= event.x + self.zoomed_vertex[1] 
                Y= event.y + self.zoomed_vertex[0] #- 100
                X=X/k#*(self.zoom.get()+120)/self.sRadius
                Y=Y/k#*(self.zoom.get()+120)/self.sRadius
                #Y+=100
                self.parent.measure_points_cart[self.index.get()].append([X,Y])
                rho=np.sqrt((X-center_x)*(X-center_x) + (Y-center_y)*(Y-center_y))
                #rho=int(rho)
                rho= int( rho *(self.zoom.get())/self.sRadius )
                b=self.get_angle_index(X,Y)#X+center_x,Y+center_y)
                self.parent.measure_points[self.index.get()].append( [b, rho] )
                self.build_measures()
                self.zoom_shift1=None 
                save_list([self.parent.measure_points,self.parent.measure_points_cart],self.parent.pullback_dir + "/FCMeasurement.txt")
            elif self.parent.current_class.get()==6:
                self.starting_point=(event.x,event.y)
                self.starting_cal_center=(self.parent.cal_center_x.get(),self.parent.cal_center_y.get())
            else:
                self.zoom_shift1 = np.array([event.x, event.y ]) #-100
                    
                
    def mouse_callback_release_right(self,event):
        if not self.zoom_on.get():
            b=self.get_angle_index(event.x,event.y)
            self.stopping_angle.set(b)
            if(self.parent.current_class.get()==1 or self.parent.current_class.get()==2 or self.parent.current_class.get()==3 ):
                if self.stopping_angle.get()<self.starting_angle.get():
                    self.parent.lipid[self.index.get() , self.starting_angle.get():] =0#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                    self.parent.lipid[self.index.get() , 0:self.stopping_angle.get()] =0#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                    self.parent.guide[self.index.get() , self.starting_angle.get():] =0#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                    self.parent.guide[self.index.get() , 0:self.stopping_angle.get()] =0#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                    self.parent.calcium[self.index.get() , self.starting_angle.get():] =0#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                    self.parent.calcium[self.index.get() , 0:self.stopping_angle.get()] =0#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                    self.parent.labels[self.index.get()-0,self.starting_angle.get():]= non_lipid_color_index
                    self.parent.labels[self.index.get()-0,0:self.stopping_angle.get()]= non_lipid_color_index
                else :
                    self.parent.lipid[self.index.get() , self.starting_angle.get():self.stopping_angle.get()] =0#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                    self.parent.guide[self.index.get() , self.starting_angle.get():self.stopping_angle.get()] =0#self.parent.labels.shape[1]-self.stopping_angle.get():self.parent.labels.shape[1]-self.starting_angle.get()]=1
                    self.parent.calcium[self.index.get() , self.starting_angle.get():self.stopping_angle.get()] =0
                    self.parent.labels[self.index.get()-0,self.starting_angle.get():self.stopping_angle.get()]= non_lipid_color_index

                self.find_guide_arc(self.index.get())
                save_list(self.parent.guide_arcs,self.parent.pullback_dir +'/GuideArcs.txt')
                save_list(self.parent.plaque_arcs,self.parent.pullback_dir +'/LipidArcs.txt')
            if (self.parent.current_class.get() == 4):
                self.calculate_area(self.index.get())
                np.save(self.parent.pullback_dir +'/Area.npy',self.parent.areas)
                
                self.parent.measure_canvas.update()
            self.calculate_plaque_angle(self.index.get())
            np.save(self.parent.pullback_dir +'/PlaqueAngle.npy',self.parent.plaque_angle)
            self.update()

    def on_mouse_motion_callback_right(self,event):
        if not self.zoom_on.get():
            if (self.parent.current_class.get()==6):
                self.parent.cal_center_x.set( int(self.starting_cal_center[0] + (event.x - self.starting_point[0])*(self.zoom.get()+120)/self.sRadius))
                self.parent.cal_center_y.set( int(self.starting_cal_center[1] + (event.y - self.starting_point[1])*(self.zoom.get()+120)/self.sRadius))
                np.save(self.parent.pullback_dir + '/CalibrationCenter.npy',(self.parent.cal_center_x.get(),self.parent.cal_center_y.get()))
                self.show_calibration()
                
            if (self.parent.current_class.get()==4):
                c=0#a buffo
                '''self.cut_history()

                center_x = self.sRadius
                center_y = self.sRadius
                rho=np.sqrt((event.x-center_x)*(event.x-center_x) + (event.y-center_y)*(event.y-center_y))
                d= rho - self.starting_rho.get()
                d= int(d *(self.zoom.get()+120)/self.sRadius)

                amplitude=np.abs(d)*3
                for i in range(amplitude):
                    j=i-int(amplitude/2)
                    #w=1
                    #if(j!=0):
                    w=np.power(np.e,-np.power(j/amplitude*2,2))#np.power(1/(np.abs(j//5)+1),1)
                    if(self.starting_angle.get()+j>0 and self.starting_angle.get()+j<self.parent.Nr):
                        self.parent.lumen[self.index.get(),self.starting_angle.get()+j]= self.old_intima[0][self.index.get(),self.starting_angle.get()+j] + int(w*d)
                self.add_shot()
                self.stopping_rho.set(self.old_intima[self.shots.get()][self.index.get(),self.starting_angle.get()] + d)
                self.build_intima()'''
                #np.save(self.parent.lumen_path,self.parent.lumen)
            else:
                b=self.get_angle_index(event.x,event.y)
                self.update_line(b)
        else:
            if self.parent.current_class.get()==6:
                R=self.zoomed_section.shape[0]/2
                
                x0=self.starting_point[0]/(self.zoom.get())*R - self.zoomed_vertex[1]
                y0=self.starting_point[1]/(self.zoom.get())*R - self.zoomed_vertex[0]
                
                x1=event.x/(self.zoom.get())*R - self.zoomed_vertex[1]
                y1=event.y/(self.zoom.get())*R - self.zoomed_vertex[0]
                
                xt=(x1-x0)*(self.zoom.get())/R
                yt=(y1-y0)*(self.zoom.get())/R
                
                self.parent.cal_center_x.set( int( self.starting_cal_center[0]+xt))
                self.parent.cal_center_y.set( int( self.starting_cal_center[1]+yt))
                np.save(self.parent.pullback_dir + '/CalibrationCenter.npy',(self.parent.cal_center_x.get(),self.parent.cal_center_y.get()))
                self.show_calibration() 


            else:    
                self.zoom_shift2 = np.array([event.x, event.y ])#-100
                if self.zoom_shift1 is None:
                    self.zoom_shift1=self.zoom_shift2
                self.zoom_shift = self.zoom_shift2-self.zoom_shift1
                self.move_zoom(self.zoom_shift)
                self.update()
            
    def undo(self):
        if(self.shots.get()>0):
            self.shots.set(self.shots.get()-1)
            self.parent.lumen[self.index.get()]= self.old_intima[self.shots.get()][self.index.get()]#np.copy(self.old_intima[self.shots.get()])
            self.parent.fc[self.index.get()]= self.old_fc[self.shots.get()][self.index.get()]#np.copy(self.old_intima[self.shots.get()])
            self.parent.fct= self.old_fct[self.shots.get()]#np.copy(self.old_intima[self.shots.get()])
            save_list(self.parent.fct,self.parent.pullback_dir +'/Fct.txt')
        
            self.fixed_points=list(self.old_fixed_points[self.shots.get()])
            self.fixed_points_fc=list(self.old_fixed_points_fc[self.shots.get()])
            #print(type(self.fixed_points))
            #print(self.fixed_points.shape)
            self.build_intima()
            self.build_fc()
            self.refresh_points()


class ResultCanvas(Canvas):
    def __init__(self, parent,data,mode, *args, **kwargs):
        Canvas.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        self.data = data
        self.mode = mode
        self.show_longitudinal=False
        #print(self.data.shape, np.transpose(self.data).shape)
        self.data=np.rot90(self.data)
        self.cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.copper(), N=number_of_colors)
        #self.data=resize(self.data, (self.parent.ann_canvas_h, self.parent.ann_canvas_w ))
        self.img = Image.fromarray(self.cmap(self.data, bytes=True))
        self.photoImage = ImageTk.PhotoImage(self.img)
        self.create_image(0, 0, image=self.photoImage, anchor='nw', tag='result_canvas_image')
        self.create_line(0, 0, 0, 0, fill='green', width=1, tag='result_canvas_section_line')
        self.create_line(0, 0, 0, 0, fill='blue', width=1, tag='result_canvas_radius_line')
        #self.create_line(0, 0, 0, 0, fill='cyan2', width=1, tag='LBI_start_line')
        #self.create_line(0, 0, 0, 0, fill='cyan2', width=1, tag='LBI_end_line')
    '''        
    def update(self):
        img = Image.fromarray(self.data)
        photoImage = ImageTk.PhotoImage(img)
        self.itemconfig('result_canvas_image', image=photoImage)
    '''    
    def update(self,i):
        if i==0:
            if self.show_longitudinal:
                self.parent.create_longitudinal()
                self.data = self.parent.longitudinal
                self.data = resize(self.data, (self.parent.ann_canvas_h, self.parent.ann_canvas_w))
            else:
                self.data = np.rot90(self.parent.labels)
                if self.parent.current_section_canvas.classification.get()==1 :
                    self.data = (number_of_colors-self.data)/4
                    self.data = resize(self.data, (self.parent.ann_canvas_h, self.parent.ann_canvas_w))
                #print(self.data*4)
                #if self.parent.current_section_canvas.classification.get()==1 :
                    self.data= (number_of_colors - self.data*4)#.astype(np.int)
                    self.data = np.rint(self.data).astype(np.int)
                 #self.data = (number_of_colors -  np.rint(self.data * 4)).astype(np.int)
                 #   self.img = Image.fromarray(self.data)  # image extension *.png,*.jpg
                    #new_width = self.parent.ann_canvas_w
                    #new_height = self.parent.ann_canvas_h
                    #self.img = self.img.resize((new_width, new_height), Image.ANTIALIAS)
                    #print(self.data)
                
        elif i==1:
            #self.data=self.parent.annotation
            self.data = np.rot90(self.parent.annotation)
            #self.data[ self.data==non_lipid_color_index ] = unknown_color_index
            #self.data[ self.data==lipid_color_index ] = non_lipid_color_index
            
            #self.data[ self.data==calcium_color_index ] = non_lipid_color_index
            #print(self.data)
            self.data = (number_of_colors - self.data) / 4
            self.data = resize(self.data, (self.parent.ann_canvas_h, self.parent.ann_canvas_w))
            self.data = (number_of_colors - self.data * 4)#.astype(np.int)
            self.data = np.rint(self.data).astype(np.int)
        elif i == 2:
            # self.data=self.parent.annotation
            self.data = np.rot90(self.parent.single_annotation)
            self.data = (number_of_colors - self.data) / 4
            self.data = resize(self.data, (self.parent.ann_canvas_h, self.parent.ann_canvas_w))
            self.data = (number_of_colors - self.data * 4)#.astype(np.int)
            self.data = np.rint(self.data).astype(np.int)
            #print(self.data)
        elif i==3:
            self.parent.create_longitudinal()
            self.data = self.parent.longitudinal
            self.data = resize(self.data, (self.parent.ann_canvas_h, self.parent.ann_canvas_w))
            
        if self.parent.current_section_canvas.classification.get()==0 :
            self.cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.parent.current_section_canvas.copper_continue(), N=2*number_of_colors)
        else :
            self.cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.parent.current_section_canvas.copper(), N=number_of_colors)
        
        ratio=self.parent.ann_canvas_w/self.parent.Nf
        self.data[:,:int(self.parent.interv_of_pertinence[0][0]*ratio)]=unknown_color_index
        #print(self.parent.interv_of_pertinence)
        for n in range(len(self.parent.interv_of_pertinence)-1) :
            self.data[:, int(self.parent.interv_of_pertinence[n][1]*ratio) : int(self.parent.interv_of_pertinence[n+1][0]*ratio) ]=unknown_color_index
        self.data[:,int(self.parent.interv_of_pertinence[-1][1]*ratio):]=unknown_color_index
        self.img = Image.fromarray(self.cmap(self.data, bytes=True))
        self.photoImage = ImageTk.PhotoImage(self.img)
        self.itemconfig('result_canvas_image', image=self.photoImage)

    @staticmethod
    def copper():
        x = np.linspace(0, 1, number_of_colors)
        r = 30
        R = np.arctan(r * x) / np.arctan(r)
        g = 10
        G = np.arctan(g * x) / np.arctan(g)
        b = 0.01
        B = np.arctan(b * x) / np.arctan(b)

        cmap = np.vstack((R, G, B))
        cmap = cmap.transpose()

        cmap[non_lipid_color_index, :] = RED
        cmap[lipid_color_index, :] = YELLOW
        cmap[calcium_color_index, :] = WHITE
        cmap[unknown_color_index, :] = BLACK
        cmap[deep_plaque_color_index, :] = ORANGE
        

        return cmap

    def on_mouse_motion_callback(self, event):
        section_index =  int(event.x /self.parent.ann_scale_w)
        if not self.show_longitudinal:
            self.parent.current_section_canvas.radius_index.set(int(event.y/self.parent.ann_scale_h))
        if section_index > self.parent.Nf-1:
            section_index = self.parent.Nf-1
        elif section_index < 0: #0:
            section_index = 0   #0

        if self.parent.current_section_canvas.radius_index.get() > self.parent.Nr:
            self.parent.current_section_canvas.radius_index.set( self.parent.Nr )
        elif self.parent.current_section_canvas.radius_index.get() < 0:
            self.parent.current_section_canvas.radius_index.set(0)
        #self.coords('result_canvas_section_line', section_index, 0, section_index, self.parent.Nr-1)
        #self.coords('result_canvas_radius_line', 0, radius_index, self.parent.Nf, radius_index)
        #self.parent.prediction_canvas.coords('result_canvas_radius_line',  self.parent.current_section_canvas.radius_index.get(), 0, self.parent.current_section_canvas.radius_index.get(), self.parent.Nf)
        self.parent.prediction_canvas.coords('result_canvas_section_line',  event.x, 0, event.x , self.parent.ann_canvas_h)
        #self.parent.annotation_canvas.coords('result_canvas_radius_line',  self.parent.current_section_canvas.radius_index.get(), 0, self.parent.current_section_canvas.radius_index.get(), self.parent.Nf)
        self.parent.annotation_canvas.coords('result_canvas_section_line', event.x, 0, event.x , self.parent.ann_canvas_h//2)
        self.parent.info_canvas.coords('result_canvas_section_line', event.x, 0, event.x , self.parent.ann_canvas_h//2)
        self.parent.current_section_canvas.index.set(section_index)
    
        self.parent.current_section_canvas.update_line(self.parent.current_section_canvas.radius_index.get())
        self.parent.current_section_canvas.old_intima=[np.copy(self.parent.lumen)]
        self.parent.current_section_canvas.old_fc=[np.copy(self.parent.fc)]
        self.parent.current_section_canvas.old_fct=[np.copy(self.parent.fct)]
        self.parent.current_section_canvas.shots.set(0)
        self.parent.current_section_canvas.fixed_points=[]
        self.parent.current_section_canvas.update()

        
    
    def mouse_callback_press_right(self,event):  
        if self.mode==0 :
            self.parent.insert_thr()
            
        else:
            self.parent.button_canvas.custom_info_bar()
              
    def update_radius(self):
        self.parent.prediction_canvas.coords('result_canvas_radius_line',  self.parent.current_section_canvas.radius_index.get(), 0, self.parent.current_section_canvas.radius_index.get(),self.parent.Nf)
        self.parent.annotation_canvas.coords('result_canvas_radius_line',  self.parent.current_section_canvas.radius_index.get(), 0, self.parent.current_section_canvas.radius_index.get(), self.parent.Nf)
        self.parent.current_section_canvas.update_line(self.parent.current_section_canvas.radius_index.get())
        #self.parent.current_section_canvas.update()

class TopDestroyer(Toplevel):
    def __init__(self,parent, *args, **kwargs):
        Toplevel.__init__(self, parent, *args, **kwargs)
        self.parent=parent
    def destroy(self):
        #self.parent.var=[]
        #self.delete('extremal_values')
        self.withdraw()
        #self.destroy()
        
class Longitudinal_scale(Canvas):
    def __init__(self, parent, *args, **kwargs):
        Canvas.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        self.i0=0
        self.i1=self.parent.ann_canvas_w
        self.var=[]
        self.pa_thr=180
        self.area_thr=3.5
        self.fc_thr=75
        self.top=TopDestroyer(parent=self)
        self.top.title("Manage intervals")
        self.top.withdraw()
        #self.var.append(Entry(self.top,width=3, borderwidth=0, highlightthickness=0))
        #self.var.append(Entry(self.top,width=3, borderwidth=0, highlightthickness=0))
        #self.var[0].insert(END, 0)
        #self.var[1].insert(END, 400)
                
                
    def refresh_measure(self):

        self.delete('righello','taccone','tacchetta','numeri')

        ratio=self.parent.Nf/self.parent.ann_canvas_w*0.2
        
        h=2.5/ratio#int(str(self.parent.ann_canvas_w*ratio/2.5)[:str(self.parent.ann_canvas_w*ratio/2.5).find('.')-1]+'0')#self.parent.ann_canvas_w/16
        #print(self.parent.ann_canvas_w,h,self.parent.ann_canvas_w/16)

        self.font_size = int(self.parent.sh*0.017)
        self.font_object = font.Font(family='Helvetica', size=self.font_size)
        
        self.create_line(0, 0, self.parent.ann_canvas_w  , 0, fill='green', width=2, tag='righello')
        self.create_line(0, 0, 0 , 10, fill='green', width=2, tag='taccone')
        self.create_text(0, 10,text='mm',fill='green', font=self.font_object,anchor=NW, tag='numeri')
        for i in range(1,int(self.parent.ann_canvas_w*ratio/2.5)):
            if i%4==0 and i<int(self.parent.ann_canvas_w*ratio/2.4)-2:
                self.create_line(i*h, 0, i*h , 10, fill='green', width=2, tag='taccone')
                self.create_text(i*h, 10,text=str(int(i*h*ratio)),fill='green', font=self.font_object,anchor=NW,tag='numeri')
            else:
                self.create_line(i*h, 0, i*h , 5, fill='green', width=1, tag='tacchetta')
        self.create_line(self.parent.ann_canvas_w , 0, self.parent.ann_canvas_w  ,10, fill='green', width=2, tag='taccone')
        self.create_text(self.parent.ann_canvas_w -20, 10,text=str(int(self.parent.ann_canvas_w*ratio)),fill='green', font=self.font_object,anchor=NW,tag='numeri')
        self.refresh_triangles()
        
    def refresh_triangles(self):
        n=0
        for interv in self.parent.interv_of_pertinence :
        
            self.i0=interv[0]*self.parent.ann_canvas_w/self.parent.Nf
            self.i1=interv[1]*self.parent.ann_canvas_w/self.parent.Nf
            self.delete('tr0-'+str(n))
            self.delete('tr1-'+str(n))
            self.create_polygon(self.triangle(self.i0),fill='blue', width=2,tags=['tr0-'+str(n),'triangles'])
            self.create_polygon(self.triangle(self.i1),fill='blue', width=2,tags=['tr1-'+str(n),'triangles'])
            n+=1
            
    def triangle(self,l):
        return [l-5,0,l+5,0,l,10]
        
    def mouse_callback_press_left(self,event):
    
        d=10000
        self.nearest=-1
        for n in range(2*len(self.parent.interv_of_pertinence)):
            
            dt=np.abs(event.x- int(self.parent.interv_of_pertinence[n//2][n%2]) * self.parent.ann_canvas_w/self.parent.Nf)
            if d > dt :
                self.nearest=n 
                d=dt
        print(self.nearest)
    
    def on_mouse_motion_callback(self,event):
        ''''n=0
        bn=0
        b=False
        for interv in self.parent.interv_of_pertinence :
            if interv[0]*self.parent.ann_canvas_w/self.parent.Nf<event.x and interv[1]*self.parent.ann_canvas_w/self.parent.Nf>event.x :
                self.i0=int(interv[0]*self.parent.ann_canvas_w/self.parent.Nf)
                self.i1=int(interv[1]*self.parent.ann_canvas_w/self.parent.Nf)
                b=True
                bn=n
            n+=1
        if b:
            if np.abs(event.x-self.i0)<np.abs(event.x-self.i1) :
                self.i0=event.x
                self.coords('tr0-'+str(bn-1),self.triangle(event.x))
                self.var[2*bn-2].delete(0, END)
                self.var[2*bn-2].insert(END, int(event.x*self.parent.Nf/self.parent.ann_canvas_w))
                #self.var[2*n-2].set(int(event.x*self.parent.Nf/self.parent.ann_canvas_w))
            else:
                self.i1=event.x
                self.coords('tr1-'+str(bn-1),self.triangle(event.x))
                self.var[2*bn-1].delete(0, END)
                self.var[2*bn-1].insert(END, int(event.x*self.parent.Nf/self.parent.ann_canvas_w))
                #self.var[2*n-1].set(int(event.x*self.parent.Nf/self.parent.ann_canvas_w))
        else:'''
        #if self.var:
        '''try:
            self.var[0]
        except NameError:
            print('dadada')
            self.var = []
        else:
            print('dododododo')
        '''
        #if self.var :
        x=int(event.x*self.parent.Nf/self.parent.ann_canvas_w)
        if (self.nearest == 0 and x>0 and x<int(self.var[1].get()) ) or (self.nearest == len(self.var)-1 and x>int(self.var[len(self.var)-2].get()) and x<self.parent.Nf ) or (not(self.nearest == 0 or self.nearest == len(self.var)-1) and x>int(self.var[self.nearest-1].get()) and x<int(self.var[self.nearest+1].get()) ): 
            self.var[self.nearest].delete(0, END)
            self.coords('tr' +str(self.nearest%2)+'-'+str(self.nearest//2),self.triangle(event.x))
            self.var[self.nearest].insert(END, x)
            self.apply()
    
    def mouse_callback_release_left(self,event):
        self.parent.warnings( A=float(self.area_thr), a=float(self.pa_thr), t=float(self.fc_thr))
        self.parent.annotation = self.parent.load_labels_classification()
        self.parent.prediction_canvas.update(0)
        self.parent.info_canvas.update(2)
        self.parent.annotation_canvas.update(1)

    def mouse_callback_press_right(self,event):            
       self.add_interval()
    
    def add_interval(self):
        self.top.deiconify()
        self.var=[]
        r=0
        self,
        for interv in self.parent.interv_of_pertinence :
            #self.var.append(IntVar(value=interv[0]))
            #self.var.append(IntVar(value=interv[1]))
            self.var.append(Entry(self.top,width=3, borderwidth=0, highlightthickness=0))
            self.var.append(Entry(self.top,width=3, borderwidth=0, highlightthickness=0))
            text0 = Label(self.top, width=5, text=" From frame ")
            #i0=Entry(self.top,width=3, borderwidth=0, highlightthickness=0)
            #i0.insert(END, interv[0])
            text1 = Label(self.top, width=5, text=" to frame ")#, anchor='w')
            #i1=Entry(self.top,width=3, borderwidth=0, highlightthickness=0)
            #i1.insert(END, interv[1])
            
            text0.grid(row=r,column=0,ipadx=20,ipady=5)
            #i0.grid(row=r,column=1)
            self.var[2*r].insert(END, interv[0]+1)
            self.var[2*r].grid(row=r,column=1)
            text1.grid(row=r,column=2,ipadx=13,ipady=5)
            #i1.grid(row=r,column=3)
            self.var[2*r+1].insert(END, interv[1]+1)
            self.var[2*r+1].grid(row=r,column=3)
            print(r)
            
            Button(self.top, text="Delete interval", command = lambda r=r : self.delete_interval(r) ).grid(row=r,column=4,ipadx=15)
            r+=1
        button_apply = Button(self.top, text="Apply", command = lambda : self.apply())
        button_new = Button(self.top, text="Add new interval", command = lambda : self.add_new_interval())
        button_apply.grid(row=r,column=2,columnspan=2)
        button_new.grid(row=r,column=0,columnspan=2)
        
    def apply(self):
        for i in range(len(self.parent.interv_of_pertinence)) :
            #print( self.var[2*i].get(), self.var[2*i+1].get() )
            self.parent.interv_of_pertinence[i] = ( int(self.var[2*i].get())-1 , int(self.var[2*i+1].get())-1 )
        save_list(self.parent.interv_of_pertinence,self.parent.pullback_dir + '/Intervals.txt')
        self.refresh_triangles()
        self.parent.warnings( A=float(self.area_thr), a=float(self.pa_thr), t=float(self.fc_thr))
        self.parent.annotation = self.parent.load_labels_classification()
        self.parent.prediction_canvas.update(0)
        self.parent.info_canvas.update(2)
        self.parent.annotation_canvas.update(1)

    def add_new_interval(self):
        self.parent.interv_of_pertinence.append((np.min([self.parent.interv_of_pertinence[-1][1]+1,self.parent.Nf]),self.parent.Nf))
        for widget in self.top.grid_slaves():
            widget.grid_forget()
        #self.top.destroy()
        self.add_interval()
        self.apply()
        
    def delete_interval(self,r):
        print(r,len(self.parent.interv_of_pertinence))
        del self.parent.interv_of_pertinence[r]
        for widget in self.top.grid_slaves():
            if int(widget.grid_info()["row"]) >= r :
                widget.grid_forget()
        if not self.parent.interv_of_pertinence :
            self.parent.interv_of_pertinence=[(0,self.parent.Nf)]
        #self.top.destroy()
        self.add_interval()
        

class MeasureCanvas(Canvas):
    def __init__(self, parent, *args, **kwargs):
        Canvas.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        self.font_size= int(self.parent.sh*0.023)
        self.font_object = font.Font(family='Helvetica', size=self.font_size)
        self.spaceline=self.font_size+int(self.font_size*0.7)
        self.last_line=5*self.spaceline
        c=30
        self.create_text(c, self.spaceline, text='', tags='show_patient', fill='green', font=self.font_object,anchor=NW)
        self.create_text(c, 2*self.spaceline, text='', tags='show_index', fill='green', font=self.font_object,anchor=NW)
        self.create_text(c, 3*self.spaceline, text='', tags='show_area', fill='green', font=self.font_object,anchor=NW)
        self.create_text(c, 4*self.spaceline, text='', tags='show_angle', fill='green', font=self.font_object,anchor=NW)
        self.create_text(c, 5*self.spaceline, text='', tags='show_fc', fill='green', font=self.font_object,anchor=NW)
        self.create_text(c, 7*self.spaceline, text='', tags='Max_diameter', fill='green', font=self.font_object, anchor=NW)
        self.create_text(c, 8*self.spaceline, text='', tags='Min_diameter', fill='green', font=self.font_object, anchor=NW)
        self.create_text(c, 6 * self.spaceline, text='', tags='Max_LCBI', fill='green', font=self.font_object, anchor=NW)
        #self.create_text(0, 150, text='', tags='show_thr', fill='green', font=self.font_object,anchor=NW)
        #self.create_text(0, 180, text='', tags='show_thr_ca', fill='green', self.font=font_object,anchor=NW)

    def update(self):
        self.itemconfig('show_patient', text='Patient : '+ os.path.basename(self.parent.pullback_dir))
        self.itemconfig('show_index', text='Frame : '+ str(self.parent.current_section_canvas.index.get()+1))
        self.itemconfig('show_area', text= 'Lumen Area : ' + str(round(self.parent.areas[self.parent.current_section_canvas.index.get()]* self.parent.calibration_scale* self.parent.calibration_scale/ ((818.5/160)*(818.5/160)),2)) + ' mm^2' )
        self.itemconfig('show_angle', text= 'Plaque Angle : ' + str( int( self.parent.plaque_angle[self.parent.current_section_canvas.index.get()] ) ) + '°' )#+ str( int( self.calculate_plaque_angle(self.index.get()) ) ) + '°' )#
        self.itemconfig('show_fc', text='FC Thickness : ' + str(round(self.parent.fct[self.parent.current_section_canvas.index.get()][0] * self.parent.calibration_scale )) + 'm-6')
        #self.itemconfig('show_thr', text= 'thr : ' + str(self.parent.current_section_canvas.thr.get() ) )
        #self.itemconfig('show_thr_ca', text= 'thr ca : ' + str(self.parent.current_section_canvas.thr_ca.get() ) )
        self.itemconfig('Max_diameter', text= 'Max diameter : ' + str(round(self.parent.diameters[1,self.parent.current_section_canvas.index.get(),0] * self.parent.calibration_scale)) + 'm-6')
        self.itemconfig('Min_diameter', text= 'Min diameter : ' + str(round(self.parent.diameters[0,self.parent.current_section_canvas.index.get(),0] * self.parent.calibration_scale)) + 'm-6')
        self.itemconfig('Max_LCBI', text='Max LCBI : ' + str(round(np.max(self.parent.lbi))))
        #print(round(self.parent.diameters[1,self.parent.current_section_canvas.index.get(),0] * self.parent.calibration_scale))

    def print_diameters(self):
        #self.del_diameters()
        #self.last_line += self.spaceline

        #self.last_line += self.spaceline
        self.itemconfig('Max_diameter', text='Max diameter : ' + str(round(self.parent.diameters[1, self.parent.current_section_canvas.index.get(), 0] * self.parent.calibration_scale)) + 'm-6')
        self.itemconfig('Min_diameter', text='Min diameter : ' + str(round(self.parent.diameters[0, self.parent.current_section_canvas.index.get(), 0] * self.parent.calibration_scale)) + 'm-6')

    def del_diameters(self):
        self.delete('diameters')
        self.last_line-=2*self.spaceline
        

class GentleButton(Button):
    def __init__(self, parent,explain, *args, **kwargs):
        Button.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        self.flag=0
        self.explain=explain
    def on_mouse_motion_callback(self,event):
        if self.flag:
            self.tw.destroy()
        self.flag=1
        x=event.x+self.winfo_rootx()
        y=event.y+self.winfo_rooty()
        #xp=
        #yp=
        #print(x,y,self.parent.winfo_rootx(),self.parent.winfo_rooty())
        
        self.tw = Toplevel(self.parent.parent)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = Label(self.tw, text=self.explain, justify='left',
                       background='yellow', relief='solid', borderwidth=1,
                       font=("times", "12", "normal"))
        label.pack(ipadx=1)
        '''
        self.parent.create_rectangle(x,y,x+200,y+30,fill='white',tags='fondo')
        #self.parent.tag_raise('fondo')
        self.font_size= int(self.parent.parent.sh*0.013)
        self.font_object = font.Font(family='Helvetica', size=self.font_size)
        self.parent.create_text(event.x,event.y, text='spiegazione', tags='info', fill='green', font=self.font_object,anchor=NW)
        '''
    def leave(self,event):
        if self.tw:
            self.tw.destroy()
            self.flag=0


class ButtonCanvas(Canvas):
    def __init__(self, parent, *args, **kwargs):
        Canvas.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        font_object = font.Font(family='Helvetica', size='20')
        #self.play_img=Image.open(os.path.dirname(os.path.realpath('__file__')) + "/play.jpg")#'C:\\Users\\Franz\\Desktop\\play.jpg')
        #self.play_photoImage = ImageTk.PhotoImage(self.play_img)
        button_play=Button(self, text="Play",  command=lambda: self.parent.play_pullback())#(label="Stop", onvalue=False, offvalue=True, variable=self.stop, command=lambda: self.play_pullback()) #image=self.play_photoImage,
        button_stop=Button(self, text="Stop", command=lambda: self.parent.stop_pullback())#(label="Stop", onvalue=False, offvalue=True, variable=self.stop, command=lambda: self.play_pullback())
        self.button_lumen=GentleButton(self, explain='Find Lumen Area in this frame. \n You can move the blue points after the analysis.', text="Lumen Area", command=lambda: self.parent.analyze_lumen(True))
        self.button_lipid = GentleButton(self, explain='Find Lipid and Calcium plaques in this frame',text="Lipid Plaque", command=lambda: self.parent.analyze_lipid(1))
        self.button_fc=GentleButton(self, explain='Find minimal fibrous cap thickness in this frame',text="FC", command=lambda: self.parent.analyze_fc(True))
        self.button_arc=GentleButton(self, explain='Select clockwise extremal points \n of the part you want modify. \n Select lumen or fibrous cap from \n the Modify Menu' , text="Modfy LA/FC", command=lambda: self.parent.current_section_canvas.select_arc_manually())
        self.button_diam=GentleButton(self, explain='Find min and max diameters of the lumen area', text="Show diameters", command=lambda: self.parent.current_section_canvas.calculate_diameter(self.parent.current_section_canvas.index.get()))
        self.button_delete_measures=GentleButton(self, explain='Delete the measures made on this frame', text="Delete measures", command=lambda: self.delete_measures())
        self.button_delete_points=GentleButton(self, explain='Delete points of the LA/FC or mesure by clicking on it ', text="Delete points", command=lambda: self.delete_points())
        self.button_longitudinal=GentleButton(self, explain='Switch between analysis results or longitudinal view', text="Longitudinal View", command=lambda: self.swich_longitudinal())
        #self.button_result_continue=Button(self, text="Show lipid probabilities", command=lambda: self.swich_results_mode())
        self.button_show_lumen=GentleButton(self, explain='Hide or show lines on the image', text="Hide lumen", command=lambda: self.show_lumen())
        #self.button_info=Button(self, text="Custom info bar", command=lambda: self.custom_info_bar())
        self.button_build=GentleButton(self, explain='Improve labels results', text="Smart build", command=lambda: self.parent.topological_build())
        self.button_macrophages=GentleButton(self, explain='Look for macrophages in this frame', text="Macrophages")#, command=lambda: self.parent.analyze_macrophages())

        self.button_LBI = GentleButton(self, explain='compute max lipid core burden index', text="LBI" , command=lambda: self.parent.analyze_LBI())

        self.zoom_button = Checkbutton(self, text="Zoom", onvalue=1, offvalue=False, variable=self.parent.current_section_canvas.zoom_on, command=lambda: self.parent.reset_zoom())
        self.measure_button = Checkbutton(self, text="Measure", onvalue=1, offvalue=False,variable=self.parent.current_section_canvas.measure_on, command=lambda: self.parent.measure_this())

        self.button_lumen.grid(row=0,column=0,pady=5,sticky='ew')
        self.button_arc.grid(row=0,column=1,pady=5,sticky='ew')
        self.button_show_lumen.grid(row=1,column=0,pady=5,sticky='ew')
        self.button_diam.grid(row=1,column=1,pady=5,sticky='ew')
        self.button_lipid.grid(row=2, column=0,pady=5,sticky='ew')
        self.button_fc.grid(row=2,column=1,pady=5,sticky='ew')
        self.button_delete_measures.grid(row=3,column=0,pady=5,sticky='ew')
        self.button_delete_points.grid(row=3,column=1,pady=5,sticky='ew')
        self.button_longitudinal.grid(row=4,column=0,pady=5,sticky='ew')
        #self.button_info.grid(row=4,column=1)
        #self.button_build.grid(row=5,column=0)
        #self.button_macrophages.grid(row=4,column=1,pady=5,sticky='ew')
        self.zoom_button.grid(row=4, column=1, pady=5, sticky='ew')
        self.measure_button.grid(row=5, column=1, pady=5, sticky='ew')

        button_play.grid(row=7,column=0,pady=5)
        button_stop.grid(row=7,column=1,pady=5)
        self.button_LBI.grid(row=5,column=0,pady=5,sticky='ew')


        #self.button_build.grid(row=5,column=1,pady=5,sticky='ew')
        #self.button_result_continue.grid(row=4,column=1)

        self.button_build.bind("<Enter>", self.button_build.on_mouse_motion_callback)
        self.button_build.bind("<Leave>", self.button_build.leave)

        self.button_lumen.bind("<Enter>", self.button_lumen.on_mouse_motion_callback)
        self.button_lumen.bind("<Leave>", self.button_lumen.leave)

        self.button_arc.bind("<Enter>", self.button_arc.on_mouse_motion_callback)
        self.button_arc.bind("<Leave>", self.button_arc.leave)

        self.button_show_lumen.bind("<Enter>", self.button_show_lumen.on_mouse_motion_callback)
        self.button_show_lumen.bind("<Leave>", self.button_show_lumen.leave)

        self.button_diam.bind("<Enter>", self.button_diam.on_mouse_motion_callback)
        self.button_diam.bind("<Leave>", self.button_diam.leave)

        self.button_lipid.bind("<Enter>", self.button_lipid.on_mouse_motion_callback)
        self.button_lipid.bind("<Leave>", self.button_lipid.leave)

        self.button_fc.bind("<Enter>", self.button_fc.on_mouse_motion_callback)
        self.button_fc.bind("<Leave>", self.button_fc.leave)

        self.button_delete_measures.bind("<Enter>", self.button_delete_measures.on_mouse_motion_callback)
        self.button_delete_measures.bind("<Leave>", self.button_delete_measures.leave)

        self.button_delete_points.bind("<Enter>", self.button_delete_points.on_mouse_motion_callback)
        self.button_delete_points.bind("<Leave>", self.button_delete_points.leave)

        self.button_longitudinal.bind("<Enter>", self.button_longitudinal.on_mouse_motion_callback)
        self.button_longitudinal.bind("<Leave>", self.button_longitudinal.leave)

        self.button_macrophages.bind("<Enter>", self.button_macrophages.on_mouse_motion_callback)
        self.button_macrophages.bind("<Leave>", self.button_macrophages.leave)

        self.button_LBI.bind("<Enter>", self.button_LBI.on_mouse_motion_callback)
        self.button_LBI.bind("<Leave>", self.button_LBI.leave)
        
    def swich_longitudinal(self):
        self.parent.prediction_canvas.show_longitudinal= not self.parent.prediction_canvas.show_longitudinal
        button_text=['Longitudinal View','Predictions']
        self.button_longitudinal.config(text=button_text[int(self.parent.prediction_canvas.show_longitudinal)])
        self.parent.prediction_canvas.update(0)
        
    def delete_measures(self):
        self.parent.measure_points[self.parent.current_section_canvas.index.get()] = []
        self.parent.measure_points_cart[self.parent.current_section_canvas.index.get()] = []
        self.parent.current_section_canvas.update()
        
    def delete_points(self):
        self.parent.current_section_canvas.deleting_points.set( not self.parent.current_section_canvas.deleting_points.get() )
        button_text=['Delete Points','Done']
        self.button_delete_points.config(text=button_text[int(self.parent.current_section_canvas.deleting_points.get())])
        
    def show_lumen(self):
        self.parent.current_section_canvas.show_lumen_fc=(self.parent.current_section_canvas.show_lumen_fc+1)%2
        button_text=['Show lumen','Hide Lumen']
        self.button_show_lumen.config(text=button_text[self.parent.current_section_canvas.show_lumen_fc])
        if  self.parent.current_section_canvas.show_lumen_fc :
            self.parent.current_section_canvas.create_line(0, 0, 0, 0, fill='blue', width=1, tag='section_canvas_line')
        self.parent.current_section_canvas.update()
        
        
    def swich_results_mode(self):
        self.parent.current_section_canvas.classification.set((self.parent.current_section_canvas.classification.get()+1)%2)
        button_text=['Show classification','Show lipid probabilities']
        self.button_result_continue.config(text=button_text[self.parent.current_section_canvas.classification.get()])
        if self.parent.current_section_canvas.classification.get()==0:
            self.parent.labels=self.parent.ricostruisci_labels_continue_guida(self.parent.preds, c=1, deg=64, stride=16)
            self.parent.labels=self.parent.labels/2 + 0.5
            #print('continue mode',np.max(self.parent.labels),np.min(self.parent.labels))
        else:
            self.parent.annotation=self.parent.load_labels_classification ()#self,thr=self.parent.current_section_canvas.thr.get(),thr_ca=self.parent.current_section_canvas.thr_ca.get() )
        
        self.parent.current_section_canvas.update()
    
    def custom_info_bar(self):
        top=Toplevel()
        top.title("Custom Info Bar")
        #msg = Message(self.top, text="select thr")
        #msg.pack()
        
        self.var_fc = IntVar(value=1)
        cb_fc = Checkbutton(top, text=" FCT threshold : ", variable=self.var_fc)
        #lab_fct = Label(self.tip, width=5, text=" FCT threshold : ")#, anchor='w')
        self.fc_thr=Entry(top)
        self.fc_thr.insert(END, '75')
        unit_fct = Label(top, width=5, text=" m-6")#, anchor='w')
        
        #lab_area = Label(self.tip, width=5, text=" Lumen area : ")#, anchor='w')
        self.var_area = IntVar(value=0)
        cb_area = Checkbutton(top, text=" Lumen area : ", variable=self.var_area)
        self.area_thr=Entry(top)
        self.area_thr.insert(END, '3.5')
        unit_area = Label(top, width=5, text=" mm2")#, anchor='w')
        
        #lab_pa = Label(self.tip, width=5, text=" Plaque angle : ")#, anchor='w')
        self.var_pa = IntVar(value=0)
        cb_pa = Checkbutton(top, text=" Plaque angle : ", variable=self.var_pa)
        self.pa_thr=Entry(top)
        self.pa_thr.insert(END, '180')
        unit_pa = Label(top, width=5, text="°")#, anchor='w')
        
        button_apply = Button(top, text="Apply changes", command = lambda : self.refresh_info_bar())
        button_quit = Button(top, text="Quit", command = top.destroy)
        
        cb_fc.grid(row=0,column=0)
        self.fc_thr.grid(row=0,column=1)
        unit_fct.grid(row=0,column=2)
        cb_area.grid(row=1,column=0)
        self.area_thr.grid(row=1,column=1)
        unit_area.grid(row=1,column=2)
        cb_pa.grid(row=2,column=0)
        self.pa_thr.grid(row=2,column=1)
        unit_pa.grid(row=2,column=2)
        
        button_apply.grid(row=3,column=0,columnspan=2)
        button_quit.grid(row=3,column=2)

    def refresh_info_bar(self):
        self.parent.load_single_annotation(float(self.area_thr.get()),self.var_area.get(),float(self.pa_thr.get()),self.var_pa.get(),float(self.fc_thr.get()),self.var_fc.get())
        self.parent.warnings( A=float(self.area_thr.get()), a=float(self.pa_thr.get()), t=float(self.fc_thr.get()))
        self.parent.annotation = self.parent.load_labels_classification()
        self.parent.scale_canvas.pa_thr=float(self.pa_thr.get())
        self.parent.scale_canvas.fc_thr=float(self.fc_thr.get())
        self.parent.scale_canvas.area_thr=float(self.area_thr.get())
        
        bar_label=''
        bar_ann_label=''
        bar_ann_label+='FCT < ' +str(float(self.fc_thr.get())) + ' micron'
        bar_ann_label+=' and '
        bar_ann_label+='Lumen Area < ' +str(float(self.area_thr.get())) + ' mm2'
        bar_ann_label+=' and '
        bar_ann_label+='Plaque Angle > ' +str(int(self.pa_thr.get())) + '°'
        if self.var_fc.get():
            bar_label+='FCT < ' +str(float(self.fc_thr.get())) + ' micron'
            if  self.var_area.get() or self.var_pa.get() :
                bar_label+=' or '
        if self.var_area.get():
            bar_label+='Lumen Area < ' +str(float(self.area_thr.get())) + ' mm2'
            if  self.var_pa.get() :
                bar_label+=' or '
        if self.var_pa.get():
            bar_label+='Plaque Angle > ' +str(int(self.pa_thr.get())) + '°'
        self.parent.annotation_label.config(text=bar_ann_label)
        self.parent.annotation_canvas.update(1)
        self.parent.info_bar_label.config(text=bar_label)
        self.parent.info_canvas.update(2)
            
        
class Main(Tk):
    def __init__(self):
        Tk.__init__(self)
        #self.attributes('-fullscreen', True)
        self.title('OCT+')
        self.create_canvases()
        self.create_menu()
        #self.import_pullback_callback()
        self.configure(background='black')
        self.grid_canvases()
        #self.pack_canvases()
        self.bind_events()
        #self.mainloop()
    
    def create_canvases(self):
        
        self.sw = int(self.winfo_screenwidth()*0.94)
        self.sh = int(self.winfo_screenheight()*0.86)
        #print(self.sh,self.sw)
        self.canvas_h=self.sh
        self.canvas_w=self.sh
        self.vborder_w = 10
        self.ann_canvas_h = int(self.sh*3/9)
        self.ann_canvas_w = self.sw - self.sh - 3*self.vborder_w
        self.yscrollbar = Scrollbar(self, orient=VERTICAL)
        self.current_section_canvas = SectionCanvas(self, width=self.canvas_w, height=self.canvas_h,  bg='black', highlightthickness=0) #(self, width=self.Np/1.45, height=self.Nr*1.6,  bg='white')
        #self.lungitudinal=self.create_longitudinal()
        self.info_canvas = ResultCanvas(self, data=np.ones([self.ann_canvas_w, self.ann_canvas_h // 6])*(unknown_color_index/number_of_colors), mode=2,
                                              width=self.ann_canvas_w, height=self.ann_canvas_h // 6, bg='black',
                                              highlightthickness=0)

        self.annotation_canvas = ResultCanvas(self,data=np.ones([self.ann_canvas_w,self.ann_canvas_h//6])*(unknown_color_index/number_of_colors), mode=1, width=self.ann_canvas_w, height=self.ann_canvas_h//6,  bg='black', highlightthickness=0)
        self.prediction_canvas = ResultCanvas(self, data=np.ones([self.ann_canvas_w, self.ann_canvas_h])*(non_lipid_color_index/number_of_colors),  mode=0, width=self.ann_canvas_w, height=self.ann_canvas_h,  bg='black', highlightthickness=0, yscrollcommand=self.yscrollbar.set)
        self.scale_canvas=Longitudinal_scale(self, width=self.ann_canvas_w, height=self.ann_canvas_h//8,  bg='black', highlightthickness=0)
        #self.yscrollbar.config(command=self.prediction_canvas.yview)
        self.button_canvas = ButtonCanvas(self, width=self.ann_canvas_w//3, height=int(self.sh*4/9), bg='black', highlightthickness=0)
        self.measure_canvas = MeasureCanvas(self, width=self.ann_canvas_w//3*2, height=int(self.sh*4/9), bg='black', highlightthickness=0)
        self.black_border_canvas = Canvas(self, width=self.ann_canvas_w//3*2, height=int(self.sh*1/15), bg='black', highlightthickness=0)
        #self.current_section_canvas.update()
        self.black_vertborder_canvas = Canvas(self, width=self.vborder_w, height=self.canvas_h, bg='black', highlightthickness=0)
        self.black_vertborder_canvas2 = Canvas(self, width=self.vborder_w, height=self.canvas_h, bg='black',highlightthickness=0)
        self.black_vertborder_canvas1 = Canvas(self, width=self.vborder_w, height=self.canvas_h, bg='black',highlightthickness=0)

    
    '''   
    def create_canvases(self):
        self.yscrollbar = Scrollbar(self, orient=VERTICAL)            
        self.current_section_canvas = SectionCanvas(self, width=self.Np/1.45, height=self.Nr*1.6,  bg='white') 
        self.annotation_canvas = ResultCanvas(self,data=self.annotation, width=self.Nr, height=self.labels.shape[0],  bg='white',yscrollcommand=self.yscrollbar.set)
        self.prediction_canvas = ResultCanvas(self, data=self.labels, width=self.Nr, height=self.labels.shape[0],  bg='white', yscrollcommand=self.yscrollbar.set)
        #self.yscrollbar.config(command=[self.annotation_canvas.yview,self.prediction_canvas.yview])
        self.yscrollbar.config(command=self.prediction_canvas.yview)
    '''  
    def create_menu(self):
        menubar = Menu(self.master)
        filemenu = Menu(menubar, tearoff=0)
        #helpmenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Import Pullback", command=self.import_new_pullback)
        filemenu.add_separator()
        filemenu.add_command(label="Save analysis", command=self.save_analysis)
        filemenu.add_command(label="Load analysis", command=self.load_analysis)
        filemenu.add_command(label="Compare Analysis", command=lambda: self.compare())
        filemenu.add_separator()
        filemenu.add_command(label="Analyze patients set", command=lambda: self.analyze_set())
        filemenu.add_separator()
        
        filemenu.add_command(label="Thresholds", command=lambda: self.insert_thr())
        filemenu.add_separator()
        filemenu.add_command(label="Guide gradients", command=lambda: self.accept_guide())
        filemenu.add_separator()
        filemenu.add_command(label="Manual calibration", command=lambda: self.manual_calibration())
        #filemenu.add_separator()
        #filemenu.add_command(label="Exit", command=self.master.destroy)
        
        menubar.add_cascade(label="Options", menu=filemenu)
        
        filemenu2 = Menu(menubar, tearoff=0)
        #helpmenu = tk.Menu(menubar, tearoff=0)
        single_frame = BooleanVar()
        single_frame.set(True)

        filemenu2.add_command(label="All", command=lambda: self.analyze_all())

        filemenu2.add_command(label="Missing informations", command=lambda: self.preprocessing())

        filemenu2.add_checkbutton(label="only this frame", onvalue=1, offvalue=False, variable=single_frame)
        
        filemenu2.add_command(label="Guide", command= lambda : self.analyze_guide())
        
        filemenu2.add_command(label="Lumen", command= lambda : self.analyze_lumen(single_frame.get()) )

        filemenu2.add_command(label="Lipid Plaque", command= lambda : self.analyze_lipid(single_frame.get()) )
        
        filemenu2.add_command(label="Fibrous cap", command= lambda : self.analyze_fc(single_frame.get()) )
        #, command=lambda: self.decrease_spicchio_index())
        #filemenu.add_separator()
        #filemenu.add_command(label="Exit", command=self.master.destroy)
        menubar.add_cascade(label="Analyze", menu=filemenu2)
        
        
        self.current_class = IntVar( master=self , value=0)
        self.choise_class= Menu(master=menubar, tearoff=0)
        self.choise_class.add_radiobutton(label="None", variable=self.current_class, value=0)  # , command= lambda : print("lipid selected " + str(self.current_class.get())))

        self.choise_class.add_radiobutton(label="Lipid Plaque",variable=self.current_class , value=1)#, command= lambda : print("lipid selected " + str(self.current_class.get())))
        self.choise_class.add_radiobutton(label="Artifact",variable=self.current_class , value=2)#, command= lambda : print("artifact selected " + str(self.current_class.get())))
        self.choise_class.add_radiobutton(label="Calcium Plaque",variable=self.current_class, value=3 )#, value=3, command= lambda : print("caclium selected " + str(self.current_class.get())))
        self.choise_class.add_radiobutton(label="Lumen",variable=self.current_class , value=4, command= lambda : self.current_section_canvas.update())
        self.choise_class.add_radiobutton(label="Fibrus Cap",variable=self.current_class , value=5, command= lambda : self.current_section_canvas.update())
        self.choise_class.add_radiobutton(label="Calibration",variable=self.current_class , value=6, command= lambda : self.current_section_canvas.update())
        

        menubar.add_cascade(label= "Modify",menu=self.choise_class)

        menubar.add_checkbutton(label="Zoom", onvalue=1, offvalue=0, variable=self.current_section_canvas.zoom_on,command=lambda : self.reset_zoom())
        menubar.add_checkbutton(label="Measure", onvalue=1, offvalue=False, variable=self.current_section_canvas.measure_on, command=lambda: self.measure_this())

        self.config(menu=menubar)


    def pack_canvases(self):
        #self.current_section_canvas.chk.pack(side=LEFT)
        #self.yscrollbar = Scrollbar(self.prediction_canvas)
        #self.yscrollbar.config(command=self.prediction_canvas.yview)


        self.annotation_canvas.pack(side=BOTTOM)
        self.prediction_canvas.pack(side=BOTTOM)
        self.current_section_canvas.pack(side=LEFT)
        #self.stop = BooleanVar()
        #self.stop.set(True)
        button_p=Button(self, text="Play", command=lambda: self.play_pullback())#(label="Stop", onvalue=False, offvalue=True, variable=self.stop, command=lambda: self.play_pullback())
        button_s=Button(self, text="Stop", command=lambda: self.stop_pullback())#(label="Stop", onvalue=False, offvalue=True, variable=self.stop, command=lambda: self.play_pullback())
        button_s.pack(side=BOTTOM)
        button_p.pack(side=BOTTOM)
        #self.play = BooleanVar()
        #self.play.set(False)
        #menubar.add_checkbutton(label="Play", onvalue=True, offvalue=False, variable=self.play, command=lambda: self.stop_pullback())
    
        #self.yscrollbar.pack(side=LEFT)
        #self.yscrollbar.pack(side=RIGHT)

    def grid_canvases(self):
        self.info_bar_label=Label(master=self,text='FCT < 75 micron',bg='black',fg='green')
        self.annotation_label=Label(master=self,text='FCT < 75 micron and Lumen Area < 3.5 mm2 and Plaque angle > 180°',bg='black',fg='green')
        self.black_vertborder_canvas.grid(row=0, column=0, rowspan=7)
        self.current_section_canvas.grid(row=0, column=1, rowspan=7 )
        self.black_vertborder_canvas1.grid(row=0, column=2, rowspan=7)
        self.button_canvas.grid(row=0, column=3)
        self.measure_canvas.grid(row=0, column=4)
        self.info_bar_label.grid(row=1, column=3, columnspan=2 ,sticky=W,padx=25)
        self.info_canvas.grid(row=2, column=3, columnspan=2 )
        self.annotation_label.grid(row=3, column=3, columnspan=2 ,sticky=W,padx=25 )
        self.annotation_canvas.grid(row=4, column=3, columnspan=2 )
        self.scale_canvas.grid(row=5, column=3, columnspan=2 )
        self.prediction_canvas.grid(row=6, column=3, columnspan=2 )
        self.black_border_canvas.grid(row=7, column=3, columnspan=2 )
        self.black_vertborder_canvas2.grid(row=0, column=5, rowspan=7 )

        self.prediction_canvas.tk.call('tk', 'scaling', 1)

    def yview(self, *args):
        self.annotation_canvas.yview(args)
        self.prediction_canvas.yview(args)
    
    def bind_events(self):
        self.bind("<o>", lambda e: self.import_pullback_callback())
        self.bind("<q>", lambda e: self.quit())
        self.bind("<Right>", lambda e: self.increase_section_index())
        self.bind("<Left>", lambda e: self.decrease_section_index())
        self.bind("<k>", lambda e: self.increase_section_index())
        self.bind("<j>", lambda e: self.decrease_section_index())
        
        self.bind("<z>", lambda e: self.increase_section_zoom())
        self.bind("<x>", lambda e: self.decrease_section_zoom())
        self.bind("<p>", lambda e: self.play_pullback())
        self.bind("<s>", lambda e: self.stop_pullback())
        self.scale_canvas.bind("<B1-Motion>", self.scale_canvas.on_mouse_motion_callback)
        self.info_canvas.bind("<B1-Motion>", self.info_canvas.on_mouse_motion_callback)
        self.annotation_canvas.bind("<3>", self.annotation_canvas.mouse_callback_press_right)
        self.info_canvas.bind("<3>", self.info_canvas.mouse_callback_press_right)
        self.prediction_canvas.bind("<3>", self.prediction_canvas.mouse_callback_press_right)
        self.scale_canvas.bind("<3>", self.scale_canvas.mouse_callback_press_right)
        self.scale_canvas.bind("<1>", self.scale_canvas.mouse_callback_press_left)
        self.scale_canvas.bind("<ButtonRelease-1>", self.scale_canvas.mouse_callback_release_left)
        
        self.annotation_canvas.bind("<B1-Motion>", self.annotation_canvas.on_mouse_motion_callback)
        self.prediction_canvas.bind("<B1-Motion>", self.prediction_canvas.on_mouse_motion_callback)
        self.bind("<m>", lambda e: self.increase_radius_index())
        self.bind("<n>", lambda e: self.decrease_radius_index())
        self.bind("<h>", lambda e: self.increase_spicchio_index())
        self.bind("<g>", lambda e: self.decrease_spicchio_index())
        self.bind("<r>", lambda e: self.current_section_canvas.regulirize_intima(10, self.Nr))
        self.bind("<f>", lambda e: self.current_section_canvas.lumen_ds(self.current_section_canvas.index.get()),True)#boundary(self.current_section_canvas.index.get(),self.current_section_canvas.lumen_weights()))#find_intima())
        self.bind("<a>", lambda e: self.current_section_canvas.select_arc_manually())
        self.bind("<l>", lambda e: self.current_section_canvas.print_dist())
        self.bind("<Control-z>", lambda e: self.current_section_canvas.undo())
        self.current_section_canvas.bind("<Button-1>", self.current_section_canvas.mouse_callback_press_left)
        self.current_section_canvas.bind("<ButtonRelease-1>", self.current_section_canvas.mouse_callback_release_left)
        self.current_section_canvas.bind("<3>", self.current_section_canvas.mouse_callback_press_right)
        self.current_section_canvas.bind("<ButtonRelease-3>", self.current_section_canvas.mouse_callback_release_right)
        self.current_section_canvas.bind("<B1-Motion>", self.current_section_canvas.on_mouse_motion_callback_left)
        self.current_section_canvas.bind("<B3-Motion>", self.current_section_canvas.on_mouse_motion_callback_right)
        self.bind("<b>", lambda e: self.add_bad())
        self.bind("<v>", lambda e: self.add_good())
        self.bind("<space>", lambda e:  self.v_stop.set(1))
        self.bind("<w>", lambda e: self.save_frame())
        #self.bind("<s>", lambda e: self.save_bad())

    def save_frame(self):
        ps = self.current_section_canvas.postscript(colormode='color')
        img = Image.open(io.BytesIO(ps.encode('utf-8')))
        img.save(self.pullback_dir + '/frame.jpg')
    
    def save_chemiogram(self,path=None, patient=None):    
        if path==None :
            path=self.pullback_dir
        if patient is None:
            file_name = '/Chemiogram.jpg'
        else:
            file_name = '/Chemiogram '+patient+'.jpg'
            print('yeah',file_name)
        ps = self.prediction_canvas.postscript(colormode='color')
        img = Image.open(io.BytesIO(ps.encode('utf-8')))
        img.save(path + file_name)
    

    def save_analysis(self):
        d=datetime.now()
        folder_name=str(d.day)+'_'+str(d.month)+'_'+str(d.year)
        folder_path=os.path.join(self.pullback_dir,folder_name)
        os.makedirs(folder_path)
        np.save(os.path.join(folder_path,"GuideArtifact.npy"),self.guide)
        np.save(os.path.join(folder_path,"Lumen.npy"),self.lumen)
        np.save(os.path.join(folder_path,"LipidPlaque.npy"),self.lipid)
        np.save(os.path.join(folder_path,"CalciumPlaque.npy"),self.calcium)
        np.save(os.path.join(folder_path,"FibrusCap.npy"),self.fc)
        np.save(os.path.join(folder_path, "annotation.npy"),self.annotation)

        save_list(self.guide_arcs,os.path.join(folder_path, 'GuideArcs.txt'))
        np.save(os.path.join(folder_path, 'Area.npy'),self.areas)
        save_list(self.plaque_arcs,os.path.join(folder_path, 'LipidArcs.txt'))
        np.save(os.path.join(folder_path, 'PlaqueAngle.npy'),self.plaque_angle)
        save_list(self.fct,os.path.join(folder_path, 'Fct.txt'))
        save_list( [self.measure_points,self.measure_points_cart], os.path.join(folder_path,"FCMeasurement.txt") )
        
    def load_analysis(self):
        folder_path = askdirectory(initialdir=self.pullback_dir)
        self.pullback_dir=folder_path
        self.guide=np.load(os.path.join(folder_path,"GuideArtifact.npy"),mmap_mode='r+')
        self.lumen=np.load(os.path.join(folder_path,"Lumen.npy"),mmap_mode='r+')
        self.lipid=np.load(os.path.join(folder_path,"LipidPlaque.npy"),mmap_mode='r+')
        self.calcium=np.load(os.path.join(folder_path,"CalciumPlaque.npy"),mmap_mode='r+')
        self.fc=np.load(os.path.join(folder_path,"FibrusCap.npy"),mmap_mode='r+')
        self.annotation=np.load(os.path.join(folder_path, "annotation.npy"),mmap_mode='r+')

        self.guide_arcs=load_list(os.path.join(folder_path, 'GuideArcs.txt'))
        self.areas=np.load(os.path.join(folder_path, 'Area.npy'),mmap_mode='r+')
        self.plaque_arcs=load_list(os.path.join(folder_path, 'LipidArcs.txt'))
        self.plaque_angle=np.load(os.path.join(folder_path, 'PlaqueAngle.npy'),mmap_mode='r+')
        self.fct=load_list(os.path.join(folder_path, 'Fct.txt'))
        mes=load_list( os.path.join(folder_path,"FCMeasurement.txt") )
        self.measure_points=mes[0]
        self.measure_points_cart=mes[1]
        self.load_labels_classification()
        self.current_section_canvas.update()
        
    def compare(self):
        
        top=Toplevel()
        
        annotation_canvas = Canvas(top, width=self.Nr, height=self.Nf,  bg='white')
        prediction_canvas = Canvas(top, width=self.Nr, height=self.Nf,  bg='white')
        
        data=np.copy(self.labels)
        if self.current_section_canvas.classification.get()==0 :
            cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.current_section_canvas.copper_continue(), N=2*number_of_colors)
        else :
            cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.current_section_canvas.copper(), N=number_of_colors)
        img = Image.fromarray(cmap(data, bytes=True))
        self.photoImage = ImageTk.PhotoImage(img)
        prediction_canvas.create_image(0, 0, image=[], anchor='nw', tag='result_canvas_image')
        prediction_canvas.itemconfig('result_canvas_image', image=self.photoImage)
        prediction_canvas.pack(side=LEFT)

        self.data_path = askdirectory(initialdir=self.pullback_dir)
        guide=np.load(self.data_path+"/GuideArtifact.npy")
        lipid=np.load(self.data_path+"/LipidPlaque.npy")
        calcium=np.load(self.data_path+"/CalciumPlaque.npy")
        labels=np.ones(self.labels.shape,dtype=np.uint16)*non_lipid_color_index
        labels[lipid==1]=lipid_color_index
        labels[calcium==1]=calcium_color_index
        labels[guide==1]=unknown_color_index
        data=labels
        if self.current_section_canvas.classification.get()==0 :
            cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.current_section_canvas.copper_continue(), N=2*number_of_colors)
        else :
            cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.current_section_canvas.copper(), N=number_of_colors)
        img = Image.fromarray(cmap(data, bytes=True))
        self.photoImage2 = ImageTk.PhotoImage(img)
        annotation_canvas.create_image(0, 0, image=[], anchor='nw', tag='annotation_canvas_image')
        annotation_canvas.itemconfig('annotation_canvas_image', image=self.photoImage2)
        annotation_canvas.pack(side=LEFT)

        top.bind("<q>", lambda e: top.destroy())
        self.statistics(self.lipid, lipid)
        WriteOrdDictToCSV(os.path.join(self.data_path, 'test_result.csv'), Info)
        
        labels=labels.flatten()
        preds=np.copy(self.labels).flatten()
        n=[]
        o=[]
        for i in range (0,labels.size,16):
            n.append(labels[i])
            o.append(preds[i])
        print(classification_report(labels, preds))
        

    def statistics (self, l1, l2 ):
        scores = np.zeros((12, 1))
        n=0
        for i in range(l1.shape[0]):
            for j in range(l1.shape[1]):
                if (l1[i, j] == l2[i, j]):
                    if (l1[i, j] == 0):
                        scores[2, n] = scores[2, n] + 1  # TN
                    else:
                        scores[0, n] = scores[0, n] + 1  # TP
                else:
                    if (l1[i, j] == 0):
                        scores[3, n] = scores[3, n] + 1  # FN
                    elif (l1[i, j] == 1):
                        scores[1, n] = scores[1, n] + 1  # FP
        if scores[0, n] + scores[1, n] > 0:
            scores[4, n] = np.divide(scores[0, n], (scores[0, n] + scores[
                1, n]))  # precision(TP tra i P, percentuale di placca tra quella rilevata)
        if scores[0, n] + scores[3, n] > 0:
            scores[5, n] = np.divide(scores[0, n], (
                        scores[0, n] + scores[3, n]))  # recall   (placca trovata ta quella esistente) (TPR)
        if scores[1, n] + scores[2, n] > 0:
            scores[6, n] = np.divide(scores[1, n], (scores[1, n] + scores[2, n]))  # FPR
        TNR = np.divide(scores[2, n], (scores[2, n] + scores[1, n]))
        scores[7, n] = 2 * np.divide(np.multiply(scores[4, n], scores[5, n]), (scores[4, n] + scores[5, n]))  # f1
        scores[8, n] = np.divide(scores[5, n], scores[6, n])  # PL
        scores[9, n] = np.divide(1 - scores[5, n], 1 - scores[6, n])  # NL
        scores[10, n] = np.divide(scores[8, n], scores[9, n])  # DOR
        # scores[11,n]=np.divide(np.multiply(scores[0,n],scores[2,n])-np.multiply(scores[1,n],scores[3,n]),np.sqrt(np.multiply(scores[0,n]+scores[1,n],scores[0,n]+scores[3,n],scores[2,n]+scores[1,n],scores[2,n]+scores[3,n])))
        scores[11, n] = (scores[0, n] * scores[2, n] - scores[1, n] * scores[3, n]) / np.sqrt((scores[0, n] + scores[1, n]) * (scores[0, n] + scores[3, n]) * (scores[2, n] + scores[1, n]) * (scores[2, n] + scores[3, n]))

        Info['TP'] = scores[0, n]
        Info['FP'] = scores[1, n]
        Info['TN'] = scores[2, n]
        Info['FN'] = scores[3, n]
        Info['FPR'] = scores[6, n]
        Info['TNR'] = TNR
        Info['precision'] = scores[4, n]
        Info['recall'] = scores[5, n]
        Info['f1'] = scores[7, n]
        Info['MCC'] = scores[11, n]

    
    def insert_thr(self):
        self.top=Toplevel()
        self.top.title("Select thresholds")
        #msg = Message(self.top, text="select thr")
        #msg.pack()
        
        self.input_w=Entry(self.top)
        lab = Label(self.top, width=10, text="   thr  : ", anchor='w')
        #self.input_var = StringVar()
        #self.input_w.config(textvariable=self.input_var)
        self.input_w.insert(END, '0.65')
        lab.pack(side=LEFT)
        self.input_w.pack(side=LEFT)
        
        lab2 = Label(self.top, width=10, text="   thr_ca  : ", anchor='w')
        self.input_w_ca=Entry(self.top)
        #self.input_var = StringVar()
        #self.input_w.config(textvariable=self.input_var)
        self.input_w_ca.insert(END, '0.4')
        lab2.pack(side=LEFT)
        self.input_w_ca.pack(side=LEFT) 
        
        button3 = Button(self.top, text="Quit", command = self.top.destroy)
        button3.pack(side=BOTTOM)
        button2 = Button(self.top, text="Calculate FC", command = lambda: self.adapt_fc())
        button2.pack(side=BOTTOM)
        button = Button(self.top, text="Done", command= lambda: self.insert_thr_value())
        button.pack(side=BOTTOM)
        
        
        
    def create_longitudinal (self):
        r=self.current_section_canvas.radius_index.get()
        #self.longitudinal=np.zeros([self.Nf,2*self.Np])
        #for f in range (self.Nf):
        #pr=self.pullback[:,r,:][:,::-1]
        #print(pr[0,:4],self.pullback[:,r,:][0,-4:],self.pullback[:,r,:][0,:4])
        self.longitudinal=np.concatenate([self.pullback[:,r,:][:,::-1], self.pullback[:,(r+int(self.Nr/2))%self.Nr,:] ], axis=1)
        self.longitudinal=np.transpose(self.longitudinal)/np.max(self.longitudinal) + 0.005
        
    @staticmethod
    def linear_bw():
        x = np.linspace(0, 1, number_of_colors)
        #R = x*0.6#np.arctan(r*x)/np.arctan(r)
        #G = 0.9*np.ones(number_of_colors)#np.arctan(g*x)/np.arctan(g)
        #B = (1-x)#np.arctan(b*x)/np.arctan(b)

        r = 30
        R = np.arctan(r * (1 - x)) / np.arctan(r)
        g = 30
        G = np.arctan(g * x) / np.arctan(g)
        b = 0.000000000
        B = np.arctan(b * x) / np.arctan(g)

        cmap = np.vstack((R, G, B))
        cmap = cmap.transpose()

        cmap[calcium_color_index, :] = RED
        cmap[unknown_color_index, :] = RED

        return cmap
    
        
    def insert_thr_value(self):
        self.current_section_canvas.thr.set(float(self.input_w.get()))
        self.current_section_canvas.thr_ca.set(float(self.input_w_ca.get()))
        #self.build_labels()
        self.topological_build()
        self.current_section_canvas.update()
        
    def adapt_fc(self):
        d=self.current_section_canvas.fc_weights()
        for f in range(self.Nf):
            self.current_section_canvas.calculate_plaque_angle(f)
            self.current_section_canvas.fc_boundary(f,d)
        np.save(self.pullback_dir +'/PlaqueAngle.npy',self.plaque_angle)
        save_list(self.plaque_arcs,self.pullback_dir +'/LipidArcs.txt')
        save_list(self.fct,self.pullback_dir +'/Fct.txt')
        

    def manual_calibration(self):
        self.cal_win=Toplevel()
        self.cal_win.title("Manual calibration")
        #msg = Message(self.top, text="select thr")
        #msg.pack()
        self.cal_radius=Entry(self.cal_win)
        labr = Label(self.cal_win, width=5, text="   Radius  : ", anchor='w')
        self.cal_radius.insert(END, str(int(self.calibration)))
        labr.grid(row=0,column=0)
        self.cal_radius.grid(row=0,column=1)
        
        self.cal_x=Entry(self.cal_win)
        labx = Label(self.cal_win, width=5, text="   Center X  : ", anchor='w')
        self.cal_x.insert(END, str(int(self.cal_center_x.get())))
        labx.grid(row=1,column=0)
        self.cal_x.grid(row=1,column=1)
        
        
        self.cal_y=Entry(self.cal_win)
        laby = Label(self.cal_win, width=5, text="   Center X  : ", anchor='w')
        self.cal_y.insert(END, str(int(self.cal_center_y.get())))
        laby.grid(row=2,column=0)
        self.cal_y.grid(row=2,column=1)
        
        
        
        self.guide_line_button = Button(self.cal_win, text="Apply", command = lambda : self.apply_calibration())
        self.guide_line_button.grid(row=0,column=2)

        button3 = Button(self.cal_win, text="Quit", command = self.cal_win.destroy)
        button3.grid(row=1,column=2)
        
        
    def apply_calibration(self):
        
        self.cal_center_x.set(int(self.cal_x.get()))
        self.cal_center_y.set(int(self.cal_y.get()))
        self.calibration=int(self.cal_radius.get())
        self.calibration_scale=450/self.calibration 
        np.save(self.pullback_dir + '/Calibration.npy',self.calibration)
        np.save(self.pullback_dir + '/CalibrationCenter.npy',(self.cal_center_x.get(),self.cal_center_y.get()))
        self.current_section_canvas.show_calibration()
        #self.current_section_canvas.update()

    def accept_guide(self):
        self.tip=Toplevel()
        self.tip.title("Guide validation")
        #msg = Message(self.top, text="select thr")
        #msg.pack()
        self.input_guide_thr=Entry(self.tip)
        lab = Label(self.tip, width=5, text="   cut  : ", anchor='w')
        self.input_guide_thr.insert(END, '200')
        self.guide_canvas = Canvas(self.tip, width=self.Nr+160, height=self.Nf,  bg='black')
        self.guide_canvas.bind("<Button-1>", self.fix_point_guide)
        self.guide_canvas.create_image(0, 0, image=[], anchor='nw', tag='gradients')
        lab.pack(side=LEFT)
        self.input_guide_thr.pack(side=LEFT)
        self.guide_canvas.pack(side=TOP)
        self.make_guide_gradients()
        button3 = Button(self.tip, text="Quit", command = self.tip.destroy)
        button3.pack(side=BOTTOM)
        self.sel_guide_line=0
        self.guide_line_button = Button(self.tip, text="Right side line selected", command = lambda : self.switch_line())
        self.guide_line_button.pack(side=BOTTOM)
        button2 = Button(self.tip, text="Calculate Guide", command= lambda: self.analyze_guide())
        button2.pack(side=BOTTOM)
        button1 = Button(self.tip, text="Adapt gradients", command = lambda: self.insert_guide_thr_value())
        button1.pack(side=BOTTOM)
        jumps, jumps_f = self.jump_and_wine()
        for i in range(self.Nf):
            for j in range(self.Nr):
                if (jumps[i,j]==1):
                    self.guide_canvas.create_oval(j-3,i-3,j+3,i+3, tags=("jumps"), fill='white')
                if (jumps_f[i,j]==1):
                    self.guide_canvas.create_oval(j-3,i-3,j+3,i+3, tags=("jumps_F"), fill='blue')

    def make_guide_gradients(self):
        a=100
        nb=20
        
        I0= np.sum(self.pullback[:,:,int(self.guide_thr):],axis=2)
        I=np.concatenate([I0[:,-a:],I0,I0[:,0:a]],axis=1)
    
        G=self.current_section_canvas.intensities_grad(I)
        G=G[:,nb:-nb]
        
        cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.linear_bw(), N=number_of_colors)
        #img = Image.fromarray(cmap( np.maximum( I0/np.max(I0),self.guide_lines ) , bytes=True))#I0/np.max(I0), bytes=True))#G
        img = Image.fromarray(cmap( np.maximum( I0/np.max(I0),self.guide_lines ) , bytes=True))#I0/np.max(I0), bytes=True))#G
        self.photoImage = ImageTk.PhotoImage(img) 
        self.guide_canvas.itemconfig('gradients', image=self.photoImage)

    
    def insert_guide_thr_value(self):
        self.guide_thr=self.input_guide_thr.get()
        self.make_guide_gradients()
        
    def switch_line(self):
        self.sel_guide_line=(self.sel_guide_line+1)%2
        side=["Right","Left"]
        self.guide_line_button.config(text=side[self.sel_guide_line] + " side line selected")
        
    def fix_point_guide(self,event):
        x=event.x
        y=event.y
        #print(x,y)
        self.guide_fix[self.sel_guide_line,y,:]=1
        self.guide_fix[self.sel_guide_line,y,x]=0
        col=["blue","green"]
        self.guide_canvas.create_oval(x-3,y-3,x+3,y+3, tags=("guide_points"), fill=col[self.sel_guide_line])        
    
        '''def play_stop(self):
            
            self.timeline=sched.scheduler(time.time(),time.sleep())
            self.timeline.enter(self, delay, priority, action, argument, kwargs)
        '''    
        
    def jump_and_wine(self):

        s=4
        h=15
        d=0.3
        l_min=80#90
        I_thr=20#30
        T = time.time()
        #G=self.pullback
        I0=self.pullback
        I=np.concatenate([I0[:,:,-h:],I0,I0[:,:,0:h]],axis=2)
        '''
        C=I
        C[:,:,:]=0
        C[I<30]=1
        i0=np.sum(C)
        C[:,:,:]=0
        C[ I<50]=1
        C[I<30]=0
        i1=np.sum(C)
        C[:,:,:]=0
        C[I<70]=1
        C[I<50]=0
        i2=np.sum(C)
        C[:,:,:]=0
        C[I<100]=1
        C[I<70]=0
        i3=np.sum(C)
        C[:,:,:]=0
        C[I>100]=1
        i3=np.sum(C)
        print('intensities',i0,i1,i2,i3)
        '''
        I[I<I_thr]=0
        
        G=np.zeros(self.pullback.shape)#self.parent.pullback.shape[1:])

        K=np.arange(2*h+1)-h
        K=-1/(np.sqrt(2*np.pi)*s*s*s) * np.multiply(K , np.exp(-1/2*np.multiply(K,K)/(s*s)))
        for t in range(G.shape[0]):
            for z in range(G.shape[1]):
                G[t,z,:]=np.convolve(I[t,z,:],K,'valid')
        #G=G[:,:,h:-h]
        #G=G-np.min(G)
        #G=G/np.max(G)
        #G=-G
        print('end gradients', (time.time() - T) / 60)

        d=0.3#2.5#0.5#3#1.5
        G=-G
        G[G<d]=0#-G[G<0]
        #G=I0
        #G[G<60]=0
        Gsum= np.sum(G[:,:,l_min:],axis=2)
        print('end sums', (time.time() - T) / 60)

        P = np.arange(0, self.Np)
        P[:l_min] = 0
        E = np.dot(G,P)
        L = np.divide(E,Gsum)
        #L[np.isnan(L)] = 0
        mask = np.isnan(L)
        L[mask] = np.interp(np.flatnonzero(mask), np.flatnonzero(~mask), L[~mask])

        P = np.multiply(P,P)
        V = np.dot(G,P)
        V = np.divide(V,Gsum)
        V = V - np.multiply(L,L)
        #V[np.isnan(V)] = 1
        #print('min',np.min(V,axis=0))
        #V[V<1] = 1
        mask = np.isnan(V)
        V[mask] = np.interp(np.flatnonzero(mask), np.flatnonzero(~mask), V[~mask])

        print(V.shape,np.max(V),np.min(V),np.mean(V))
        plt.imshow(V)
        plt.show()
        
        V=np.log2(V)
        plt.imshow(V)
        plt.show()
        #V=V-np.min(V)
        #V=V/np.max(V)
        V=V-np.mean(V)
        V=V/np.std(V)
        V[V<-1]=-1
        V[V>1]=1
        V=(V+1)/2
        plt.imshow(V)
        plt.show()
        print(V.shape, np.max(V), np.min(V), np.mean(V))
        cmap = mpl.colors.LinearSegmentedColormap.from_list([], self.linear_bw(), N=number_of_colors)
        
        #img = Image.fromarray(cmap( np.maximum( I0/np.max(I0),self.guide_lines ) , bytes=True))#I0/np.max(I0), bytes=True))#G
        img = Image.fromarray(cmap( V, bytes=True))#I0/np.max(I0), bytes=True))#G
        self.photoImage = ImageTk.PhotoImage(img)
        self.guide_canvas.itemconfig('gradients', image=self.photoImage)
        '''
        L = np.zeros(G.shape[:2])

        for i in range(0,self.Nf):
            for j in range(0,self.Nr):
                e=0
                P=np.arange(0,self.Np)
                P[:l_min]=0
                e=np.dot(P,G[i,j,:])
                #print(type(e),e.size,e)
                #for k in range(l_min,self.Np):
                #    e+=k*G[i,j,k]
                if Gsum[i,j]:
                    e/=Gsum[i,j]
                e=int(e)
                L[i,j]=e
        #print('lumen bound',np.max(L),np.min(L))
        '''


        print('end E', (time.time() - T) / 60)

        self.lumen=L
        L_old=L
        self.current_section_canvas.update()
        h=7
        s=7
        J=np.zeros(L.shape)
        for i in range(h):
            L=np.concatenate([np.reshape(L[0,:],[1,L.shape[1]]) , L ,np.reshape(L[-1,:],[-1,L.shape[1]])],axis=0)
        K=np.arange(2*h+1)-h
        K=-1/(np.sqrt(2*np.pi)*s*s*s) * np.multiply(K , np.exp(-1/2*np.multiply(K,K)/(s*s)))
        for t in range(J.shape[1]):
            J[:,t]=np.convolve(L[:,t],K,'valid')
            #if t%100==0:
            #    print(L[:,t],J[:,t])
        #J=J[-h-h,:]
        print('end jumps', (time.time() - T) / 60)

        print(np.max(J),np.min(J))
        Jumps=J
        delta=3
        Jumps [np.abs(Jumps)<delta]=0
        Jumps [np.abs(Jumps)>delta]=1
        self.labels[Jumps==1]=400
        self.prediction_canvas.update(0)

        h=7
        s=7
        J=np.zeros(L_old.shape)
        L=np.concatenate([L_old[:,:h] , L_old ,L_old[:,-h:]],axis=1)
        K=np.arange(2*h+1)-h
        K=-1/(np.sqrt(2*np.pi)*s*s*s) * np.multiply(K , np.exp(-1/2*np.multiply(K,K)/(s*s)))
        for t in range(J.shape[0]):
            J[t,:]=np.convolve(L[t,:],K,'valid')
        
        print('end jumps', (time.time() - T) / 60)
        print(np.max(J),np.min(J))
        Jumps_f=J
        delta=1.5
        Jumps_f [np.abs(Jumps_f)<delta]=0
        Jumps_f [np.abs(Jumps_f)>delta]=1
        self.labels[Jumps_f==1]=200
        self.prediction_canvas.update(0)

        
        return Jumps, Jumps_f
        
    def start_play(self):
        #print(self.current_section_canvas.index.get(),time.time())
        
        global play
        if play and self.current_section_canvas.index.get()<self.Nf:
            self.increase_section_index()
            self.update()
            #self.after(4000,self.start_play())
            self.start_play()
            
    def play_pullback(self):
        global play
        play=True
        self.start_play()
    
    def stop_pullback(self):
        global play
        play=False
    
    def add_bad(self):
        self.bad.append(0+permutation[self.current_section_canvas.index.get()-0])
        save_list(self.bad,model_file + '/bad.txt')
        print("adding  "+str(0+permutation[self.current_section_canvas.index.get()-0]))
    
        
    def add_good(self):
        self.good.append(0+permutation[self.current_section_canvas.index.get()-0])
        save_list(self.good,model_file + '/good.txt')
        print("adding  "+str(0+permutation[self.current_section_canvas.index.get()-0]))
    
    
    def increase_spicchio_index(self):
        if(self.current_section_canvas.spicchio.get()<(self.Nr-DEG)/STRIDE):
            #print(self.current_section_canvas.spicchio.get())
            self.current_section_canvas.spicchio.set(self.current_section_canvas.spicchio.get()+1)
        #self.current_section_canvas.update_circle_arc()

    def decrease_spicchio_index(self):
        if(self.current_section_canvas.spicchio.get()>0):
            #print(self.current_section_canvas.spicchio.get())
            self.current_section_canvas.spicchio.set(self.current_section_canvas.spicchio.get()-1)
        #self.current_section_canvas.update_circle_arc()
    
    def increase_radius_index(self):
        #if self.current_section_canvas.radius_index.get()<self.Nr : 
        self.current_section_canvas.radius_index.set(self.current_section_canvas.radius_index.get()+1)
        #self.current_section_canvas.update()
        #self.prediction_canvas.update_radius()
        self.current_section_canvas.update_line(self.current_section_canvas.radius_index.get())

        if self.prediction_canvas.show_longitudinal:
            self.prediction_canvas.update(0)

    def decrease_radius_index(self):
        #if self.current_section_canvas.radius_index.get()>0 : 
        self.current_section_canvas.radius_index.set(self.current_section_canvas.radius_index.get()-1)
        #self.annotation_canvas.coords('result_canvas_radius_line',  self.parent.current_section_canvas.radius_index.get(), 0, self.parent.current_section_canvas.radius_index.get(), self.parent.labels.shape[0])
        #self.prediction_canvas.coords('result_canvas_radius_line',  self.parent.current_section_canvas.radius_index.get(), 0, self.parent.current_section_canvas.radius_index.get(), self.parent.labels.shape[0])
        #self.prediction_canvas.update_radius()
        self.current_section_canvas.update_line(self.current_section_canvas.radius_index.get())

        if self.prediction_canvas.show_longitudinal:
            self.prediction_canvas.update(0)

    def increase_section_index(self):
        if self.current_section_canvas.index.get()<self.Nf :
            self.current_section_canvas.index.set(self.current_section_canvas.index.get()+1)
            self.info_canvas.coords('result_canvas_section_line',  self.current_section_canvas.index.get()*self.ann_scale_w, 0,  self.current_section_canvas.index.get()*self.ann_scale_w, self.ann_canvas_h//2)
            self.annotation_canvas.coords('result_canvas_section_line',  self.current_section_canvas.index.get()*self.ann_scale_w, 0,  self.current_section_canvas.index.get()*self.ann_scale_w, self.ann_canvas_h//2)
            self.prediction_canvas.coords('result_canvas_section_line',  self.current_section_canvas.index.get()*self.ann_scale_w, 0,  self.current_section_canvas.index.get()*self.ann_scale_w, self.ann_canvas_h)
            self.current_section_canvas.old_intima=[np.copy(self.lumen)]
            self.current_section_canvas.old_fc=[np.copy(self.fc)]
            self.current_section_canvas.old_fct=[np.copy(self.fct)]
            
            self.current_section_canvas.fixed_points=[]
            self.current_section_canvas.fixed_points_coord=[]
            self.selected_arcs=[]
            self.old_selected_arcs=[np.copy(self.current_section_canvas.selected_arcs)]
            self.old_fixed_points=[np.copy(self.current_section_canvas.fixed_points)]

            self.current_section_canvas.fixed_points_fc=[]
            self.current_section_canvas.old_fixed_points_fc=[np.copy(self.current_section_canvas.fixed_points_fc)]
            self.current_section_canvas.fixed_points_coord_fc=[]
            self.current_section_canvas.selected_arcs_fc=[]
            self.current_section_canvas.old_selected_arcs_fc=[np.copy(self.current_section_canvas.selected_arcs_fc)]
            
            self.current_section_canvas.fct=[0,np.nan,np.nan]
            self.current_section_canvas.plaque_arcs=[]
            self.current_section_canvas.shots.set(0)
            if self.current_section_canvas.zoom_on.get():  
                self.current_section_canvas.zommed_immage()
            self.current_section_canvas.update()
            #self.current_section_canvas.update_circle_arc()
            self.measure_canvas.del_diameters()

    def decrease_section_index(self):
        if self.current_section_canvas.index.get()>0 :
            self.current_section_canvas.index.set(self.current_section_canvas.index.get()-1)
            self.info_canvas.coords('result_canvas_section_line',  self.current_section_canvas.index.get()*self.ann_scale_w, 0,  self.current_section_canvas.index.get()*self.ann_scale_w, self.ann_canvas_h//2)
            self.annotation_canvas.coords('result_canvas_section_line',  self.current_section_canvas.index.get()*self.ann_scale_w, 0,  self.current_section_canvas.index.get()*self.ann_scale_w, self.ann_canvas_h//2)
            self.prediction_canvas.coords('result_canvas_section_line',  self.current_section_canvas.index.get()*self.ann_scale_w, 0,  self.current_section_canvas.index.get()*self.ann_scale_w, self.ann_canvas_h)
            self.current_section_canvas.old_intima=[np.copy(self.lumen)]
            self.current_section_canvas.old_fc=[np.copy(self.fc)]
            self.current_section_canvas.old_fct=[np.copy(self.fct)]
            self.current_section_canvas.fixed_points=[]
            self.current_section_canvas.old_fixed_points=[np.copy(self.current_section_canvas.fixed_points)]
            self.current_section_canvas.fixed_points_coord=[]
            self.current_section_canvas.selected_arcs=[]
            self.current_section_canvas.old_selected_arcs=[np.copy(self.current_section_canvas.selected_arcs)]
            
            self.current_section_canvas.fixed_points_fc=[]
            self.current_section_canvas.old_fixed_points_fc=[np.copy(self.current_section_canvas.fixed_points_fc)]
            self.current_section_canvas.fixed_points_coord_fc=[]
            self.current_section_canvas.selected_arcs_fc=[]
            self.current_section_canvas.old_selected_arcs_fc=[np.copy(self.current_section_canvas.selected_arcs_fc)]
            
            self.current_section_canvas.fct=[0,np.nan,np.nan]
            self.current_section_canvas.plaque_arcs=[]
            self.current_section_canvas.shots.set(0)
            if self.current_section_canvas.zoom_on.get():  
                self.current_section_canvas.zommed_immage()
            self.current_section_canvas.update()
            #self.current_section_canvas.update_circle_arc()
            self.measure_canvas.del_diameters()
            
    def increase_section_zoom(self):
        if self.current_section_canvas.zoom.get()<1000 :
            self.current_section_canvas.zoom.set(self.current_section_canvas.zoom.get()+100)
            self.current_section_canvas.update()
            #self.current_section_canvas.update_circle_arc()

    def decrease_section_zoom(self): 
        if self.current_section_canvas.zoom.get()>100 :
            self.current_section_canvas.zoom.set(self.current_section_canvas.zoom.get()-100)
            self.current_section_canvas.update()
            #self.current_section_canvas.update_circle_arc()

    def load_labels_classification (self,thr=0.65,thr_ca=0.4 ):
        
        #calcium_annotation=all_calcium_annotations[0:self.Nf,:]
        self.labels=np.zeros([self.Nf, self.Nr],dtype=np.uint16)
        #self.labels = self.labels.astype(np.uint16)
        self.labels[self.lipid == 0] = non_lipid_color_index
        self.labels[self.lipid == 1] = lipid_color_index
        self.labels[self.calcium==1]=calcium_color_index
        self.labels[self.guide==1]=unknown_color_index
        #self.mark_deep_plaque()
        #self.prediction_canvas.update(0)
        #self.mark_deep_plaque_n()


        self.annotation[self.annotation == 0] = unknown_color_index
        self.annotation[self.annotation == 1] = non_lipid_color_index
        
        return self.annotation

    def mark_deep_plaque(self, depth=100 ):
        #self.dplaq=np.zeros([self.Nf, self.Nr],dtype=np.uint16)
        #self.labels[(((self.fc-self.lumen)* 0.88 / 196 * 1000) >depth) & (self.lipid == 1)] = deep_plaque_color_index
        mfct=np.zeros([self.Nf,self.Nr])
        fct=np.array([round(self.fct[i][0] * self.calibration_scale) for i in range (self.Nf)])
        for i in range (self.Nr):
            mfct[:,i]=fct
        self.labels[(mfct >depth) & (self.lipid == 1)] = deep_plaque_color_index
        #print(fct)
        #print(self.fc,(self.fc-self.lumen)* 0.88 / 196*1000)

    def mark_deep_plaque_n(self, depth=100, df=2, dt=15 ):
        thin_fc=np.zeros([self.Nf,self.Nr])
        #for f in range(self.Nf):
        #    self.current_section_canvas.calculate_fct_cart(f)
        #print(self.all_fct)
        all_fct=self.all_fct * self.calibration_scale
        all_fct[self.all_fct==0]=1000
        for f in  range(self.Nf):
            for t in range (self.Nr):
                if self.lipid[f,t]==1 and all_fct[f,t]<depth :
                    if t-dt>0 and t +dt <self.Nr:
                        thin_fc[np.max([0,f-df]):np.min([self.Nf,f+df]),t-dt:t+dt]=1
                    elif t-dt<0:
                        thin_fc[np.max([0,f-df]):np.min([self.Nf,f+df]),(t-dt)%self.Nr:self.Nr]=1
                        thin_fc[np.max([0,f-df]):np.min([self.Nf,f+df]),0:t+dt]=1
                    elif t+dt>self.Nr:
                        thin_fc[np.max([0,f-df]):np.min([self.Nf,f+df]),t-dt:self.Nr]=1
                        thin_fc[np.max([0,f-df]):np.min([self.Nf,f+df]),0:(t+dt)%self.Nr]=1 
        #print('thin fc', sum(sum(thin_fc==1))/(self.Nf*self.Nr), sum(sum(self.lipid == 1))/(self.Nf*self.Nr),  sum(sum((thin_fc==1) & (self.lipid == 1)))/sum(sum(self.lipid==1)))
        self.labels[( ~(thin_fc==1)) & (self.lipid == 1)] = deep_plaque_color_index
            
    def load_or_create_labels(self,path,i):
        
        if not os.path.exists(path):
            np.save(path,np.zeros([self.Nf, self.Nr],dtype=np.int))
        else:
            self.flags[i]=1
        return np.load(path,mmap_mode='r+')
        #if not os.path.exists():
        #    os.makedirs(directory)99
        
    def import_new_pullback(self):
        #print('lumen type',type(self.lumen))
        self.import_pullback_callback()
        #print('lumen type',type(self.lumen))
        
    def import_pullback_callback(self,data_path=None):

        if data_path:
            self.pullback_dir = data_path
            for file in os.listdir(data_path):
                if '.oct'in file :
                    self.pullback_path=os.path.join(data_path , file)
                    
        else:
            #self.pullback_path = '/home/eugenio/workspace/documets di pasqualino/Documents/OCTdatabase/gui test 12/pullback.oct'
            #self.pullback_path = '/media/eugenio/Hitachi/Pazienti CLIMA/Prof. Ozaki/FHU LAD 62/PB1/{FAFEC34B-20AA-4AA1-84C2-8E4B0A7A3BE2}.oct'
            self.pullback_path = askopenfilename(initialdir='/media/eugenio/Hitachi',#"/home/eugenio/workspace/documets di pasqualino/Documents/OCTdatabase",#'/media/fisici/Hitachi/Pazienti CLIMA',
                                                 defaultextension='.oct', filetypes=[('OCT files', '*.oct')])
            self.pullback_dir= os.path.dirname(os.path.abspath(self.pullback_path))
        
        if os.path.exists(self.pullback_dir+"/pullback.npy"):
            self.pullback=np.load(self.pullback_dir+"/pullback.npy")#[0:350,:,:]
        else :
            self.pullback = skimage.external.tifffile.imread(self.pullback_path) #io
            np.save(self.pullback_dir+"/pullback.npy",self.pullback)
        self.pullback_path=self.pullback_dir+"/pullback.npy"
        self.patient_name=os.path.dirname(self.pullback_dir)
        self.Nf=self.pullback.shape[0]
        self.Nr=self.pullback.shape[1]
        self.Np=self.pullback.shape[2]

        self.ann_scale_h=self.ann_canvas_h/self.Nr
        self.ann_scale_w=self.ann_canvas_w/self.Nf

        self.ff=0
        self.lf=self.Nf

        self.interv_of_pertinence=[(0,self.Nf)]
        self.guide_path= self.pullback_dir+"/GuideArtifact.npy"
        self.lumen_path=self.pullback_dir+"/Lumen.npy"
        self.lipid_path=self.pullback_dir+"/LipidPlaque.npy"
        self.calcium_path=self.pullback_dir+"/CalciumPlaque.npy"
        self.fc_path=self.pullback_dir+"/FibrusCap.npy"
        self.all_fct_path=self.pullback_dir+"/FCT.npy"
        self.annotation_path = self.pullback_dir + "/annotation.npy"

        self.flags=[0,0,0,0,0]

        self.guide=self.load_or_create_labels(self.guide_path,0)
        self.lumen=self.load_or_create_labels(self.lumen_path,1)
        self.lipid=self.load_or_create_labels(self.lipid_path,2)
        self.calcium=self.load_or_create_labels(self.calcium_path,4)
        self.fc=self.load_or_create_labels(self.fc_path,3)
        self.all_fct=self.load_or_create_labels(self.all_fct_path,4)
        
        self.guide_arcs= [ [] for i in range(self.Nf) ]
        self.plaque_arcs= [ [] for i in range(self.Nf) ]
        self.fct=[ [0,np.nan,np.nan] for i in range(self.Nf) ]

        self.measure_points=[ [ ] for i in range(self.Nf) ]
        self.measure_points_cart = [[] for i in range (self.Nf) ]

        self.plaque_angle=np.zeros(self.Nf)
        self.areas=np.zeros(self.Nf)
        self.diameters=np.zeros([2,self.Nf,3])
        self.annotation=self.load_or_create_labels(self.annotation_path,4)
        self.annotation=self.load_labels_classification ()#self.preds)
        
        self.guide_thr=200
        self.guide_fix=np.zeros((2,self.Nf,self.Nr))
        self.guide_lines= np.zeros((self.Nf,self.Nr)) 
        
        self.macrophages=[]
        if os.path.exists(self.pullback_dir+"/macrophages.ixt"):
            self.macrophages = load_list(self.pullback_dir+'/macrophages.txt')
        print(self.pullback.shape)

        
        self.v_stop=IntVar(master=self , value=0)
        #self.current_section_canvas.update()
        #self.annotation_canvas.update()
        #self.prediction_canvas.update()
        
        self.current_section_canvas.zoom = IntVar(master=self, value=self.Np)#self.parent.Np)
        self.current_section_canvas.old_intima = [np.copy(self.lumen)]
        self.current_section_canvas.old_fc = [np.copy(self.fc)]
        self.current_section_canvas.old_fct = [np.copy(self.fct)]
        
        self.scale_canvas.i1=self.ann_canvas_w
        #self.current_section_canvas.update()

        #self.preprocessing()
        self.scale_canvas.delete('triangles')

        self.calibration=0
        self.calibration_scale=0.88 / 196 * 1000
        self.cal_center_x= IntVar(master=self, value=0)
        self.cal_center_y= IntVar(master=self, value=0)
        #self.calibrate()

        self.load_preanal()


        self.warnings()
        self.current_section_canvas.show_calibration()

        #self.mark_deep_plaque_n()
        
        self.measure_canvas.print_diameters()
        self.current_section_canvas.update()
        for widget in self.scale_canvas.top.grid_slaves():
            widget.grid_forget()
        self.scale_canvas.add_interval()
        self.scale_canvas.refresh_measure()
        self.scale_canvas.top.withdraw()
        
        ''' per fare controlli sul fc
        n=218
        print(self.all_fct[n])
        m=[(self.lipid==0)|(self.all_fct<1)]
        #m=[self.all_fct==0]
        m=[self.lipid==0]
        #self.current_section_canvas.index.set()
        all_fctm=np.ma.array(self.all_fct , mask=m)
        print(all_fctm[n],all_fctm.min(axis=1)[n])
        print(all_fctm.argmin(axis=1)[n],self.fct[n])
        '''
        #self.labels[(self.all_fct>0) & (self.all_fct<1000)]=unknown_color_index

        #self.current_section_canvas.calculate_plaque_angle(128)

        #print(self.pullback[179, 100,:])
        #print(self.pullback[179, 200, :])
        #print(self.pullback[179, 300, :])
        '''
        for inter in self.interv_of_pertinence:
            for f in range(inter[0],inter[1]):
                self.current_section_canvas.calculate_area(f)
                np.save(self.pullback_dir +'/Area.npy',self.areas)
        '''
        #self.write_results()
    
    def reset_zoom(self):
        self.current_section_canvas.measure_on.set(False)
        self.current_class.set(0)

        if not self.current_section_canvas.zoom_on.get():
            self.current_section_canvas.zoom_now.set(1)
            self.current_section_canvas.zoomed_vertex=[0,0] 
            self.current_section_canvas.update()
            
    def measure_this(self):
        self.current_class.set(0)
        #if not self.current_section_canvas.zoom_on.get():
            #print("ciao")

    def analyze_set(self):
        main_dir = askdirectory()
        self.walk_down(main_dir)
    
    def walk_down(self,dir): 
        
        if next(os.walk(dir))[1] :
            for folder in next(os.walk(dir))[1]:
                folder_path=os.path.join(dir,folder)
                self.walk_down( folder_path )
        else:
            n=0
            for file in os.listdir(dir):
                print(dir[-3:-1])
                if '.oct' in file :
                    n+=1
                    if (not dir[-3:-1]=='PB'):
                        data_path=os.path.join(dir,'PB'+str(n))
                        os.mkdir(data_path)
                        shutil.move(os.path.join(dir,file) , os.path.join(data_path,file))
                    else:
                        try:
                            self.import_pullback_callback(dir)
                            self.analyze_all()
                        except:
                            pass
                            
            if (not dir[-3:-1]=='PB') and n>0 :
                self.walk_down( dir )
                            
    def analyze_all(self):
        self.flags = [0, 0, 0, 0, 0]
        if os.path.exists(self.pullback_dir + "/Prediction.npy"):
            os.remove(self.pullback_dir + "/Prediction.npy")
        self.preprocessing()
        self.current_section_canvas.update()


    def analyze_guide(self):

        I,lines=self.current_section_canvas.guide_boundary_double(0,self.Nf)
        save_list(self.guide_arcs,self.pullback_dir +'/GuideArcs.txt')
        self.guide_arcs=load_list(self.pullback_dir +'/GuideArcs.txt')
        #self.guide = self.guide.astype(np.int)
        self.current_section_canvas.update()
        #self.accept_guide()
        #self.wait_window(self.tip)
        #print("ho aspettato")


    def analyze_lumen(self, b):

        d=self.current_section_canvas.lumen_weights()
        if b :
            self.current_section_canvas.lumen_ds(self.current_section_canvas.index.get(), single_frame=True)
            #self.current_section_canvas.boundary(self.current_section_canvas.index.get(),d)
            self.current_section_canvas.regulirize_intima(10, self.Nr)
            self.current_section_canvas.calculate_area(self.current_section_canvas.index.get())
            np.save(self.pullback_dir +'/Area.npy',self.areas)
            self.current_section_canvas.update()
            #self.analyze_lipid(b)
            #self.analyze_fc(b)
            self.current_class.set(4)

        else :
           # self.current_section_canvas.boundary(0,d)
           
            #for f in range(0,self.Nf):
            for inter in self.interv_of_pertinence:
                for f in range(inter[0],inter[1]):
                
                    #self.current_section_canvas.lumen_boundary_c(f,d)
                    self.current_section_canvas.lumen_ds(f)
                    self.current_section_canvas.index.set(f)
                    #self.current_section_canvas.regulirize_intima(10, self.Nr)
                    self.current_section_canvas.calculate_area(f)
            self.current_section_canvas.index.set(0)

            np.save(self.pullback_dir +'/Area.npy',self.areas)
            self.current_section_canvas.update() 
                    
        
    def analyze_lipid(self,b):
        

        if b:
            self.identify_plaque_frame()
            self.current_section_canvas.calculate_plaque_angle(self.current_section_canvas.index.get())
            self.analyze_fc(b)
            self.current_section_canvas.update()
            self.current_class.set(1)

        else:
            if os.path.exists(self.pullback_dir + "/Prediction.npy"):
                os.remove( self.pullback_dir  +  "/Prediction.npy")
            self.identify_plaque()
            for f in range(self.Nf):
                self.current_section_canvas.calculate_plaque_angle(f)
        save_list(self.plaque_arcs,self.pullback_dir +'/LipidArcs.txt')
        np.save(self.pullback_dir +'/PlaqueAngle.npy',self.plaque_angle)
        self.current_section_canvas.update()

            
    def analyze_fc (self,b):

        if b:
            d=self.current_section_canvas.fc_weights()
            self.current_section_canvas.fc_boundary(self.current_section_canvas.index.get(),d)
            save_list(self.fct,self.pullback_dir +'/Fct.txt')
            self.all_fct = self.all_fct.astype(np.int)
            np.save(self.all_fct_path,self.all_fct)
        
            self.current_class.set(5)

            #np.save(self.pullback_dir+"/FibrusCap.npy",self.fc)

        else:
            
            d=self.current_section_canvas.fc_weights()
            #for f in range(self.Nf):
            for inter in self.interv_of_pertinence:
                for f in range(inter[0],inter[1]):
                    self.current_section_canvas.fc_boundary(f,d)
            save_list(self.fct,self.pullback_dir +'/Fct.txt')
            self.all_fct = self.all_fct.astype(np.int)
            np.save(self.all_fct_path,self.all_fct)
        
            #np.save(self.pullback_dir+"/FibrusCap.npy",self.fc)
        self.current_section_canvas.update()

    def analyze_macrophages(self):
        
        l=15
        span=2*l+10

        d=300#14.2
        L=200
        self.macrophages=[]#np.zeros(self.pullback.shape,dtype='uint8')
        self.mph_frame_index=[]
        #first_on_fr=0
        #last_on_fr=0
        for f in range(self.current_section_canvas.index.get(),self.current_section_canvas.index.get()+1):#
            img=np.copy(self.pullback[f])
            for i in range(self.Nr):
                if self.guide[f,i]==1:
                    img[i,:]=0
                else:
                    img[i,:self.lumen[f,i]]=0
            s_min=np.min(img[np.nonzero(img)])#not img[f]==0])
            s_max=np.max(img[np.nonzero(img)])#not img[f]==0])
            img=(img-s_min)/(s_max-s_min)        
            for i in range(2*l,self.Nr-2*l):
                for j in range(self.lumen[f,i]+l,np.min([self.lumen[f,i]+L , self.Np-l-span-1] )):
                    
                    Q1=img[i-l:i+l+1, j-l:j+l+1]#self.pullback[f, i-l:i+l+1, j-l:j+l+1]
                    Q2=img[i-l:i+l+1, j+span-l:j+span+l+1]#self.pullback[f, i-l:i+l+1, j+span-l:j+span+l+1]
                    N=Q1.size
                    '''
                    Q1=Q1.flatten()
                    Q2=Q2.flatten()
                    E1 = np.sum(Q1)
                    E2 = np.sum(Q2)
                    
                    #V1 = np.dot(Q1,Q1)
                    #V2 = np.dot(Q2,Q2)
                    V1=np.multiply(E1,E1)
                    V1=np.sum(V1)
                    V2=np.multiply(E2,E2)
                    V2=np.sum(V2)

                    s1=np.sqrt(V1*N-E1*E1)/N
                    s2=np.sqrt(V2*N-E2*E2)/N
                    '''
                    s1=np.std(Q1[np.nonzero(Q1)])
                    s2=np.std(Q2[np.nonzero(Q2)])
                    NSD=np.sqrt(N/(N-1))*s1
                    NSDR=s1/s2
                    print(NSDR,NSD)
                    if NSDR > d and NSD>0.024:
                        self.macrophages.append((f,i,j))
                        self.labels[f,i]=unknown_color_index
            #self.mph_frame_index
        
        save_list(self.macrophages, self.pullback_dir+'/macrophages.txt')
        print(len(self.macrophages))
        self.prediction_canvas.update(0)
        self.current_section_canvas.build_macrophages()
        print('factotum')

    def analyze_macrophages_old(self):

        l = 15
        span = 2 * l + 10

        d = 14.2
        L = 200
        self.macrophages = []  # np.zeros(self.pullback.shape,dtype='uint8')
        self.mph_frame_index = []
        # first_on_fr=0
        # last_on_fr=0
        for f in range(self.current_section_canvas.index.get(), self.current_section_canvas.index.get() + 1):  #
            img = np.copy(self.pullback[f])
            for i in range(self.Nr):
                if self.guide[f, i] == 1:
                    img[i, :] = 0
                else:
                    img[i, :self.lumen[f, i]] = 0
            s_min = np.min(img[np.nonzero(img)])  # not img[f]==0])
            s_max = np.max(img[np.nonzero(img)])  # not img[f]==0])
            img = (img - s_min) / (s_max - s_min)
            for i in range(2 * l, self.Nr - 2 * l, 3):
                for j in range(self.lumen[f, i] +  l, np.min([self.lumen[f, i] + L, self.Np - l - span - 1]),5):
                    e1=0
                    s1=0

                    e2=0
                    s2=0
                    N=4*l*l
                    e1=np.mean(img[i - l:i + l + 1, j - l:j + l + 1])
                    e2 = np.mean(img[i - l:i + l + 1, j + span - l:j + span + l + 1])
                    '''for t in range(i - l,i + l + 1):
                        for r in range(j - l,j + l + 1):
                            e1+=img[t,r]
                        for r in range(j + span - l,j + span + l + 1):
                            e2 += img[t, r]
                    e1/=N
                    e2/=N'''
                    for t in range(i - l, i + l + 1):

                        for r in range(j - l, j + l + 1):
                            s1 += np.power(img[t, r]-e1,2)
                        for r in range(j + span - l, j + span + l + 1):
                            s2 += np.power(img[t, r]-e2,2)
                    s1/=N-1
                    s2/=N-1
                    s1 = np.sqrt(s1)
                    s2 = np.sqrt(s2)
                    NSD=s1/(s_max-s_min)

                    print(s1 / s2,NSD)
                    if s1 / s2 > d and  NSD > 0.024:
                        self.macrophages.append((f, i, j))
                        self.labels[f, i] = unknown_color_index
            # self.mph_frame_index

        save_list(self.macrophages, self.pullback_dir + '/macrophages.txt')
        print(len(self.macrophages))
        self.prediction_canvas.update(0)
        self.current_section_canvas.build_macrophages()
        print('factotum')

    def analyze_LBI(self,superf=0, depth=20, thr=400 ):
        if self.Nf>450:
            depth=40
        print('LBI depth',depth)
        if superf:
            self.lbis=np.zeros(self.Nf)
        else:
            self.lbi=np.zeros(self.Nf)
        for interv in self.interv_of_pertinence:
            d=interv[1]-interv[0]

            #self.lbi=np.zeros(d-depth)
            for i in range(d-depth):
                if superf:
                    self.lbis[i+interv[0]] = self.LBI_superf(i+interv[0],i+interv[0]+depth)
                else:
                    self.lbi[i+interv[0]] = self.LBI(i+interv[0],i+interv[0]+depth )
        if superf :
            np.save(self.pullback_dir + '/LBIs.npy',self.lbis)
        else:
            np.save(self.pullback_dir + '/LBI.npy',self.lbi)
        self.mlbi=np.max(self.lbi)
        
        self.prediction_canvas.delete('LBI_start_line')
        self.prediction_canvas.delete('LBI_end_line')
        where=np.argmax(self.lbi)
        print('max lbi',self.mlbi,' at frame',where)
        start=where*self.ann_scale_w
        end=(where+depth)*self.ann_scale_w
        self.prediction_canvas.create_line(start, 0, start, self.ann_canvas_h, fill='cyan2', width=1, tag='LBI_start_line')
        self.prediction_canvas.create_line(end, 0, end, self.ann_canvas_h, fill='cyan2', width=1, tag='LBI_end_line')
        #self.prediction_canvas.coords('result_canvas_section_line',  self.current_section_canvas.index.get()*self.ann_scale_w, 0,  self.current_section_canvas.index.get()*self.ann_scale_w, self.ann_canvas_h)
            
        #print(self.mlbi, np.argmax(self.lbi)+interv[0])

        #print(np.where(self.lbi>thr))
    
    def LBI(self,i0,i1):
        return np.sum(self.lipid[i0:i1,:]==1)/(self.Nr*(i1-i0)) *1000
    
    def LBI_superf(self,i0,i1):
        return np.sum(self.labels[i0:i1,:]==lipid_color_index)/(self.Nr*(i1-i0)) *1000
    
    
    def warningsold(self, A=3.5, a=180, t=75):

        for f in range(self.Nf):
            if (self.areas[f] < A and self.plaque_angle[f] > a and  round(self.fct[f][0] * self.calibration_scale) < t):
                self.annotation[f, :] = np.ones(self.Nr)
                #print('bad frame' , f)

    def load_single_annotation(self, A=3.5, A_flag=False, a=180, a_flag=False, t=75 ,t_flag=True):
        self.single_annotation=np.ones([self.Nf, self.Nr])*unknown_color_index
        for f in range(self.Nf):
            if ( round(self.fct[f][0] * self.calibration_scale) < t and round(self.fct[f][0])>0 ) and t_flag:
                self.single_annotation[f,:]=non_lipid_color_index
            if round(self.areas[f]) < A and A_flag:
                self.single_annotation[f,:]=non_lipid_color_index
            if round(self.plaque_angle[f]) > a and a_flag:
                self.single_annotation[f,:]=non_lipid_color_index

    def warnings_old(self, A=3.5, a=180, t=75, l= 10):
        self.annotation [:,:]= 0
        for interv in self.interv_of_pertinence:
            for f in range(interv[0],interv[1]):
                if ( round(self.fct[f][0] * self.calibration_scale < t and round(self.fct[f][0])>0 )):
                    k0 = np.max([0,f-l,interv[0]])
                    k1 = np.min([self.Nf, f + l,interv[1]])
                    farea=0
                    fangle=0
                    i1=k0-1
                    i2=f
                    for k in range(k0,k1):
                        if self.areas[k] < A:
                            farea=1
                            if i1==k0-1:
                                i1=k
                            if k>i2 :
                                i2=k
                        if self.plaque_angle[f] > a:
                            fangle=1
                            if i1==k0-1:
                                i1=k
                            if k>i2:
                                i2=k
                        if i1>f or i1==k0-1:
                            i1=f-1
                    if farea==1 and fangle==1 :
                        self.annotation[i1:i2, :] = 1

    def warnings(self, A=3.5, a=180, t=75, l= 10):
        self.annotation [:,:]= 0

        lbi_thr=398
        if self.Nf>500:
            lbi_depth=40
        else:
            lbi_depth=20

        for interv in self.interv_of_pertinence:
            lp = np.where(self.lbi[interv[0]:interv[1]-lbi_depth] > lbi_thr)
            print(lp,np.size(lp))
            for i in range(np.size(lp)):
                lp0=interv[0]+lp[0][i]
                lp1=lp0 + lbi_depth
                print(lp0,lp1)

                wlumen = 0
                wfct = 0

                for f in range(lp0,lp1):

                    if ( round(self.fct[f][0] * self.calibration_scale <= t and round(self.fct[f][0])>0 )):
                        wfct=1

                    if self.areas[f] <= A:
                        wlumen=1

                if wfct==1 and wlumen==1:
                    self.annotation[lp0:lp1, :] = 1




    def calibrate_old(self):
        m=20
        M=130
        f=134
        df=47
        sigma=10
        eps=4
        nf=0
        calibrations=[]
        
        while nf<100 and (not (nf>6 and sigma<4)):
            print('frame', nf)
            central=np.copy(self.pullback[f,:,0:M])
            #print((sum(sum(central>100)) + sum(sum(central<0)) )/central.size , '  eliminati')
            #central[central>100]=0
            #central[central<0]=0
            print((sum(sum((central>0) & (np.std(central,axis=0)>50)) ))/central.size , '  eliminati')
            central[(central>0) & (np.std(central,axis=0)>50)]=0
            central[(central>-1) & (np.mean(central,axis=0)<15)]=0
            part_sum=np.zeros(M-m)
            s=np.sum(central[:,0:m])
            max_s=0
            for r in range(0,M-m):
                old_s=s
                part_sum[r]=s
                s+=np.sum(self.pullback[f,:,r])
                if (s-old_s>max_s):
                    max_s=s-old_s
            print('max circ',max_s)
            #print(part_sum)
            
            n=1
            h=4
            flat=False
            while n<M-m-h and flat==False:
                if(part_sum[M-m-n]- part_sum[M-m-n-h]<max_s/4):
                    flat=True
                n+=1
            
            print('flat ' , M - n)
            
            while n<M-m-h and flat==True:
                #print(n,M-m-n,M-m-n-h)
                if(part_sum[M-m-n]- part_sum[M-m-n-h]>max_s*0.3):
                    flat=False
                n+=1
            
            #n-=h
            print('calibrate', M-n)
            '''
            plt.imshow(central)
            plt.show()

            self.current_section_canvas.index.set(f)
            self.current_section_canvas.update()
            self.calibration=M-n
            self.current_section_canvas.show_calibration()
            
            img=np.copy(central)
            img[:,M-n]=np.max(central)
            plt.imshow(img)
            plt.show()
            '''
            calibrations.append(M-n)
            sigma=np.std(np.asarray(calibrations))
            nf+=1
            f=(f+df)%self.Nf
            print(sigma)
        
        print(sigma,nf)
        while sigma > 2 :
            c=np.mean(np.asarray(calibrations))
            calibrations=[x for x in calibrations if np.abs(x - c)<sigma ]
            sigma=np.std(np.asarray(calibrations))
            print(len(calibrations),sigma)
           
        '''
        if nf==100:
            for i,calibration in enumerate(calibrations):
                if np.abs(calibration - c)>sigma:
                    del calibrations[i]
                    print(calibration,calibrations[i])
        '''
            
        self.calibration=np.rint(np.mean(np.asarray(calibrations)))
        print('final calibration',self.calibration)
        self.current_section_canvas.show_calibration()
        self.calibration_scale=450/self.calibration 
        

    def calibrate(self):
        m=20
        M=130
        f=134
        df=47
        sigma=10
        eps=4
        nf=0
        calibrations=[]
        
        while nf<100 and (not (nf>6 and sigma<4)):
            #print('frame', nf)
            central=np.copy(self.pullback[f,:,0:M])

            #print((sum(sum((central>0) & (np.std(central,axis=0)>50)) ))/central.size , '  eliminati')
            central[(central>0) & (np.std(central,axis=0)>50)]=0
            central[(central>-1) & (np.mean(central,axis=0)<15)]=0
            
            #central=central/np.max(central[:,m:])
            central[central>0]=1
            #print(np.max(central),np.min(central),np.mean(central))
            
            part_sum=np.zeros(M-m)
            s=np.sum(central[:,0:m])
            for r in range(0,M-m):
                old_s=s
                part_sum[r]=s
                s+=np.sum(central[:,r+m])
                #print(np.sum(central[:,r+m]),0.7*self.Nr)
            #print(part_sum)
            
            n=1
            h=1
            #flat=False
            '''while n<M-m-h and flat==False:
                if(part_sum[M-m-n]- part_sum[M-m-n-h]<max_s/4):
                    flat=True
                n+=1
            '''
            #print('flat ' , M - n)

            flat=True            
            while n<M-m-h and flat==True:
                if(part_sum[M-m-n]- part_sum[M-m-n-h]>0.7*self.Nr):
                    flat=False
                n+=1
            
            #n-=h
            #print('calibrate', M-n)
            '''
            plt.imshow(central)
            plt.show()

            self.current_section_canvas.index.set(f)
            self.current_section_canvas.update()
            self.calibration=M-n
            self.current_section_canvas.show_calibration()
            
            img=np.copy(central)
            img[:,M-n]=np.max(central)
            plt.imshow(img)
            plt.show()
            '''
            
            calibrations.append(M-n)
            sigma=np.std(np.asarray(calibrations))
            nf+=1
            f=(f+df)%self.Nf
            #print(sigma)
        
        #print(sigma,nf)
        while sigma > 4 :
            c=np.mean(np.asarray(calibrations))
            calibrations=[x for x in calibrations if np.abs(x - c)<sigma ]
            sigma=np.std(np.asarray(calibrations))
            #print(len(calibrations),sigma)
           
        self.calibration=np.rint(np.mean(np.asarray(calibrations)))
        print('final calibration',self.calibration)
        self.current_section_canvas.show_calibration()
        self.calibration_scale=450/self.calibration 
        
    def write_results(self):
        
        wb = Workbook()
        result=wb[wb.sheetnames[0]]
        result.title = "Results"
        #_cell=result.cell(1,1)
        #_cell.style.font.size = 8
        #_cell.style.font.bold = True
        #_cell.style.fill.fill_type = Fill.FILL_SOLID
        #_cell.style.fill.start_color.index = Color.DARKYELLW

        resume = wb.create_sheet("Resume")
                
        columns=['Frame','Lumen Area','Max Diameter','Min Diameter','Diameters Ratio','Lipid Arc','Superficial Plaque %', 'FCT' , 'FCT Mean']
        
        for j in range(1,10):
            result.cell(row = 1, column = j).value = columns[j-1]

        fill=PatternFill(fill_type= FILL_SOLID,start_color='FFF993')
        for cell in result[1]:
            result[str(cell.coordinate)].font = Font(color='000000', italic=True, bold=True)
            result[str(cell.coordinate)].fill = fill#Fill(color= '')
        
        letters=['A','B','C','D','E','F','G','H','I']    
        for i in range(len(columns)):
            result.column_dimensions[letters[i]].width = len(columns[i])+3
        
        keys={}
        i=1
        for cell in  result[1]:
            keys[cell.value]=i
            i+=1
        
        for i in range(self.Nf):
            result.cell(i+2,keys['Frame']).value = i+1
            result.cell(i+2,keys['Lumen Area']).value = int(self.areas[i] * self.calibration_scale* self.calibration_scale/ ((818.5/160)*(818.5/160))  *100)/100
            result.cell(i+2,keys['Max Diameter']).value = int(self.diameters[1,i,0]* self.calibration_scale/1000 * 100)/100
            result.cell(i+2,keys['Min Diameter']).value = int(self.diameters[0,i,0]* self.calibration_scale/1000 * 100)/100
            result.cell(i+2,keys['Diameters Ratio']).value = int(self.diameters[0,i,0]/self.diameters[1,i,0] * 100)/100
            
            result.cell(i+2,keys['Lipid Arc']).value = int(self.plaque_angle[i]*100)/100
            
            if sum(self.lipid[i,:])>0:
                
                result.cell(i+2,keys['Superficial Plaque %']).value = int(sum(self.labels[i,:]==lipid_color_index)/sum(self.lipid[i,:])*100) 
                
                all_fct=self.all_fct[i] * self.calibration_scale
                all_fctm=np.ma.array(all_fct , mask=[ (all_fct==0) | (all_fct>999)])
                
                
                if str(all_fctm.mean()).replace('.','').isnumeric():
                    result.cell(i+2,keys['FCT']).value =  int(self.fct[i][0] * self.calibration_scale)
                    result.cell(i+2,keys['FCT Mean']).value =  int(all_fctm.mean())    
                else :
                    result.cell(i+2,keys['FCT']).value =  'NA'
                    result.cell(i+2,keys['FCT Mean']).value = 'NA'
            
            else :
                result.cell(i+2,keys['Superficial Plaque %']).value = 'None'
                result.cell(i+2,keys['FCT']).value = 'None'
                result.cell(i+2,keys['FCT Mean']).value = 'None'
        
        
        columns=['Min Lumen Area','Min Lumen Area Frame','Max Lipid Arc','Max Lipid Arc Frame','Superficial Plaque %', 'Min FCT','Min FCT Frame','Vulnarable Plaque']
        
        for j in range(1,9):
            resume.cell(row = 1, column = j).value = columns[j-1]

        fill=PatternFill(fill_type= FILL_SOLID,start_color='FFF993')
        for cell in result[1]:
            resume[str(cell.coordinate)].font = Font(color='000000', italic=True, bold=True)
            resume[str(cell.coordinate)].fill = fill#Fill(color= '')
            
        letters=['A','B','C','D','E','F','G','H']    
        for i in range(len(columns)):
            resume.column_dimensions[letters[i]].width = len(columns[i])+3
            
            
        keys={}
        i=1
        for cell in  resume[1]:
            keys[cell.value]=i
            i+=1
        
        
        
        mfct=1000
        mfct_fr=-1
        marea=1000
        marea_fr=-1
        mla=0
        mla_fr=-1
        vulnerable=0
        lipid=np.zeros([self.Nf,self.Nr])
        labels=np.zeros([self.Nf,self.Nr])
        for inter in self.interv_of_pertinence:
            for f in range(inter[0],inter[1]):
                if  self.fct[f][0]<mfct and (not self.fct[f][0]==0):
                    mfct= self.fct[f][0]
                    mfct_fr=f
                if self.areas[f]<marea and (not self.areas[f]==0):
                    marea=self.areas[f]
                    marea_fr=f
                if self.plaque_angle[f]>mla:
                    mla=self.plaque_angle[f]
                    mla_fr=f    
                if self.annotation[f,0]==non_lipid_color_index:
                    vulnerable=1
                lipid[f,:]=self.lipid[f,:]
                labels[f,:]=self.labels[f,:]
                
        resume.cell(2,keys['Min Lumen Area']).value = int(marea * self.calibration_scale* self.calibration_scale/ ((818.5/160)*(818.5/160))  *100)/100
        resume.cell(2,keys['Min Lumen Area Frame']).value = marea_fr+1
        
        resume.cell(2,keys['Max Lipid Arc']).value = int(mla)
        resume.cell(2,keys['Max Lipid Arc Frame']).value  = mla_fr +1
        
        resume.cell(2,keys['Superficial Plaque %']).value  = int(sum(sum(labels==lipid_color_index))/sum(sum(lipid))*100)
        
        if mfct_fr>-1:
            resume.cell(2,keys['Min FCT']).value = int(mfct*self.calibration_scale)
            resume.cell(2,keys['Min FCT Frame']).value = mfct_fr+1
        else:
            resume.cell(2,keys['Min FCT']).value = 'None'
            resume.cell(2,keys['Min FCT Frame']).value = 'None'
        
        resume.cell(2,keys['Vulnarable Plaque']).value = vulnerable 
        
        wb.save(self.pullback_dir+'/Resume.xlsx')
            
        
    def load_preanal(self):

        if os.path.exists(self.pullback_dir + '/Calibration.npy'):
            self.calibration = np.load(self.pullback_dir + '/Calibration.npy')
            self.calibration_scale=450/self.calibration
        
        if  os.path.exists(self.pullback_dir + '/CalibrationCenter.npy'):
            cal_center=np.load(self.pullback_dir + '/CalibrationCenter.npy')
            self.cal_center_x.set(cal_center[0])
            self.cal_center_y.set(cal_center[1])
            
        if  os.path.exists(self.pullback_dir + '/GuideArcs.txt'):
            self.guide_arcs = load_list(self.pullback_dir + '/GuideArcs.txt')

        if os.path.exists(self.pullback_dir + '/Area.npy'):
            self.areas = np.load(self.pullback_dir + '/Area.npy')

        if  os.path.exists(self.pullback_dir + '/diameters.npy'):
            self.diameters = np.load(self.pullback_dir + '/diameters.npy')

        if  os.path.exists(self.pullback_dir + "/Prediction.npy"):
            self.preds = np.load(self.pullback_dir + "/Prediction.npy")
        else:
            self.preds = np.NaN

        if os.path.exists(self.pullback_dir + '/LipidArcs.txt') :
           self.plaque_arcs = load_list(self.pullback_dir + '/LipidArcs.txt')

        if os.path.exists(self.pullback_dir + '/PlaqueAngle.npy'):
            self.plaque_angle = np.load(self.pullback_dir + '/PlaqueAngle.npy')
            #print ('plaque',self.plaque_angle)

        if os.path.exists(self.pullback_dir + '/Fct.txt'):
            self.fct = load_list(self.pullback_dir + '/Fct.txt')

        if os.path.exists(self.pullback_dir + "/FCMeasurement.txt"):
            mes = load_list(self.pullback_dir + "/FCMeasurement.txt")
            self.measure_points = mes[0]
            self.measure_points_cart = mes[1]
        if os.path.exists(self.pullback_dir + '/Intervals.txt'):
            self.interv_of_pertinence = load_list(self.pullback_dir + '/Intervals.txt')
            self.scale_canvas.var=[]
            n=0
            for interv in self.interv_of_pertinence:
                self.scale_canvas.var.append(Entry(self.scale_canvas,width=3, borderwidth=0, highlightthickness=0))
                self.scale_canvas.var.append(Entry(self.scale_canvas,width=3, borderwidth=0, highlightthickness=0))
                self.scale_canvas.var[n].insert(END, interv[0])
                self.scale_canvas.var[n+1].insert(END, interv[1])
                n+=1
        if os.path.exists(self.pullback_dir + '/Fct.txt') and os.path.exists(self.pullback_dir + '/PlaqueAngle.txt') and os.path.exists(self.pullback_dir + '/Area.npy') :
            self.warnings()
            
        self.annotation = self.load_labels_classification()
        
        if os.path.exists(self.pullback_dir + '/LBI.npy') :
            self.lbi=np.load(self.pullback_dir + '/LBI.npy')
        
        if os.path.exists(self.pullback_dir + '/LBIs.npy') :
            self.lbis=np.load(self.pullback_dir + '/LBIs.npy')
            
        self.annotation_canvas.update(1)
        self.load_single_annotation()
        self.info_canvas.update(2)
        self.current_section_canvas.old_intima = [np.copy(self.lumen)]
        self.current_section_canvas.old_fc = [np.copy(self.fc)]
        self.current_section_canvas.old_fct = [np.copy(self.fct)]

    def preprocessing(self):
        p=1
        T=time.time()
        T0=T
        
        if not os.path.exists(self.pullback_dir + '/Calibration.npy'):
            self.calibrate()
            np.save(self.pullback_dir + '/Calibration.npy',self.calibration)
            
            
        if self.flags[0]==0 or (not os.path.exists(self.pullback_dir + '/GuideArcs.txt')):
            I,lines=self.current_section_canvas.guide_boundary_double(0,self.Nf)
            #self.accept_guide()
            #self.wait_window(self.tip)
            save_list(self.guide_arcs,self.pullback_dir +'/GuideArcs.txt')
        self.guide_arcs=load_list(self.pullback_dir +'/GuideArcs.txt')
        #self.guide = self.guide.astype(np.int)
        print('end guide', (time.time()-T)/60)

        #self.lumen = self.lumen.astype(np.int)
        
        T=time.time()
        if self.flags[1]==0 :
            for inter in self.interv_of_pertinence:
                for f in range(inter[0],inter[1]):
                    self.current_section_canvas.lumen_ds(f)
                    self.current_section_canvas.calculate_area(f)
                    self.current_section_canvas.calculate_diameter(f)
                    self.current_section_canvas.delete("max_diameter")
                    self.current_section_canvas.delete("min_diameter")
            self.lumen = self.lumen.astype(np.int)
            np.save(self.lumen_path,self.lumen)
            np.save(self.pullback_dir +'/Area.npy',self.areas)
            np.save(self.pullback_dir + '/diameters.npy', self.diameters)

        if  not os.path.exists(self.pullback_dir+'/Area.npy'):
            for inter in self.interv_of_pertinence:
                for f in range(inter[0],inter[1]):
                    self.current_section_canvas.calculate_area(f)
                    np.save(self.pullback_dir +'/Area.npy',self.areas)
        else:
            self.areas=np.load(self.pullback_dir +'/Area.npy')

        if  not os.path.exists(self.pullback_dir+'/diameters.npy'):
            for inter in self.interv_of_pertinence:
                for f in range(inter[0],inter[1]):
                    self.current_section_canvas.calculate_diameter(f)
                    np.save(self.pullback_dir +'/diameters.npy',self.diameters)
        else:
            self.diameters=np.load(self.pullback_dir +'/diameters.npy')
            
                
        print('end lumen ', (time.time()-T)/60)
        T=time.time()
        #plt.imshow(np.asarray(np.maximum(  I,np.max(I)*lines)))
        #plt.show()
        
        self.identify_plaque()
        if self.flags[2]==0 or (not os.path.exists(self.pullback_dir + '/LipidArcs.txt')) or (not os.path.exists(self.pullback_dir + '/PlaqueAngle.txt')):
            for inter in self.interv_of_pertinence:
                for f in range(inter[0],inter[1]):
                    self.current_section_canvas.calculate_plaque_angle(f)
            save_list(self.plaque_arcs,self.pullback_dir +'/LipidArcs.txt')
            np.save(self.pullback_dir +'/PlaqueAngle.npy',self.plaque_angle)
        self.plaque_angle=np.load(self.pullback_dir +'/PlaqueAngle.npy')
        self.plaque_arcs=load_list(self.pullback_dir +'/LipidArcs.txt')

        print('end plaque', (time.time()-T)/60)
        if self.flags[3]==0 or (not os.path.exists(self.pullback_dir + '/Fct.txt')):
            d=self.current_section_canvas.fc_weights()
            for inter in self.interv_of_pertinence:
                for f in range(inter[0],inter[1]):
                    self.current_section_canvas.fc_boundary(f,d)
            save_list(self.fct,self.pullback_dir +'/Fct.txt')
            self.all_fct = self.all_fct.astype(np.int)
            np.save(self.all_fct_path,self.all_fct)
        print('end fct', (time.time()-T)/60)
        
        self.fct=load_list(self.pullback_dir +'/Fct.txt')
        print('end processing', (time.time()-T0)/60)
        
        if os.path.exists(self.pullback_dir + "/FCMeasurement.txt"):
            mes=load_list(self.pullback_dir + "/FCMeasurement.txt")
            self.measure_points=mes[0]
            self.measure_points_cart=mes[1]
        
        
            
        self.warnings()
        self.annotation=self.load_labels_classification ()
        
        #if not os.path.exists(self.pullback_dir + '/LBIs.npy') :
            #self.analyze_LBI(superf=1)
            #self.lbis=self.lbi
        #else:
            #self.lbis=np.load(self.pullback_dir + '/LBIs.npy')

        if not os.path.exists(self.pullback_dir + '/LBI.npy') :
            self.analyze_LBI(superf=0)
        else:
            self.lbi=np.load(self.pullback_dir + '/LBI.npy')

        self.annotation_canvas.update(1)
        self.load_single_annotation()
        self.info_canvas.update(2)
        self.current_section_canvas.old_intima=[np.copy(self.lumen)]
        self.current_section_canvas.old_fc=[np.copy(self.fc)]
        self.current_section_canvas.old_fct=[np.copy(self.fct)]

    
    def identify_plaque(self):

        if not os.path.exists(self.pullback_dir + "/Prediction.npy"):
            model = load_model(os.path.dirname(os.path.realpath('__file__')) + "/PlaccaModel.h5")

            X_test, lumen_test = self.pullback_n_labels_part_min_multiclass(0, self.Nf, depth=400, deg=64, stride=16)

            X_test = np.reshape(X_test, np.append(X_test.shape, 1))
            lumen_test = np.reshape(lumen_test, np.append(lumen_test.shape, 1))

            self.preds = model.predict(X_test, verbose=1)
            #self.preds= model.predict([X_test,lumen_test],verbose=1)
            np.save(self.pullback_dir + "/Prediction.npy", self.preds)
            self.topological_build()
            self.flags[2]=0

        else:
            self.preds = np.load(self.pullback_dir + "/Prediction.npy")
            self.annotation = self.load_labels_classification()

    def build_labels(self):
        p=pred_to_label_multiclass2(self.preds, thr=self.current_section_canvas.thr.get(),thr_ca=self.current_section_canvas.thr_ca.get())
        self.labels=self.ricostruisci_labels_guide_multi(p,deg=64, stride=16)
        
        #self.labels=self.labels_continuity(self.labels, self.Nf, self.Nr, stride=16)
        #self.labels=self.calcium_continuity(self.labels, self.Nf, self.Nr, stride=16)
        
        
        self.lipid[self.labels == 1] = 1
        self.lipid[self.labels != 1] = 0
        self.calcium[self.labels == 3]=1
        self.calcium[self.labels != 3]=0
        self.annotation = self.load_labels_classification()
    
    def identify_plaque_frame(self):
        f=self.current_section_canvas.index.get()
        model = load_model(os.path.dirname(os.path.realpath('__file__')) + "/PlaccaModel.h5")
        X_test, lumen_test = self.pullback_n_labels_part_min_multiclass(f, f+1, depth=400, deg=64, stride=16)
        X_test = np.reshape(X_test, np.append(X_test.shape, 1))
        lumen_test = np.reshape(lumen_test, np.append(lumen_test.shape, 1))
        preds = model.predict(X_test, verbose=1)
        #preds= model.predict([X_test,lumen_test],verbose=1)
        p=pred_to_label_multiclass2(preds, thr=self.current_section_canvas.thr.get(),thr_ca=self.current_section_canvas.thr_ca.get())
        self.labels[f]=self.ricostruisci_labels_guide_multi(p, deg=64, stride=16, single_frame=True)[f]
        self.lipid[f][self.labels[f] == 1] = 1
        self.lipid[f][self.labels[f] != 1] = 0
        self.calcium[f][self.labels[f] == 3]=1
        self.calcium[f][self.labels[f] != 3]=0
        self.annotation = self.load_labels_classification()
                
    def ricostruisci_labels_guide_multi(self,Y, deg=64, stride=8,single_frame=False):
                
        if single_frame:
            nSlices=1
            k0=int (self.Nr/stride)*self.current_section_canvas.index.get()
        else:
            nSlices=self.Nf
            k0=0 
        Nr=self.Nr
        n4Slice= int (Nr/stride) 
        N=int ((Nr-deg)/stride) +1
        nTot= n4Slice*nSlices
        count=0
        labels=np.zeros([self.Nf, Nr])
        for k in range(nTot):
            
            k_shift=k+k0
            if(k%n4Slice*stride+deg<=Nr):
                if sum(self.guide[k_shift//n4Slice,  k_shift%n4Slice*stride: k_shift%n4Slice*stride+deg])>0 :
                    count+=1
                    #print(count) 19501
                    #if k%n4Slice ==0 :
                    #labels [k//n4Slice,  k%n4Slice*stride : k%n4Slice*stride+deg] = 2
                    #else :
                    #    labels [k//n4Slice,  k%n4Slice*stride + deg - stride : k%n4Slice*stride+deg] = 2
                else :
                    if(Y[k-count]==2) :
                        labels [k_shift//n4Slice,  k_shift%n4Slice*stride : k_shift%n4Slice*stride+deg] = 3
                    
                    elif(Y[k-count]==1) :
                        labels [k_shift//n4Slice,  k_shift%n4Slice*stride : k_shift%n4Slice*stride+deg] = 1
            else :
                if sum(self.guide[k_shift//n4Slice,  k_shift%n4Slice*stride:  Nr]) + sum(self.guide[k_shift//n4Slice, 0 : (k_shift%n4Slice*stride+deg)%Nr ]) > 0 :
                    count+=1
                    #print(count) 19501
                    #if k%n4Slice == 0 :
                    #labels [k//n4Slice,  k%n4Slice*stride : Nr] = 2
                    #labels [k//n4Slice,  0 : (k%n4Slice*stride+deg)%Nr] = 2
                    #else :
                    #    labels [k//n4Slice,  k%n4Slice*stride + deg - stride : k%n4Slice*stride+deg] = 2
                else :
                    if(Y[k-count]==2) :
                        labels [k_shift//n4Slice,  k_shift%n4Slice*stride : Nr] = 3
                        labels [k_shift//n4Slice,  0 : (k_shift%n4Slice*stride+deg)%Nr] = 3
                    
                    elif(Y[k-count]==1) :
                        labels [k_shift//n4Slice,  k_shift%n4Slice*stride : Nr] = 1
                        labels [k_shift//n4Slice,  0 : (k_shift%n4Slice*stride+deg)%Nr] = 1
         
       #('rec',nTot,count)
        return labels
        
    def ricostruisci_labels_continue_guida(self,Y, c=1, deg=64, stride=16, single_frame=False):
        
        if single_frame:
            nSlices=1
        else:
            nSlices=self.Nf
        Nr=self.Nr
        n4Slice= int (Nr/stride) 
        N=int ((Nr-deg)/stride) +1
        nTot= n4Slice*nSlices
        count=0
        p=pred_to_label_multiclass2(self.preds, thr=self.current_section_canvas.thr.get(),thr_ca=self.current_section_canvas.thr_ca.get())
        labels=np.zeros([nSlices, Nr])
        for k in range(nTot):
    
            if(k%n4Slice*stride+deg<=Nr):
                if sum(self.guide[k//n4Slice,  k%n4Slice*stride: k%n4Slice*stride+deg])>0 :
                    count+=1
                    #labels [k//n4Slice,  k%n4Slice*stride : k%n4Slice*stride+deg] = 2
                else :
                    labels [k//n4Slice,  k%n4Slice*stride : k%n4Slice*stride+deg] = Y[k-count,c]
            else :
                if sum(self.guide[k//n4Slice,  k%n4Slice*stride:  Nr]) + sum(self.guide[k//n4Slice, 0 : (k%n4Slice*stride+deg)%Nr ]) > 0 :
                    count+=1
                    #labels[k//n4Slice,  k%n4Slice*stride:  Nr] = 2 
                    #labels[k//n4Slice, 0 : (k%n4Slice*stride+deg)%Nr ] = 2
                else :
                    labels [k//n4Slice,  k%n4Slice*stride : ] = Y[k-count,c]
                    #if k%n4Slice*stride+stride< Nr:
                    #    labels [k//n4Slice,  k%n4Slice*stride : k%n4Slice*stride+stride] = Y[k-count,c]
                    #else:
                    #    labels [k//n4Slice,  k%n4Slice*stride : Nr] = Y[k-count,c]
                        #labels [k//n4Slice,  0 : (k%n4Slice*stride+stride)%Nr] = Y[k-count,c]
         
        labels[self.guide == 2] = -number_of_colors + unknown_color_index
        return labels

    def pullback_n_labels_part_min_multiclass(self, s, e, classes=3, depth=400, deg=64, stride=8, mixed_mode=False):

        nSlices = e - s
        Nr = self.Nr
        n4Slice = int(Nr / stride)
        N = int((Nr - deg) / stride) + 1
        nTot = n4Slice * nSlices
        X = np.zeros((nTot, deg, depth))
        lumen = np.zeros((nTot, 64))

        massimo = 0
        count = 0
        count_guide = 0
        for i in range(s,s+nSlices):
            i_shift=i-s
            for j in range(n4Slice):
                if (j < N):
                    if (sum(self.guide[i, j * stride:j * stride + deg]) == 0):
                        count += 1
                        m = np.min(self.lumen[i, j * stride:j * stride + deg])
                        lumen[i_shift * n4Slice + j - count_guide, :] = self.lumen[i, j * stride:j * stride + deg]
                        max_depth = self.Np - m
                        if (max_depth < depth):
                            X[i_shift * n4Slice + j - count_guide, :, 0:max_depth] = self.pullback[i, j * stride:j * stride + deg, m:self.Np]
                        else:
                            X[i_shift * n4Slice + j - count_guide, :, :] = self.pullback[i, j * stride:j * stride + deg, m:m + depth]
                    else:
                        count_guide += 1
                else:

                    if (sum(self.guide[i, j * stride:]) + sum(self.guide[i, 0:deg + j * stride - self.Nr]) == 0):
                        count += 1
                        m = np.min(
                            [np.min(self.lumen[i, j * stride:]), np.min(self.lumen[i, 0:deg + j * stride - self.Nr])])
                        lumen[i_shift * n4Slice + j - count_guide, 0:self.Nr - j * stride] = self.lumen[i, j * stride:]
                        lumen[i_shift * n4Slice + j - count_guide, self.Nr - j * stride:] = self.lumen[i, 0:deg + j * stride - self.Nr]

                        max_depth = self.Np - m
                        if (max_depth < depth):
                            X[i_shift * n4Slice + j - count_guide, 0:self.Nr - j * stride, 0:max_depth] = self.pullback[i,j * stride:, m:self.Np]
                            X[i_shift * n4Slice + j - count_guide, self.Nr - j * stride:, 0:max_depth] = self.pullback[i, 0:deg + j * stride - self.Nr,m:self.Np]

                        else:
                            X[i_shift * n4Slice + j - count_guide, 0:self.Nr - j * stride, :] = self.pullback[i, j * stride:,m:m + depth]
                            X[i_shift * n4Slice + j - count_guide, self.Nr - j * stride:, :] = self.pullback[i, 0:deg + j * stride - self.Nr, m:m + depth]
                    else:
                        count_guide += 1
        X = X[0:count, :, :]
        lumen = lumen[0:count, :]
        
        #print('count guide spezz', count_guide,nTot,count)

        return X, lumen
    
    def topological_build(self):
        #T=build_labels_topology(self.preds, self.guide, self.Nf, self.Nr)
        #labels_pred = ricostruisci_from_topology(T ,self.guide, self.Nf, self.Nr, 64, 16, self.current_section_canvas.thr.get(),self.current_section_canvas.thr_ca.get() )
        
        T,guide_counter=build_matrix_topology(self.preds, self.guide, self.Nf, self.Nr)
        labels_pred = ricostruisci_from_topology_matrix(T ,guide_counter,self.guide, self.Nf, self.Nr, 64, 16, self.current_section_canvas.thr.get(),self.current_section_canvas.thr_ca.get() )
        
        self.lipid[labels_pred == 1] = 1
        self.lipid[labels_pred != 1] = 0
        self.calcium[labels_pred == 3]=1
        self.calcium[labels_pred != 3]=0

        for inter in self.interv_of_pertinence:
                for f in range(inter[0],inter[1]):
                    self.current_section_canvas.calculate_plaque_angle(f)
        save_list(self.plaque_arcs,self.pullback_dir +'/LipidArcs.txt')
        np.save(self.pullback_dir +'/PlaqueAngle.npy',self.plaque_angle)


        self.annotation = self.load_labels_classification()
        '''old_labels=np.copy(self.labels).flatten()
        self.annotation = self.load_labels_classification()
        new_labels=np.copy(self.labels).flatten()
        n=[]
        o=[]
        for i in range (0,new_labels.size,16):
            n.append(new_labels[i])
            o.append(old_labels[i])
        print(classification_report(new_labels, old_labels))#, target_names=['None','Lipid','Calcium']))
        '''
    
    def labels_continuity(self, lipid,Nf,Nr,stride):
        labels=np.zeros([Nf,Nr])
        labels[lipid==1]=1
        for f in range(3,Nf-3):
            for r in range(Nr//stride):
                if labels[f,r*stride]==1:
                    if(labels[f+1,r*stride] + labels[f-1,r*stride] + labels[f+1,((r-1)*stride)%Nr] + labels[f-1,((r-1)*stride)%Nr] + labels[f+1,((r+1)*stride)%Nr] + labels[f-1,((r+1)*stride)%Nr]+
                       labels[f+2,r*stride] + labels[f-2,r*stride] + labels[f+2,((r-1)*stride)%Nr] + labels[f+2,((r-2)*stride)%Nr] + labels[f+2,((r+1)*stride)%Nr] + labels[f+2,((r+2)*stride)%Nr] +
                       labels[f-2,((r-1)*stride)%Nr] + labels[f-2,((r-2)*stride)%Nr] + labels[f-2,((r+1)*stride)%Nr] + labels[f-2,((r+2)*stride)%Nr])==0:
                        lipid[f,r*stride:np.min([(r+1)*stride , Nr])]=0
                        #print(f,r,'printo brass')
        return lipid
        
    def calcium_continuity(self, lipid,Nf,Nr,stride):
        labels=np.zeros([Nf,Nr])
        labels[lipid==3]=1
        for f in range(3,Nf-3):
            for r in range(Nr//stride):
                if labels[f,r*stride]==1:
                    if(labels[f+1,r*stride] + labels[f-1,r*stride] + labels[f+1,((r-1)*stride)%Nr] + labels[f-1,((r-1)*stride)%Nr] + labels[f+1,((r+1)*stride)%Nr] + labels[f-1,((r+1)*stride)%Nr]+
                       labels[f+2,r*stride] + labels[f-2,r*stride] + labels[f+2,((r-1)*stride)%Nr] + labels[f+2,((r-2)*stride)%Nr] + labels[f+2,((r+1)*stride)%Nr] + labels[f+2,((r+2)*stride)%Nr] +
                       labels[f-2,((r-1)*stride)%Nr] + labels[f-2,((r-2)*stride)%Nr] + labels[f-2,((r+1)*stride)%Nr] + labels[f-2,((r+2)*stride)%Nr])==0:
                        lipid[f,r*stride:np.min([(r+1)*stride , Nr])]=0
                        #print(f,r,'printo brass')
        return lipid

    

if __name__ == "__main__":
    #pullback = np.load(pullback_path,mmap_mode='r')
    #annotation,labels=load_labels_classification ()
    #print(annotation.shape,labels.shape)
    gui=Main()#pullback,annotation,labels,[],[])#,DP.load_list(model_file+"/bad.txt"),DP.load_list(model_file+"/good.txt"))
    gui.mainloop()