import os
import numpy as np
import GuiClean as GC
from openpyxl import Workbook, load_workbook
import unicodedata
import sys
import shutil

wb = load_workbook(os.path.dirname(os.path.realpath('__file__')) + "/OCTvsNIRS 41 pazienti.xlsx")

for sheet in wb:
    print(sheet.title)

data = wb['Database Final']

keys = {}
gui = GC.Main()

i = 1
for cell in data[1]:
    keys[cell.value] = i
    print(i)
    i += 1

print(keys)

database = 'D:\Franz\oct_software\data\OCTvsNIRS'  # '/media/eugenio/Hitachi/Pazienti CLIMA'#'/media/fisici/Hitachi/Pazienti CLIMA'   #'/home/eugenio/workspace/documets di pasqualino/Documents/OCTdatabase/pazienti mace'


for lp in range(2,43):
    found = False
    c = False

    # print(data.cell(lp, keys['Patient surname']).value, data.cell(lp, keys['Patient name']).value, str(data.cell(lp, keys['Center']).value), str(data.cell(lp, keys['Acquisition date']).value),'PB'+str(data.cell(lp, keys['Pullback']).value ))
    patient_surname = data.cell(lp, keys['Patient ']).value.strip().lower()
    patient_surname = patient_surname.replace("_", " ")
    patient_surname = unicodedata.normalize('NFD', patient_surname) \
        .encode('ascii', 'ignore') \
        .decode("utf-8")

    c = True
    patients = next(os.walk(database))[1]

    for patient in patients:
        m_patient = patient.strip().lower()
        m_patient = unicodedata.normalize('NFD', m_patient) \
            .encode('ascii', 'ignore') \
            .decode("utf-8")
        if (patient_surname in m_patient):
            data_path = os.path.join(database, patient)
            found = True
    if found:
        gui.import_pullback_callback(data_path)

        print(data_path)

        '''
        if not os.path.exists(gui.pullback_dir + '/Calibration.npy'):
            gui.calibrate()
            np.save(gui.pullback_dir + '/Calibration.npy', gui.calibration)

        if gui.Nf < 500:
            lbi_depth = 20
        else:
            lbi_depth = 40

        #gui.analyze_guide()
        #gui.current_section_canvas.thr.set(0.4)
        #gui.current_section_canvas.thr_ca.set(0.4)
        #gui.topological_build()
        #gui.analyze_fc(False)
        gui.analyze_LBI(superf=0)

        lp0=np.argmax(gui.lbi)
        lp1= lp0 + lbi_depth
        '''

        if np.sum(gui.annotation)==0:
            data.cell(lp, keys['Voulnerability']).value = 0
        else:
            data.cell(lp, keys['Voulnerability']).value = 1

        '''
        data.cell(lp, keys['Max LBI']).value = np.max(gui.lbi)

        f=data.cell(lp, keys['Fr OCT']).value -1
        data.cell(lp, keys['OCT software MLA']).value = gui.areas[f] * gui.calibration_scale * gui.calibration_scale / ((818.5 / 160) * (818.5 / 160))
        data.cell(lp, keys['OCT software Lipid arc']).value = int(gui.plaque_angle[f])
        data.cell(lp, keys['OCT software FTC']).value = int(gui.fct[f][0] * gui.calibration_scale)

        data.cell(lp, keys['Fr LBI']).value = lp0
        data.cell(lp, keys['LBI depth']).value = lbi_depth
        #data.cell(lp, keys['Calibration']).value = gui.calibration_scale

        data.cell(lp, keys['mLA']).value = np.min(np.array(gui.areas)[lp0:lp1]) * gui.calibration_scale * gui.calibration_scale / ((818.5 / 160) * (818.5 / 160))
        data.cell(lp, keys['Fr mLA']).value = np.argmin(np.array(gui.areas)[lp0:lp1]) + lp0 + 1

        data.cell(lp, keys['LPA']).value = np.max(np.array(gui.plaque_angle)[lp0:lp1])
        data.cell(lp, keys['Fr LPA']).value = np.argmax(np.array(gui.plaque_angle)[lp0:lp1]) + lp0 + 1

        fct = np.array([i[0] for i in gui.fct]) * gui.calibration_scale
        fct[fct == 0] = 1000
        # print(fct)
        data.cell(lp, keys['FCT']).value = np.min(fct[lp0:lp1])
        data.cell(lp, keys['Fr FCT']).value = np.argmin(fct[lp0:lp1]) + lp0 + 1
        '''

        wb.save(os.path.dirname(os.path.realpath('__file__')) + '/OCTvsNIRS 41 pazienti.xlsx')
