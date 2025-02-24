import os
import numpy as np
import GuiClean as GC
from openpyxl import Workbook, load_workbook
import unicodedata
import sys
import shutil

# env_path=" E:\Pazienti CLIMA "#"/home/eugenio/workspace/documets di pasqualino/Documents/OCTdatabase/all_pullbacks/nuovi dati"
# model_file ="/home/eugenio/workspace/documets di pasqualino/Documents"#placcaultras"#4.7-1"#/4.4-12"#4.3/4.3-1.1"#modello2-4e-r0-all08"#modello2-mozzino-guida"#modello2e4r-5-pure"#modello2-4epoche-reg000001"#placcaultras"#modello2"#



wb = load_workbook(os.path.dirname(os.path.realpath('__file__')) + "/Clima_Results.xlsx")

for sheet in wb:
    print(sheet.title)

data = wb['selected mace lbi']
result = wb['mace lbi per patient']
keys = {}
gui = GC.Main()

i = 1
for cell in data[1]:
    keys[cell.value] = i
    print(i)
    i += 1

print(keys)
l, ls, lt, lst = [],[],[],[]
old_name=data.cell(2, keys['Patient surname']).value
lp=2
lpr=2
while lp< 82:
    name = data.cell(lp, keys['Patient surname']).value
    if name==old_name:
        l.append(data.cell(lp, keys['Max LBI']).value)
        print(data.cell(lp, keys['Max LBI']).value)
        ls.append(data.cell(lp, keys['Max LBIs']).value)
        lt.append(data.cell(lp, keys['Max LBI Thr']).value)
        lst.append(data.cell(lp, keys['Max LBIs Thr']).value)
        lp+=1
    else:
        for j in range(1, 3):
            result.cell(row=lpr, column=j).value = data.cell(row=lp-1, column=j).value
        for j in range(1, 50):
            print(data.cell(row=lp-1, column=j).value)
        print(l)
        result.cell(lpr, keys['Max LBI']).value = min(l)
        result.cell(lpr, keys['Max LBIs']).value = min(ls)
        result.cell(lpr, keys['Max LBI Thr']).value = min(lt)
        result.cell(lpr, keys['Max LBIs Thr']).value = min(lst)
        lpr+=1
        old_name=name
        l, ls, lt, lst = [],[],[],[]


wb.save(os.path.dirname(os.path.realpath('__file__')) + "/Clima_Results.xlsx")