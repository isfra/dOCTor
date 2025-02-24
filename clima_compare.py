import os
import numpy as np
import GuiClean as GC
from openpyxl import Workbook, load_workbook
import unicodedata
import sys
import shutil

#env_path=" E:\Pazienti CLIMA "#"/home/eugenio/workspace/documets di pasqualino/Documents/OCTdatabase/all_pullbacks/nuovi dati"
#model_file ="/home/eugenio/workspace/documets di pasqualino/Documents"#placcaultras"#4.7-1"#/4.4-12"#4.3/4.3-1.1"#modello2-4e-r0-all08"#modello2-mozzino-guida"#modello2e4r-5-pure"#modello2-4epoche-reg000001"#placcaultras"#modello2"#


wb=load_workbook(os.path.dirname(os.path.realpath('__file__')) + "/Clima_Results.xlsx")

for sheet in wb:
    print(sheet.title)

data=wb['MaceData']
result=wb['Result mace']
keys={}
gui=GC.Main()


i=1
for cell in  result[1]:
    keys[cell.value]=i
    print(i)
    i+=1
    
print(keys)

database='E:\Pazienti CLIMA'#'/media/eugenio/Hitachi/Pazienti CLIMA'#'/media/fisici/Hitachi/Pazienti CLIMA'   #'/home/eugenio/workspace/documets di pasqualino/Documents/OCTdatabase/pazienti mace'
counter=0
lpr=1
not_checked=0
not_checked_list=[]
not_found_list=[]
old_data_path=''
err=0

good=0
bad=0
warn_c=np.zeros(1526)
warn=np.zeros(1526)

for lp in range(2,102):
    '''
    if  isinstance(result.cell(lp,keys['mFCT Comp']).value,int):
        if result.cell(lp,keys['mFCT Comp']).value>1000:
            result.cell(lp, keys['mFCT Comp']).value='NF'

    if  isinstance(result.cell(lp,keys['mFCT Thr']).value,int):
        if result.cell(lp,keys['mFCT Thr']).value>1000:
            result.cell(lp, keys['mFCT Thr']).value='NF'


    if isinstance(result.cell(lp, keys['MLPA Comp']).value, int) and isinstance(result.cell(lp, keys['Max LP Arc']).value, int) and isinstance(result.cell(lp, keys['mFCT Comp']).value, int) and isinstance(result.cell(lp, keys['FCT']).value, int) and isinstance(result.cell(lp, keys['mLA Comp']).value, float) and isinstance(result.cell(lp, keys['mLA']).value, float):
        if result.cell(lp, keys['MLPA Comp']).value>180 and result.cell(lp, keys['mFCT Comp']).value<75 and result.cell(lp, keys['mLA Comp']).value<3.5 :
            result.cell(lp, keys['criteria comp']).value = 1
        else:
            result.cell(lp, keys['criteria comp']).value = 0

        if  result.cell(lp, keys['Max LP Arc']).value>180 and result.cell(lp, keys['FCT']).value<75 and result.cell(lp, keys['mLA']).value<3.5 :
            result.cell(lp, keys['criteria']).value = 1
        else:
            result.cell(lp, keys['criteria']).value = 0

        if isinstance(result.cell(lp, keys['MLPA Thr']).value, int) and isinstance(result.cell(lp, keys['mFCT Thr']).value,int):
            if result.cell(lp, keys['MLPA Thr']).value > 180 and result.cell(lp,keys['mFCT Thr']).value < 75 and result.cell(lp, keys['mLA Comp']).value < 3.5:
                result.cell(lp, keys['criteria thr']).value=1
            else:
                result.cell(lp, keys['criteria thr']).value = 0

    if isinstance(result.cell(lp, keys['MLPA Comp']).value, int) and isinstance(result.cell(lp, keys['Max LP Arc']).value, int):
        err0 = np.abs(result.cell(lp, keys['MLPA Comp']).value - result.cell(lp, keys['Max LP Arc']).value) / result.cell(lp,keys['Max LP Arc']).value
        result.cell(lp, keys['MLPA error']).value=err0


    if isinstance(result.cell(lp, keys['MLPA Thr']).value, int) and isinstance(result.cell(lp, keys['Max LP Arc']).value, int):
        err0 = np.abs(result.cell(lp, keys['MLPA Thr']).value - result.cell(lp, keys['Max LP Arc']).value) / result.cell(lp,keys['Max LP Arc']).value
        result.cell(lp, keys['MLPA thr error']).value=err0

    if isinstance(result.cell(lp, keys['mFCT Comp']).value, int) and isinstance(result.cell(lp, keys['FCT']).value, int):
        err0 = np.abs(result.cell(lp, keys['mFCT Comp']).value - result.cell(lp, keys['FCT']).value) / result.cell(lp,keys['FCT']).value
        result.cell(lp, keys['FCT error']).value=err0

    if isinstance(result.cell(lp, keys['mFCT Thr']).value, int) and isinstance(result.cell(lp, keys['FCT']).value, int):
        err0 = np.abs(result.cell(lp, keys['mFCT Thr']).value - result.cell(lp, keys['FCT']).value) / result.cell(lp,keys['FCT']).value
        result.cell(lp, keys['FCT thr error']).value=err0

    if isinstance(result.cell(lp, keys['mLA Comp']).value, float) and isinstance(result.cell(lp, keys['mLA']).value, float):
        err0 = np.abs(result.cell(lp, keys['mLA Comp']).value - result.cell(lp, keys['mLA']).value) / result.cell(lp,keys['mLA']).value
        result.cell(lp, keys['mLA error']).value=err0
    '''

    '''if  isinstance(result.cell(lp,keys['MLPA Comp']).value,int) and isinstance(result.cell(lp,keys['Max LP Arc']).value,int) :
        #if result.cell(lp,keys['mFCT Comp']).value>1000:
        #    result.cell(lp, keys['mFCT Comp']).value='NF'
        err0=np.abs(result.cell(lp, keys['MLPA Comp']).value - result.cell(lp, keys['Max LP Arc']).value) / result.cell(lp, keys['Max LP Arc']).value

        err+=err0
        if err0<0.2:
            good+=1
        else:
            bad+=1
        n+=1
    '''




    found=False
    c=False
    
    #if str(data.cell(lp, keys['Position']).value)=='P1' or str(data.cell(lp, keys['Position']).value)=='D1':
    if (not str(data.cell(lp, keys['LP Frames']).value)=='None') and (not data.cell(lp, keys['LP Frames']).value==None) and (not str(data.cell(lp, keys['LP Frames']).value)=='NA')   :
        #print(data.cell(lp, keys['Patient surname']).value, data.cell(lp, keys['Patient name']).value, str(data.cell(lp, keys['Center']).value), str(data.cell(lp, keys['Acquisition date']).value),'PB'+str(data.cell(lp, keys['Pullback']).value ))   
        patient_surname=data.cell(lp, keys['Patient surname']).value.strip().lower()
        patient_surname=patient_surname.replace("_"," ")
        patient_surname = unicodedata.normalize('NFD', patient_surname)\
            .encode('ascii', 'ignore')\
            .decode("utf-8")
        
        patient_name=data.cell(lp, keys['Patient name']).value.strip().lower()
        patient_name = unicodedata.normalize('NFD', patient_name)\
            .encode('ascii', 'ignore')\
            .decode("utf-8")
        
        if len(patient_surname)>3:
            if patient_surname[0:3]=='fhu':
                patient_surname+=' '
        center=str(data.cell(lp, keys['Center']).value)
        center_path=os.path.join(database,center)
        
        if os.path.isdir(center_path):
            c=True
            patients = next(os.walk(center_path))[1]
            #print('Enter in Center dir',center_path)
        
            for patient in patients:
                m_patient=patient.strip().lower()
                m_patient = unicodedata.normalize('NFD', m_patient)\
                    .encode('ascii', 'ignore')\
                    .decode("utf-8")
                
                if len(m_patient)>3 :
                    if m_patient[0:3]=='fhu'  :
                        m_patient+=' '
                        
                
                #print('test center names  :   ' ,  m_patient,   patient_surname,  (patient_surname in m_patient),    patient_name,   (patient_name in m_patient), ((patient_surname in m_patient) and len(patient_name)<2) )
                if ((patient_surname in m_patient) and (patient_name in m_patient)) or ((patient_surname in m_patient) and len(patient_name)<2) :
                    data_path=os.path.join(center_path ,patient, 'PB'+str(data.cell(lp, keys['Pullback']).value ) )
                    found=True
        
        date=str(data.cell(lp, keys['Acquisition date']).value).split(' ')
        date=date[0].split('-')
        #print(date,str(date),type(date) , date[0])
        database_date= os.path.join(database,date[0])
        
        if found==False:
            #print('Searching by date',database_date,patient_name,patient_surname)
            patients = next(os.walk(database_date))[1]
            #print(data.cell(lp, keys['Patient surname']).value , data.cell(lp, keys['Patient name']).value )
            for patient in patients:
                m_patient=patient.strip().lower()
                m_patient = unicodedata.normalize('NFD', m_patient)\
                    .encode('ascii', 'ignore')\
                    .decode("utf-8")
                
                #print('test date names  :   ' ,  m_patient,   patient_surname,  (patient_surname in m_patient),    patient_name,   (patient_name in m_patient), ((patient_surname in m_patient) and len(patient_name)<2) )
                if ((patient_surname in m_patient) and (patient_name in m_patient)) or ((patient_surname in m_patient) and len(patient_name)<2):
                    data_path=os.path.join(database_date ,patient, 'PB'+str(data.cell(lp, keys['Pullback']).value ) )  
                    found=True
        
        already_checked=False
        if os.path.exists(data_path+'/GuideArtifact.npy'):
            guide=np.load(data_path+'/GuideArtifact.npy')
            if not sum(sum(guide))==0:
                already_checked=True    
        if not  os.path.exists(data_path+'/Prediction.npy'):
            already_checked=False

        
        if (os.path.isdir(data_path) ) and (not data.cell(lp, keys['Pullback']).value==None) and already_checked :
            #if ( (data.cell(lp, keys['Pullback']).value==1  and data.cell(lp+3, keys['Pullback']).value==2) or (data.cell(lp, keys['Pullback']).value==2 and data.cell(lp+3, keys['Pullback']).value==1) ) and ( os.path.dirname(os.path.dirname(os.path.abspath(data_path))) in ['/media/eugenio/Hitachi/Pazienti CLIMA/2012']):#,'/media/eugenio/Hitachi/Pazienti CLIMA/2010','/media/eugenio/Hitachi/Pazienti CLIMA/2011']):
            print(data_path)#, next(os.walk(os.path.dirname(os.path.abspath(data_path)) ))[1] )
            patient_path=os.path.dirname(os.path.abspath(data_path))
            #print(lp,patient_surname)

            lpr += 1

            for j in range(1,21):
                result.cell(row = lpr, column = j).value = data.cell(row = lp, column = j).value

            
            try:
                new_pb=False
                if not data_path==old_data_path:
                    if os.path.exists(old_data_path + '/pullback.npy'):
                        os.remove(old_data_path + '/pullback.npy')
                    gui.import_pullback_callback(data_path)
                    new_pb=True

                old_data_path=data_path

                if not os.path.exists(gui.pullback_dir + '/Calibration.npy'):
                    gui.calibrate()
                    np.save(gui.pullback_dir + '/Calibration.npy',gui.calibration)

                LP=data.cell(lp, keys['LP Frames']).value.split('-')

                if len(LP)>1:
                    lp0=int(LP[0])-1
                    lp1=int(LP[1])-1

                if gui.Nf < 500:
                    lbi_depth = 20
                else:
                    lbi_depth = 40


                if new_pb and lp1-lp0>lbi_depth:
                    gui.current_section_canvas.thr.set(0.4)
                    gui.current_section_canvas.thr_ca.set(0.3)
                    gui.topological_build()

                    gui.analyze_LBI(superf=0)
                    gui.mark_deep_plaque_n(depth=120)


                if len(LP) > 1:
                    if lp1 - lp0 > lbi_depth + 1:
                        #print(lp0,lp1,lbi_depth,len(gui.lbi))
                        result.cell(lpr, keys['Max LBI']).value = np.max(gui.lbi[lp0 :lp1 - lbi_depth])

                if not (data.cell(lp,keys['mLA Fr']).value == 'NA' or  data.cell(lp,keys['mLA Fr']).value == 'None' ):
                    result.cell(lpr,keys['mLA Comp']).value = gui.areas[ data.cell(lp,keys['mLA Fr']).value - 1] * gui.calibration_scale* gui.calibration_scale/ ((818.5/160)*(818.5/160))


                if not (data.cell(lp,keys['Max LP Arc Fr']).value == 'NA' or  data.cell(lp,keys['Max LP Arc Fr']).value == 'None'):
                    result.cell(lpr, keys['MLPA Comp']).value = int(gui.plaque_angle[data.cell(lp, keys['Max LP Arc Fr']).value - 1])
                    result.cell(lpr,keys['mFCT Comp']).value = int(gui.fct[ data.cell(lp,keys['Max LP Arc Fr']).value -1][0] * gui.calibration_scale)


                if new_pb and lp1-lp0>lbi_depth:
                    gui.analyze_LBI(superf=1)
                if len(LP) > 1:
                    if lp1 - lp0 > lbi_depth + 1:
                        #print(lp0,lp1,lbi_depth,len(gui.lbi))
                        result.cell(lpr, keys['Max LBIs']).value = np.max(gui.lbis[lp0 :lp1 - lbi_depth])


                if new_pb and lp1 - lp0 > lbi_depth:
                    gui.current_section_canvas.thr.set(0.65)
                    gui.current_section_canvas.thr_ca.set(0.3)
                    gui.topological_build()

                    gui.analyze_LBI(superf=0)
                    gui.mark_deep_plaque_n(depth=120)

                all_fctm = np.ma.array(gui.all_fct, mask=[gui.lipid == 0])
                fct = all_fctm.min(axis=1)

                if len(LP) > 1:
                    if lp1 - lp0 > lbi_depth + 1:
                        # print(lp0,lp1,lbi_depth,len(gui.lbi))
                        result.cell(lpr, keys['Max LBI Thr']).value = np.max(gui.lbi[lp0:lp1 - lbi_depth])


                if not (data.cell(lp, keys['Max LP Arc Fr']).value == 'NA' or data.cell(lp,keys['Max LP Arc Fr']).value == 'None'):
                    result.cell(lpr, keys['MLPA Thr']).value = int(gui.plaque_angle[data.cell(lp, keys['Max LP Arc Fr']).value - 1])
                    #result.cell(lpr,keys['mFCT Comp']).value = int(gui.fct[ data.cell(lp,keys['Max LP Arc Fr']).value -1][0] * gui.calibration_scale)

                    if str(fct[data.cell(lp, keys['Max LP Arc Fr']).value - 1]).replace('.', '').isnumeric():
                        result.cell(lpr, keys['mFCT Thr']).value = int(fct[data.cell(lp, keys['Max LP Arc Fr']).value - 1] * gui.calibration_scale)
                    else:
                        result.cell(lpr, keys['mFCT Thr']).value = 'NA'

                if new_pb and lp1 - lp0 > lbi_depth:
                    gui.analyze_LBI(superf=1)
                if len(LP) > 1:
                    if lp1 - lp0 > lbi_depth + 1:
                        # print(lp0,lp1,lbi_depth,len(gui.lbi))
                        result.cell(lpr, keys['Max LBIs Thr']).value = np.max(gui.lbis[lp0:lp1 - lbi_depth])


                with open(os.path.dirname(os.path.realpath('__file__')) + "/completed_lbi_mace.txt","a") as completed_file :
                    completed_file.write( data_path + '    ' + str(data.cell(lp, keys['Position']).value) )
                    completed_file.write("\n")


            except:
        
                #counter+=1
                with open(os.path.dirname(os.path.realpath('__file__')) + "/errors_result_mace.txt","a") as errors_file:
                    errors_file.write(data_path + '    ' + str(data.cell(lp, keys['Position']).value))
                    errors_file.write("\n")
                print('ERROR IN ANALYSIS!!!!___________' ,data_path, str(data.cell(lp, keys['Position']).value))
                pass

        '''
            if (not os.path.isdir(data_path) ) and (not data.cell(lp, keys['Pullback']).value==None) and (str(data.cell(lp, keys['Position']).value)=='P1' or str(data.cell(lp, keys['Position']).value)=='D1'):
                counter+=1
                not_found_list.append(data_path)
            if (not already_checked) and (str(data.cell(lp, keys['Position']).value)=='P1' or str(data.cell(lp, keys['Position']).value)=='D1'):
                not_checked+=1
                not_checked_list.append(data_path)
                
            if len(next(os.walk(patient_path ))[1])==2 :
                gui.import_pullback_callback(  os.path.join(patient_path,next(os.walk(patient_path ))[1][0])  )                
                gui.mainloop()
                gui.import_pullback_callback(  os.path.join(patient_path,next(os.walk(patient_path ))[1][1])  )
                gui.mainloop()
            if len(next(os.walk(patient_path ))[1])==0:
                data_path=os.path.join(patient_path,'PB1')
                os.mkdir(data_path)
                for file in os.listdir(patient_path):
                    if not (file == 'PB1') :
                        shutil.move(os.path.join(patient_path,file) , os.path.join(data_path,file))
                
            if len(next(os.walk(patient_path ))[1])==1:
                
                print(data.cell(lp, keys['Pullback']).value, next(os.walk(patient_path ))[1][0], 'PB' + str(data.cell(lp, keys['Pullback']).value))    
                os.rename(os.path.join(patient_path,next(os.walk(patient_path ))[1][0]), os.path.join(patient_path,'PB' + str(data.cell(lp, keys['Pullback']).value) ) )

            for dir in next(os.walk(patient_path ))[1]:
                if dir[:2]=='PB' and dir[2]==' ':
                    
                    print(dir, dir.replace(" ",""))
                    os.rename(os.path.join(patient_path,dir), os.path.join(patient_path,dir.replace(" ","")) )
                    #sys.exit("rinominato")
        '''
         
    '''
        if not found:
            print(data.cell(lp, keys['Patient surname']).value, data.cell(lp, keys['Patient name']).value, str(data.cell(lp, keys['Center']).value), str(data.cell(lp, keys['Acquisition date']).value),'PB'+str(data.cell(lp, keys['Pullback']).value ))   
            counter+=1
            print('NOT FOUND',c)
            print('Enter in Center dir',center_path)
            if c:
                patients = next(os.walk(center_path))[1]
                for patient in patients:
                    m_patient=patient.strip().lower()
                    
                    print('test center names  :   ' ,  m_patient,   patient_surname,  (patient_surname in m_patient),    patient_name,   (patient_name in m_patient),
                          ((patient_surname in m_patient) and len(patient_name)<2) , ((patient_surname in m_patient) and (patient_name in m_patient)) or ((patient_surname in m_patient) and len(patient_name)<2))
            print('Searching by date',database_date)
            patients = next(os.walk(database_date))[1]
            for patient in patients:
                m_patient=patient.strip().lower()
                print('test date names  :   ' ,  m_patient,   patient_surname,  (patient_surname in m_patient),    patient_name,   (patient_name in m_patient),
                      ((patient_surname in m_patient) and len(patient_name)<2) , ((patient_surname in m_patient) and (patient_name in m_patient)) or ((patient_surname in m_patient) and len(patient_name)<2))
            '''
        #else:
        #    print('FOUND',data_path)
    '''
    gui.import_pullback_callback(data_path)
    #print(str(data.cell(lp, keys['Pullback'])))
    LP=data.cell(lp, keys['LP Frames']).value.split('-')
    print(LP)
    lbi_depth=20
    if len(LP)>1:
        
        lp0=int(LP[0])-1
        lp1=int(LP[1])-1
        
        if not (data.cell(lp,keys['mLA Fr']).value == 'NA' or  data.cell(lp,keys['mLA Fr']).value == 'None' ):
            data.cell(lp,keys['mLA Comp']).value = gui.areas[ data.cell(lp,keys['mLA Fr']).value ]

        if not (data.cell(lp,keys['Max LP Arc Fr']).value == 'NA' or  data.cell(lp,keys['mLA Fr']).value == 'None'):
            data.cell(lp,keys['MLPA Comp']).value = gui.plaque_angle[ data.cell(lp,keys['Max LP Arc Fr']).value ]
            data.cell(lp,keys['mFCT Comp']).value = gui.fct[ data.cell(lp,keys['Max LP Arc Fr']).value ][0]* 0.88 / 196 * 1000
        
        
        gui.analyze_LBI(superf=0)
        #print(np.max(gui.lbi[lp0:lp1]))
        #print(np.argmax(gui.lbi[lp0:lp1])+ lp0)
        #print(data.cell(lp,keys['Max LBI']).value)
        data.cell(lp,keys['Max LBI']).value = np.max(gui.lbi[lp0:lp1])
        data.cell(lp,keys['Max LBI Fr']).value = np.argmax(gui.lbi[lp0:lp1])+ lp0 +1
        data.cell(lp,keys['Mean LBI']).value = np.mean(gui.lbi[lp0:lp1-lbi_depth])
        data.cell(lp,keys['LBI all']).value = gui.LBI(lp0,lp1)
        
        data.cell(lp,keys['mLA GC']).value =  np.min( np.array(gui.areas) [lp0:lp1] )
        data.cell(lp,keys['mLA GC Fr']).value =  np.argmin( np.array(gui.areas) [lp0:lp1] ) + lp0 +1
        
        data.cell(lp,keys['MLPA GC']).value =  np.max( np.array(gui.plaque_angle) [lp0:lp1] )
        data.cell(lp,keys['MLPA GC Fr']).value =  np.argmax( np.array(gui.plaque_angle) [lp0:lp1] ) + lp0 +1
        
        fct=np.array([i[0] for i in gui.fct])* 0.88 / 196 * 1000
        fct[fct==0]=1000
        #print(fct)
        data.cell(lp,keys['mFCT GC']).value =  np.min( fct[lp0:lp1] )
        data.cell(lp,keys['mFCT GC Fr']).value =  np.argmin( fct[lp0:lp1] ) + lp0 + 1
        
        gui.analyze_LBI(superf=1)
        data.cell(lp,keys['Max LBIs']).value = np.max(gui.lbi[lp0:lp1])   
        data.cell(lp,keys['Max LBIs Fr']).value = np.argmax(gui.lbi[lp0:lp1])+ lp0 +1
        data.cell(lp,keys['Mean LBIs']).value = np.mean(gui.lbi[lp0:lp1-lbi_depth])
        data.cell(lp,keys['LBIs all']).value = gui.LBI_superf(lp0,lp1)
        
    gui.analyze_LBI(superf=0)
    data.cell(lp,keys['Max LBI GC']).value = np.max(gui.lbi)
    data.cell(lp,keys['Max LBI GC Fr']).value = np.argmax(gui.lbi) +1
    
    gui.analyze_LBI(superf=1)
    data.cell(lp,keys['Max LBIs GC']).value = np.max(gui.lbi)
    data.cell(lp,keys['Max LBIs GC Fr']).value = np.argmax(gui.lbi) +1
    
        
        
    '''
#err/=n
#print(err,good,bad)

wb.save(os.path.dirname(os.path.realpath('__file__')) + "/Clima_Results.xlsx")
#print(sum(warn),sum(warn_c),sum(np.abs(np.subtract(warn,warn_c))))
#print()
print('pullbacks non anlizzati',not_checked)
for p in not_checked_list:
    print(p)
print('pullbacks non trovati',counter)
for p in not_found_list:
    print(p)

        