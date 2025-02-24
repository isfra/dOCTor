import os
import numpy as np
import GuiClean as GC
from openpyxl import Workbook, load_workbook
import shutil

env_path="/home/eugenio/workspace/documets di pasqualino/Documents/OCTdatabase/all_pullbacks/nuovi dati"
model_file ="/home/eugenio/workspace/documets di pasqualino/Documents"#placcaultras"#4.7-1"#/4.4-12"#4.3/4.3-1.1"#modello2-4e-r0-all08"#modello2-mozzino-guida"#modello2e4r-5-pure"#modello2-4epoche-reg000001"#placcaultras"#modello2"#


wb=load_workbook(os.path.dirname(os.path.realpath('__file__')) + '/NonMaceLBI.xlsx')#"/globalLBI.xlsx")

for sheet in wb:
    print(sheet.title)

data=wb['Sheet1']
keys={}
gui=GC.Main()

i=1
for cell in  data[1]:
    keys[cell.value]=i
    print(i)
    i+=1
    
print(keys)

database='/media/fisici/Hitachi/Pazienti CLIMA' #'D:\Franz\oct_software\data\Pazienti CLIMA x matematici\San Giovanni 2015'    #'/home/eugenio/workspace/documets di pasqualino/Documents/OCTdatabase/pazienti mace'
#database='/media/fisici/Hitachi/Pazienti clima fra'

lp = 110
thr = 500

chem_path=database +'/Chemiograms'

if not os.path.exists(chem_path):
    os.mkdir(chem_path)
    
years= next(os.walk(database))[1]

for year in years:

    patients = next(os.walk(os.path.join(database, year)))[1]
    
    for patient in patients :
        #print(next(os.walk(os.path.join(database, patient)))[1])
        '''if  not next(os.walk(os.path.join(database,year,patient)))[1]:
            data_path=os.path.join(database,year,patient,'PB1')
            os.mkdir(data_path)
            #print(os.listdir(os.path.join(database, patient)))
            for file in os.listdir(os.path.join(database,year,patient)):
                #print(file)
                if not (file == 'PB1') :
                    shutil.move(os.path.join(database,year,patient,file) , os.path.join(data_path,file))
        '''
            
        for pb in next(os.walk(os.path.join(database,year,patient)))[1]:
            data_path=os.path.join(database,year,patient,pb)
            
            print(data_path)
            if os.path.exists(data_path+'/pullback.npy') and os.path.exists(data_path+'/Prediction.npy') :
                '''
                for file in os.listdir(data_path):
                    shutil.move(os.path.join(data_path,file),os.path.join(database,year,patient,file))
                if not os.listdir(data_path):
                    os.rmdir(data_path)
                 '''   
                data.cell(lp, keys['Patient surname']).value = patient
                data.cell(lp, keys['Pullback']).value = pb
                
                print(data_path)
                gui.import_pullback_callback(data_path)
                
                l=0
                for interv in gui.interv_of_pertinence:
                   l+=interv[1]-interv[0] 
                data.cell(lp, keys['Length']).value = l*0.2
                
                lbi_depth=20
        
                gui.current_section_canvas.thr.set(0.3)
                gui.current_section_canvas.thr_ca.set(0.4)
                gui.topological_build()
                gui.prediction_canvas.update(0)
                
                gui.update_idletasks()
                gui.update()
                gui.save_chemiogram(chem_path, 'Low ' + patient +' '+pb)
                
                gui.analyze_LBI(superf=0)
                
                data.cell(lp,keys['Max LBI']).value = np.max(gui.lbi)
                data.cell(lp,keys['Max LBI Fr']).value = np.argmax(gui.lbi) +1
                data.cell(lp,keys['Over LBI']).value = sum(gui.lbi>thr)
                if l-20 > 0:
                    data.cell(lp,keys['%Over LBI']).value = sum(gui.lbi>thr)/(l-20)
                else:
                    data.cell(lp,keys['%Over LBI']).value = None
                
                #data.cell(lp,keys['Mean LBI']).value = np.mean(gui.lbi[lp0:lp1-lbi_depth])
        
                gui.analyze_LBI(superf=1)
                data.cell(lp,keys['Max LBIs']).value = np.max(gui.lbi)
                data.cell(lp,keys['Max LBIs Fr']).value = np.argmax(gui.lbi) +1
                data.cell(lp,keys['Over LBIs']).value = sum(gui.lbi>thr)
                if l-20 > 0:
                    data.cell(lp,keys['%Over LBIs']).value = sum(gui.lbi>thr)/(l-20)
                else:
                    data.cell(lp,keys['%Over LBIs']).value = None
                
                #data.cell(lp,keys['Mean LBIs']).value = np.mean(gui.lbi[lp0:lp1-lbi_depth])
                #data.cell(lp,keys['LBIs all']).value = gui.LBI_superf(lp0,lp1)
                
                gui.current_section_canvas.thr.set(0.65)
                gui.current_section_canvas.thr_ca.set(0.4)
                gui.topological_build()
                gui.prediction_canvas.update(0)
                
                gui.update_idletasks()
                gui.update()
                gui.save_chemiogram(chem_path, 'High ' + patient +' '+pb)
                
                gui.analyze_LBI(superf=0)
        
                data.cell(lp,keys['Max LBI thr']).value = np.max(gui.lbi)
                data.cell(lp,keys['Max LBI thr Fr']).value = np.argmax(gui.lbi) +1
                data.cell(lp,keys['Over LBI thr']).value = sum(gui.lbi>thr)
                if l-20 > 0:
                    data.cell(lp,keys['%Over LBI thr']).value = sum(gui.lbi>thr)/(l-20)
                else:
                    data.cell(lp,keys['%Over LBI thr']).value = None
                
                gui.analyze_LBI(superf=1)
                data.cell(lp,keys['Max LBIs thr']).value = np.max(gui.lbi)
                data.cell(lp,keys['Max LBIs thr Fr']).value = np.argmax(gui.lbi) +1
                data.cell(lp,keys['Over LBIs thr']).value = sum(gui.lbi>thr)
                if l-20 > 0:
                    data.cell(lp,keys['%Over LBIs thr']).value = sum(gui.lbi>thr)/(l-20)
                else:
                    data.cell(lp,keys['%Over LBIs thr']).value = None
                
                lp+=1
    
    wb.save(os.path.dirname(os.path.realpath('__file__')) + '/NonMaceLBI.xlsx')#"/globalLBI.xlsx")
            
        